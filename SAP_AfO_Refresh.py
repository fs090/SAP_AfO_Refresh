# -*- coding: utf-8 -*-
"""
Created on 3/5/2023
Author: https://github.com/fs090


Based on:
Created on 3/8/2021
Author: Arnold Souza
Email: arnoldporto@gmail.com
"""

import os

import sys
import pathlib

import time
import timeit
from datetime import date
from datetime import datetime
from dateutil.relativedelta import relativedelta


from logging.handlers import TimedRotatingFileHandler
#import functools
from functools import wraps
from cryptography.fernet import Fernet

import win32com.client as win32
from win32com.client import constants as cst



from os import listdir 

from os.path import isfile, join
from pathlib import Path

import pandas as pd
from halo import Halo

from  psutil import process_iter
from tenacity import retry, wait_fixed, before_sleep_log, stop_after_attempt, stop_after_delay

import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox

import re

import csv
import subprocess

import numpy as np

from shutil import rmtree

import tempfile

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import NamedStyle
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import glob
from winreg import ConnectRegistry, OpenKey, DeleteKey, HKEY_CURRENT_USER, HKEY_LOCAL_MACHINE #, CloseKey, QueryInfoKey, SetValueEx

#for handling open window from BOA message
from win32gui import FindWindow, PostMessage 
from win32con import WM_CLOSE

#Threading used to monitor for pop-ups from BOA addin which can not be handled
from pythoncom import CoInitialize
from threading import Thread


import logging


class TimeoutException(Exception):
    pass

class CredentialsException(Exception):
    pass


global start_mode, update_mode, clear_excel_sessions, password_found, path_tmp, Supress_All_BOA_Messages, Err_Count, use_loaded_encrypted_pw, path_user, start_time, end_time, Run_Refresh_All_Excel, Refresh_All_Waiting_Time, Close_Excel_After_Completion, Force_Close_BOA_Messages_Window, refresh_running, Restart_Excel_after_refresh_Error, default_system, default_client, curr_boa_message_id, boa_message_severity, boa_message_text, prev_msg, file_source_config, run_times, run_time_start, run_time_end, capture_runtimes, file_path, settings_sheet_available, save_results_csv, always_refresh_error_queries, default_week_mask, week_mask


start_time = datetime.fromtimestamp(time.time())


###### Settings / Arguments initialization for several functions
#If True a Macro will be implemented in all Excel Files temporarily to handle BOA messages
clear_excel_sessions = True
#Run Refresh All in Excel after refreshing each File & Waiting time after refresh
Run_Refresh_All_Excel = True
Refresh_All_Waiting_Time = 2 #60
#Close Excel after completion
Close_Excel_After_Completion = True
#Monitor for BOA message windows and close them
Force_Close_BOA_Messages_Window = True
#Save and close Files and Excel when an error occured and restart Excel
Restart_Excel_after_refresh_Error = False
#Implement temporary VBA macro in Excel files to supress messages
Supress_All_BOA_Messages = False
#Capture runtimes in CSV
capture_runtimes = False
#Save Query results in CSV
save_results_csv = False
#Re-run queries which returned an error / no results during the last refresh
always_refresh_error_queries = False


#Initializing Values
update_mode = ""
start_mode ="Initializing"
password_found = False
#True if the file source is a configuration file
file_source_config = True
#If in interface button selected use only loaded PWs
use_loaded_encrypted_pw = False
#To check if configuration file does include settings
settings_sheet_available = True

#Default working week is from monday to friday, can be changed using default_week_mask and week_mask (e.g. monday to friday is '1111100'). If another schedule is to be taken into account change default_week_mask to true and adjust week_mask
default_week_mask = False
week_mask = '1111100'


#Initial BOA message ID
curr_boa_message_id = -1
boa_message_severity = -1
boa_message_text = ""
prev_msg = 0


#Default values for system and client
default_system = 'P72'
default_client = 100

#Optionally capture runtimes

run_times = []
run_time_start = ''
run_time_end = ''



#Used to determine if exception is an error to be sent as an email after refresh
Err_Count = 0

if len(sys.argv) > 1:
    for i in range(1, len(sys.argv)):
        if sys.argv[i] == "Scheduled" or str(sys.argv[i]).lower() == "scheduled" or str(sys.argv[i]).lower() == "arg0":
            print("Scheduled")
            start_mode = "Scheduled"
            update_mode = "Refresh"
        if sys.argv[i] == "SupressBOAMessages" or str(sys.argv[i]).lower() == "supressboamessages"or str(sys.argv[i]).lower() == "arg1":
            Supress_All_BOA_Messages = True

        if sys.argv[i] == "DontForceCloseBOAMessages" or str(sys.argv[i]).lower() == "dontforcecloseboamessages"or str(sys.argv[i]).lower() == "arg2":
            Force_Close_BOA_Messages_Window = False

        if sys.argv[i] == "DontCloseExcel" or str(sys.argv[i]).lower() == "dontcloseexcel"or str(sys.argv[i]).lower() == "arg3":
            clear_excel_sessions = False
        
        if sys.argv[i] == "DontKillExcelAfterCompletion" or str(sys.argv[i]).lower() == "dontkillexcelaftercompletion"or str(sys.argv[i]).lower() == "arg4":
            Close_Excel_After_Completion = False

        if sys.argv[i] == "DontRefresh" or str(sys.argv[i]).lower() == "dontrefresh"or str(sys.argv[i]).lower() == "arg5":
            Run_Refresh_All_Excel = False

        if sys.argv[i] == "CaptureRuntimes" or str(sys.argv[i]).lower() == "captureruntimes"or str(sys.argv[i]).lower() == "arg6":
            capture_runtimes = True

        if sys.argv[i] == "SaveResultsCSV" or str(sys.argv[i]).lower() == "saveresultscsv"or str(sys.argv[i]).lower() == "arg7":
            save_results_csv = True

        if sys.argv[i] == "RefreshErrorQueries" or str(sys.argv[i]).lower() == "refresherrorqueries"or str(sys.argv[i]).lower() == "arg8":
            always_refresh_error_queries = True
        
        if sys.argv[i] == "RestartExcelOnError" or str(sys.argv[i]).lower() == "restartexcelonerror"or str(sys.argv[i]).lower() == "arg9":
            Restart_Excel_after_refresh_Error = True


    #Scheduled Argument not found, Executable
    if start_mode =="Initializing":
        print("Excecutable")
        start_mode = "Excecutable"


#Get path of executable or from running e.g. in VS Code 
if getattr(sys, 'frozen', False):
    file_path= os.path.dirname(sys.executable)
    full_path = os.path.realpath(sys.executable)
elif __file__:
    file_path= os.path.dirname(__file__)
    full_path = os.path.realpath(__file__)


path_user = os.path.expanduser(os.getenv('USERPROFILE'))
path_tmp = tempfile.gettempdir() + os.sep


def secret_decode(token, secret_key):
            """Decrypt the secret string"""
            key = secret_key 
            f = Fernet(key)
            value = f.decrypt(bytes(token,'utf-8')).decode('utf-8')
            return value

def read_secret_file(secret_path):
    
            #Check if password file exists
            file_exists = os.path.isfile(secret_path)

            if file_exists == True:
                file = open(secret_path, 'r')
                for line in file.readlines():
                    return line
                    break
            else:
                print('No secret file file')
                return False

def read_Password_Files(path_to_file, secret_key_path):
    global df_pws
    #Reading PW Files

    if not secret_key_path == None:
        
        sec = read_secret_file(secret_key_path)
        if sec != False:
            if sec[0] =='b':
                sec = sec[1:]
    
    #Check if password file exists
    file_exists = os.path.isfile(path_to_file)
 
    if file_exists == True:
        #PW File found
        file = open(path_to_file, 'r')
    else:
        print('No password or enocded password file found')
        #Raise warning as no password file is found and user/password are not stored in the configuration file
        try:
            if user_config == False and password_config == False:
                #logger.warning("Secret file not found in " + str(secret_key_path)+ " - Make sure Secret.txt and PW_Encoded.txt are available and correct or use unenconed file Passwords.txt or consider saving user and password in the configuration file")
                print("Secret file not found in " + str(secret_key_path)+ " - Make sure Secret.txt and PW_Encoded.txt are available and correct or use unenconed file Passwords.txt or consider saving user and password in the configuration file")
        except:
            pass
        df_pws = pd.DataFrame([{0, 0, 0}])
        return False

    

    if file_exists == True:
        #pws = dict()
        pws = []
        
        for line in file.readlines():
                
                fname = re.split(r'[,;\n]',line.rstrip()) #using rstrip to remove the NewLine \n
                #Get credentials)
                #If length = 1 then file is encoded or line is empty string which will be ignored
                if len(fname) == 1 and '' not in fname:
                    if sec == False:
                        if user_config == False and password_config == False:
                            #logger.warning("Secret file not found in " + str(secret_key_path)+ " - Make sure Secret.txt and PW_Encoded.txt are available and correct or use unenconed file Passwords.txt or consider saving user and password in the configuration file")
                            print(("Secret file not found in " + str(secret_key_path)+ " - Make sure Secret.txt and PW_Encoded.txt are available and correct or use unenconed file Passwords.txt or consider saving user and password in the configuration file"))
                        df_pws = pd.DataFrame([{0, 0, 0}])
                        return False
                    else:
                        
                        line = secret_decode(line, sec)
                        fname = re.split(r'[,;\n]',line.rstrip()) #using rstrip to remove the \n
                        
                        if len(fname) == 3:
                            #globals()[f"{fname[0]}_user"] = fname[1]
                            #globals()[f"{fname[0]}_password"] = fname[2]

                            pws.append(
                            {
                                'System': fname[0],
                                'User': fname[1],
                                'Password': fname[2]
                            }
                            )
                        else:
                            #logger.warning("Could not read encoded password file, make sure to follow the syntax System;User;Password and use a separate line for each sytem")
                            print("Could not read encoded password file, make sure to follow the syntax System;User;Password and use a separate line for each sytem")
                            df_pws = pd.DataFrame([{0, 0, 0}])
                            return False
                elif len(fname) == 3:
                        pws.append(
                        {
                            'System': fname[0],
                            'User': fname[1],
                            'Password': fname[2]

                        }
                        )
                elif len(fname) == 1 and '' in fname:
                    #Empty string will be skipped
                    pass
                else:
                    #logger.warning("Could not read password file, make sure to follow the syntax System;User;Password and use a separate line for each sytem")
                    print("Could not read password file, make sure to follow the syntax System;User;Password and use a separate line for each sytem")
                    df_pws = pd.DataFrame([{0, 0, 0}])
                    return False
        
        df_pws = pd.DataFrame.from_dict(pws)           
        
    return df_pws 



def CheckFiles(source_folder, string_to_match, file_extension):
    # Check all files in source_folder
    for filename in os.listdir(source_folder):
        #Skip folders
        if os.path.isdir(os.path.join(source_folder, filename)) == False:
            #Only take into account files with matching extension
            if file_extension != None:
                if file_extension in filename.lower():
                    # Move the file if the filename contains the string to match
                    if string_to_match.lower() in filename.lower():
                        pth = os.path.join(source_folder, filename)#
                        pth = os.path.realpath(pth)
                        return pth
            else:
                    if string_to_match.lower() in filename.lower():
                        pth = os.path.join(source_folder, filename)#
                        pth = os.path.realpath(pth)
                        return pth
    return 0


def find_files(string_to_match, file_extension):
    #Search config file first in parent directory in case pyinstaller created one directory
    if string_to_match == 'Config':
        result_path =  CheckFiles(file_path, string_to_match, file_extension)
        if result_path == 0:
            result_path =  CheckFiles(Path(file_path).parent.absolute(), string_to_match, file_extension)
            if result_path == 0:
                result_path =  CheckFiles(path_user, string_to_match, file_extension)
                if result_path == 0:
                        result_path = 0

    else:
        result_path =  CheckFiles(file_path, string_to_match, file_extension)
        if result_path == 0:
            result_path =  CheckFiles(path_user, string_to_match, file_extension)
            if result_path == 0:
                result_path =  CheckFiles(Path(file_path).parent.absolute(), string_to_match, file_extension)
                if result_path == 0:
                        result_path = 0
    
    return result_path
    

def get_inital_file_paths():
    global inital_config_path, inital_password_path,inital_holiday_path, inital_secret_path,inital_encoded_path, df_passw #, df_encoded_pws
    #CheckFiles returns 0 if not found
    #Passwords
    try:
        inital_password_path =  find_files('Password','.txt')
        if inital_password_path == 0:
                inital_password_path = ""
                
    except:
        #Check user directory
        try:
            inital_password_path =  CheckFiles(path_user, 'Password','.txt')
        except:
            inital_password_path = ""   


    #Configuration
    try:
        inital_config_path =  find_files('Config','.xl')
        if inital_config_path == 0:
                inital_config_path = ""
    except:
        inital_config_path = ""
    
    #Holiday
    try:
        inital_holiday_path =  find_files('Holiday','.csv')
        if inital_holiday_path == 0:
                inital_holiday_path = ""
    except:
        #Check user directory
        try:
            inital_holiday_path =  CheckFiles(path_user, 'Holiday','.csv')
        except:
            inital_holiday_path = ""   

    #Secret
    try:
        inital_secret_path =  find_files('Secret','.txt')
        if inital_secret_path == 0:
                inital_secret_path = ""
    except:
        #Check user directory
        try:
            inital_secret_path =  CheckFiles(path_user, 'Secret','.txt')
        except:
            inital_secret_path = ""   

    #Encoded PWs
    try:
        inital_encoded_path =  find_files('Encoded','.txt')
        if inital_encoded_path == 0:
                inital_encoded_path = ""
    except:
        #Check user directory
        try:
            inital_encoded_path =  CheckFiles(path_user, 'Encoded','.txt')
        except:
            inital_encoded_path = "" 


    #Load PW if Secret + Encoded found
    if inital_secret_path != "" and inital_encoded_path != "":
        try:
            df_passw = read_Password_Files(inital_encoded_path,inital_secret_path)
        except:
            pass
    
    if inital_password_path != "":
        try:
            df_passw = read_Password_Files(inital_password_path,None)
        except:
            pass



global CONFIG_PATH



if start_mode == "Scheduled":
    retval = os.getcwd()
    os.chdir(file_path)
    #Look for xlsx files which contain "config"
    if len(glob.glob('*config*.xls*')) > 0:
        #Use first result
        c_file = glob.glob('*config*.xls*')[0]
        if len(glob.glob('*config*.xls*')) > 1:
            print('More than one configuration file found, first result taken')
    else:
        try:
            #Search in parent directory
            print('No configuration file found in current directory, searching in parent directory')
            os.chdir(pathlib.Path(file_path).parent.resolve())
            if len(glob.glob('*config*.xls*')) > 0:
                #Use first result
                c_file = glob.glob('*config*.xls*')[0]
                if len(glob.glob('*config*.xls*')) > 1:
                    print('More than one configuration file found, first result taken')
            else:
                print('No configuration file found')
                c_file = r'\Configuration.xlsx'
        except:
            print('No configuration file found')
            c_file = r'\Configuration.xlsx' 
            os.chdir(retval)
    #reset chdir
    os.chdir(retval)
    CONFIG_PATH = os.path.join(file_path,c_file)
else:
    
    
    class BOA_Interface:

      
        def close_quit(self):
            root.quit()
            sys.exit()

        
        def __init__(self,root, title, width, height):
            global update_mode
            #Basic settings
            self.root = root
            self.root.title(title)
            self.root.geometry(str(width) + 'x' + str(height)) 
            #Initalizing update mode - if window is closed code stops
            update_mode = 'No Input'

            


            #Add notebook for tabs
            main_notebook = ttk.Notebook(root)
            main_notebook.pack(pady=15)
        
            #main_notebook.grid(row=0,column=0,padx=5, pady = 5)

            #Create frames for all tabs and add them to grid
            frm_refresh = Frame(main_notebook,width=width,height=height)
            frm_configuration = Frame(main_notebook,width=width,height=height*1.5)
            frm_settings = Frame(main_notebook,width=width,height=500)

            #Add frames to notebook
            main_notebook.add(frm_refresh,text="Refresh existing configuration")
            main_notebook.add(frm_configuration,text="Create configuration file")
            main_notebook.add(frm_settings,text="Settings")




            #Control buttons
            #Add controls for refresh page

            #Labels

            self.lbl_config_path =Label(frm_refresh,text ="Path Configuration File").grid(row=0,column=0,padx=5, pady = 5)
            self.lbl_passwprd_path =Label(frm_refresh,text ="Path Password File").grid(row=1,column=0,padx=5, pady = 5)

            #Text Input fields
            self.txt_config_path = Entry(frm_refresh,width=75)
            self.txt_config_path.grid(row=0,column=1,padx=5, pady = 5)
            self.txt_config_path.insert(0,inital_config_path)

            self.txt_password_path = Entry(frm_refresh,width=75)
            self.txt_password_path.grid(row=1,column=1,padx=5, pady = 5)
            self.txt_password_path.insert(0,inital_password_path)
            

            btn_config_path = Button(frm_refresh,text="Select File",command=(lambda:self.selectpath(self.txt_config_path))).grid(row=0,column=2,padx=5, pady = 5) 

            btn_password_path = Button(frm_refresh,text="Select File",command=(lambda:self.selectpath(self.txt_password_path))).grid(row=1,column=2,padx=5, pady = 5) 
    
            btn_close = Button(frm_refresh, text="Close",command=self.close_quit).grid(row=6,column=1,padx=5, pady = 5)

            btn_start = Button(frm_refresh, text="Refresh Files",command=self.refresh).grid(row=5,column=1,padx=5, pady = 5)
        
            #Add controls for configuration page

            #Labels

            self.lbl_query_path =Label(frm_configuration,text ="Path Query File").grid(row=0,column=0,padx=5, pady = 5)
            self.lbl_dir_path =Label(frm_configuration,text ="Folder path").grid(row=1,column=0,padx=5, pady = 5)
            self.lbl_config_path =Label(frm_configuration,text ="Path Configuration File").grid(row=2,column=0,padx=5, pady = 5)
            self.lbl_passwprd_path =Label(frm_configuration,text ="Path Password File").grid(row=3,column=0,padx=5, pady = 5)

            #Text Input fields
            self.txt_query_path = Entry(frm_configuration,width=75)
            self.txt_query_path.grid(row=0,column=1,padx=5, pady = 5)
            
            self.txt_dir_path = Entry(frm_configuration,width=75)
            self.txt_dir_path.grid(row=1,column=1,padx=5, pady = 5)


            self.txt_config_path2 = Entry(frm_configuration,width=75)
            self.txt_config_path2.grid(row=2,column=1,padx=5, pady = 5)
            self.txt_config_path2.insert(0,inital_config_path)

            self.txt_password_path2 = Entry(frm_configuration,width=75)
            self.txt_password_path2.grid(row=3,column=1,padx=5, pady = 5)
            self.txt_password_path2.insert(0,inital_password_path)

            btn_query_path2 = Button(frm_configuration,text="Select File",command=(lambda:self.selectpath(self.txt_query_path))).grid(row=0,column=2,padx=5, pady = 5)  

            btn_dir_path = Button(frm_configuration,text="Select Folder",command=(lambda:self.selectdir(self.txt_dir_path))).grid(row=1,column=2,padx=5, pady = 5)  

            btn_config_path2 = Button(frm_configuration,text="Select File",command=(lambda:self.selectpath(self.txt_config_path))).grid(row=2,column=2,padx=5, pady = 5) 

            btn_password_path2 = Button(frm_configuration,text="Select File",command=(lambda:self.selectpath(self.txt_password_path))).grid(row=3,column=2,padx=5, pady = 5)   
    
            btn_close2 = Button(frm_configuration, text="Close",command=self.close_quit).grid(row=8,column=1,padx=5, pady = 5)

            btn_create_single = Button(frm_configuration, text="Create configuration file from query file",command=self.configuration_single_file).grid(row=5,column=1,padx=5, pady = 5)
            btn_create_from_dir = Button(frm_configuration, text="Create configuration file all query files in folder",command=self.configuration_multiple_files_dir).grid(row=6,column=1,padx=5, pady = 5)
            btn_create_multiple = Button(frm_configuration, text="Create configuration file from files in configuration file",command=self.configuration_multiple_files).grid(row=7,column=1,padx=5, pady = 5)


            #Add controls for settings page

            #Labels
            self.lbl_holiday_path =Label(frm_settings,text ="Path Holiday CSV File").grid(row=0,column=0,padx=5, pady = 5)
            self.lbl_holiday_path =Label(frm_settings,text ="Path Secret File").grid(row=1,column=0,padx=5, pady = 5)
            self.lbl_pass_to_encode_path =Label(frm_settings,text ="Path Passwordfile to Encode").grid(row=2,column=0,padx=5, pady = 5)
            self.lbl_encoded_path =Label(frm_settings,text ="Path Encoded Passwordfile ").grid(row=3,column=0,padx=5, pady = 5)

            #Checkboxes
            self.var_kill_excel = tk.IntVar(value=1)
            self.cb_var_kill_excel = Checkbutton(frm_settings,text='Close running Excel instances',onvalue=1, offvalue=0, variable=self.var_kill_excel, command=self.display_input)
            self.cb_var_kill_excel.grid(row=5,column=0,padx=5, pady = 5)
            #Supress all BOA Messages
            self.var_supress_boa_msg = tk.IntVar(value=0)
            self.cb_var_supress_boa_msg = Checkbutton(frm_settings,text='Supress all BOA messages',onvalue=1, offvalue=0, variable=self.var_supress_boa_msg, command=self.display_input)
            self.cb_var_supress_boa_msg.grid(row=5,column=2,padx=5, pady = 5)

            #Refresh All command
            self.var_refresh_all_excel = tk.IntVar(value=1)
            self.cb_var_refresh_all_excel = Checkbutton(frm_settings,text='Run Refresh all in Excel files after refresh',onvalue=1, offvalue=0, variable=self.var_refresh_all_excel, command=self.display_input)
            self.cb_var_refresh_all_excel.grid(row=6,column=0,padx=5, pady = 5)

            #Capture Query runtimes command
            self.var_capture_runtimes = tk.IntVar(value=0)
            self.cb_var_capture_runtimes = Checkbutton(frm_settings,text='Capture query runtimes',onvalue=1, offvalue=0, variable=self.var_capture_runtimes, command=self.display_input)
            self.cb_var_capture_runtimes.grid(row=6,column=1,padx=5, pady = 5)


            #Close Excel completely after refresh
            self.var_close_excel_after_completion = tk.IntVar(value=0)
            self.cb_var_close_excel_after_completion = Checkbutton(frm_settings,text='Close Excel after Refresh',onvalue=1, offvalue=0, variable=self.var_close_excel_after_completion, command=self.display_input)
            self.cb_var_close_excel_after_completion.grid(row=6,column=2,padx=5, pady = 5)

            #Force Close BOA Message Window
            self.var_force_close_boa_messages = tk.IntVar(value=1)
            self.cb_var_force_close_boa_messages = Checkbutton(frm_settings,text='Force close BOA pop-ups',onvalue=1, offvalue=0, variable=self.var_force_close_boa_messages, command=self.display_input)
            self.cb_var_force_close_boa_messages.grid(row=5,column=1,padx=5, pady = 5)


            #Restart Excel after refresh Error
            self.var_Restart_Excel_after_refresh_Error = tk.IntVar(value=0)
            self.cb_var_Restart_Excel_after_refresh_Error = Checkbutton(frm_settings,text='Restart Excel after Refresh Error',onvalue=1, offvalue=0, variable=self.var_Restart_Excel_after_refresh_Error, command=self.display_input)
            self.cb_var_Restart_Excel_after_refresh_Error.grid(row=7,column=0,padx=5, pady = 5)

            #Save Query results in CSV file
            self.var_save_results_csv = tk.IntVar(value=0)
            self.cb_var_save_results_csv = Checkbutton(frm_settings,text='Save Query results in CSV',onvalue=1, offvalue=0, variable=self.var_save_results_csv, command=self.display_input)
            self.cb_var_save_results_csv.grid(row=7,column=1,padx=5, pady = 5)

            #Always refresh errorenous queries
            self.var_always_refresh_error_queries = tk.IntVar(value=0)
            self.cb_var_always_refresh_error_queries = Checkbutton(frm_settings,text='Always refresh queries with errors during last refresh',onvalue=1, offvalue=0, variable=self.var_always_refresh_error_queries, command=self.display_input)
            self.cb_var_always_refresh_error_queries.grid(row=7,column=2,padx=5, pady = 5)



            #Text Input fields
            self.txt_holiday_path = Entry(frm_settings,width=75)
            self.txt_holiday_path.grid(row=0,column=1,padx=5, pady = 5)
            self.txt_holiday_path.insert(0,inital_holiday_path)

            self.txt_secret_path = Entry(frm_settings,width=75)
            self.txt_secret_path.grid(row=1,column=1,padx=5, pady = 5)
            self.txt_secret_path.insert(0,inital_secret_path)

            self.txt_pass_to_encode_path = Entry(frm_settings,width=75)
            self.txt_pass_to_encode_path.grid(row=2,column=1,padx=5, pady = 5)
            self.txt_pass_to_encode_path.insert(0,inital_password_path)

            self.txt_encoded_path = Entry(frm_settings,width=75)
            self.txt_encoded_path.grid(row=3,column=1,padx=5, pady = 5)
            self.txt_encoded_path.insert(0,inital_encoded_path)

            btn_query_path = Button(frm_settings,text="Select File",command=(lambda:self.selectpath(self.txt_holiday_path))).grid(row=0,column=2,padx=5, pady = 5)  
            btn_secret_path = Button(frm_settings,text="Select File",command=(lambda:self.selectpath(self.txt_secret_path))).grid(row=1,column=2,padx=5, pady = 5) 
            btn_new_secret= Button(frm_settings,text="Create New Secret File",command=(lambda:self.write(os.path.join(file_path, 'Secret.txt'),str(self.create_secret_key())))).grid(row=1,column=3,padx=5, pady = 5) 
            btn_pass_to_encode_path = Button(frm_settings,text="Select File",command=(lambda:self.selectpath(self.txt_secret_path))).grid(row=2,column=2,padx=5, pady = 5) 
            btn_encode_file = Button(frm_settings,text="Encode Password File",command=(lambda:self.encode_secret(self.txt_secret_path.get(), self.txt_pass_to_encode_path.get(), os.path.join(file_path, 'PW_Encoded.txt')))).grid(row=2,column=3,padx=5, pady = 5)
            btn_encoded_path = Button(frm_settings,text="Select File",command=(lambda:self.selectpath(self.txt_encoded_path))).grid(row=3,column=2,padx=5, pady = 5)  
            btn_load_pw = Button(frm_settings,text="Load and use encoded passwords",command=(lambda:read_Password_Files(self.txt_encoded_path.get(),self.txt_secret_path.get()))).grid(row=3,column=3,padx=5, pady = 5)  
            btn_load_holiday = Button(frm_settings,text="Load and use holiday file",command=(lambda:self.holiday_load(self.txt_holiday_path.get()))).grid(row=0,column=3,padx=5, pady = 5)  
            
            #encode_secret(self,self.txt_secret_path, self.txt_pass_to_encode_path, os.path.join(file_path, 'PW_Encoded.txt')

            btn_close3 = Button(frm_settings, text="Close",command=self.close_quit).grid(row=8,column=1,padx=5, pady = 5)
        
             
            self.root.mainloop()
            #self.root.protocol('WM_DELETE_WINDOW', self.close_quit()) 

        def display_input(self):
            #triggered when checkboxes are changed
                global clear_excel_sessions, Supress_All_BOA_Messages
                if self.var_kill_excel.get() == 1:
                    clear_excel_sessions = True
                else:
                    clear_excel_sessions = False
                
                global Supress_All_BOA_Messages
                if self.var_supress_boa_msg.get() == 1:
                    Supress_All_BOA_Messages = True
                else:
                    Supress_All_BOA_Messages = False

                global Run_Refresh_All_Excel
                if self.var_refresh_all_excel.get() == 1:
                    Run_Refresh_All_Excel = True
                else:
                    Run_Refresh_All_Excel = False
                
                global Close_Excel_After_Completion
                if self.var_close_excel_after_completion.get() == 1:
                    Close_Excel_After_Completion = True
                else:
                    Close_Excel_After_Completion = False

                global Force_Close_BOA_Messages_Window
                if self.var_force_close_boa_messages.get() == 1:
                    Force_Close_BOA_Messages_Window = True
                else:
                    Force_Close_BOA_Messages_Window = False               
                
                global capture_runtimes
                if self.var_capture_runtimes.get() == 1:
                    capture_runtimes = True
                else:
                    capture_runtimes = False       
                
                global save_results_csv
                if self.var_save_results_csv.get() == 1:
                    save_results_csv = True
                else:
                    save_results_csv = False       
                
                global always_refresh_error_queries
                if self.var_always_refresh_error_queries.get() == 1:
                    always_refresh_error_queries = True
                else:
                    always_refresh_error_queries = False  

                global Restart_Excel_after_refresh_Error
                if self.var_Restart_Excel_after_refresh_Error.get() == 1:
                    Restart_Excel_after_refresh_Error = True
                else:
                    Restart_Excel_after_refresh_Error = False  



        def holiday_load(self,target):
            global loaded_holiday_path, use_loaded_holidays
            loaded_holiday_path = target
            use_loaded_holidays = True
            
        def encrypted_pw_load(self,encoded_pth, secret_pth):
            global loaded_holiday_path, use_loaded_encrypted_pw
            
            #If passwords were extracted successfully set to true to prevent overwriting
            result = read_Password_Files(encoded_pth,secret_pth)
            if result != False:
                use_loaded_encrypted_pw = True


        def selectdir(self,target):
            global filepath_selected, filename_selected, filenamefull_selected
        

            filepath_selected = fd.askdirectory(
            title='Select file',
            initialdir=file_path)


            #pattern = '[\w-]+?(?=\.)'

            # searching the pattern
            #a = re.search(pattern, filepath_selected)
            try:
                pattern = '([a-zA-Z0-9\s_\\.\-\(\):\!\{\}])+(.)$'
                a = re.search(pattern, filepath_selected)
                pattern = '[a-zA-Z0-9\s_\\.\-\(\):\!\{\}]+?(?=\.)'
                a = re.search(pattern, a.group())
                print('Selected folder ' + a.group())
                filenamefull_selected = os.path.basename(filepath_selected)
                filename_selected = a.group()
            except:
                filenamefull_selected = ""
                filename_selected = ""


            if filepath_selected:
                if target == self.txt_dir_path:
                    self.change_text(self.txt_dir_path, filepath_selected)
                else:
                    pass
                
                return filepath_selected




        def selectpath(self, target):
            global filepath_selected, filename_selected, filenamefull_selected, file_source_config
        
            filetypes = (
            ('All files', '*.*'),
            ('text files', '*.txt'),
            ('Excel files', '*.xls*'),
            ('CSV files', '*.csv')
        )

            filepath_selected = fd.askopenfilename(
            title='Select file',
            initialdir=file_path,
            filetypes=filetypes)


            #pattern = '[\w-]+?(?=\.)'

            # searching the pattern
            #a = re.search(pattern, filepath_selected)
            try:
                pattern = '([a-zA-Z0-9\s_\\.\-\(\):\!\{\}])+(.)$'
                a = re.search(pattern, filepath_selected)
                pattern = '[a-zA-Z0-9\s_\\.\-\(\):\!\{\}]+?(?=\.)'
                a = re.search(pattern, a.group())
                print('Selected file ' + a.group())
                filenamefull_selected = os.path.basename(filepath_selected)
                filename_selected = a.group()
            except:
                filenamefull_selected = ""
                filename_selected = ""


            if filepath_selected:
                if target == self.txt_password_path:
                    self.change_text(self.txt_password_path, filepath_selected)
                    self.change_text(self.txt_password_path2, filepath_selected)
                elif target == self.txt_config_path:
                    self.change_text(self.txt_config_path, filepath_selected)
                    self.change_text(self.txt_config_path2, filepath_selected)
                elif  target == self.txt_query_path:
                    file_source_config = True
                    self.change_text(self.txt_query_path, filepath_selected)
                elif  target == self.txt_holiday_path:
                    self.change_text(self.txt_holiday_path, filepath_selected)
                elif target == self.txt_secret_path:
                    self.change_text(self.txt_secret_path, filepath_selected)
                elif target == self.txt_pass_to_encode_path:
                    self.change_text(self.txt_pass_to_encode_path, filepath_selected)
                elif target == self.txt_encoded_path:
                    self.change_text(self.txt_encoded_path, filepath_selected)
                
                return filepath_selected

        def change_text(self, txtbox, txt):
            txtbox.delete(0,END)
            txtbox.insert(0,txt)

        def secret_encode(self, string, secret_key):
            """Encrypt the secret string"""
            #key = Fernet.generate_key()
            key = secret_key 

            f = Fernet(key)
            token = f.encrypt(string.encode('utf-8'))
            return token

            

        def encode_Password_Files(self, path_to_file, secret_key):
            #Check if password file exists
            file_exists = os.path.isfile(path_to_file)

            if file_exists == True:
                file = open(path_to_file, 'r')
                encoded = ''
                for line in file.readlines():
                    secret_string = str(self.secret_encode(line,secret_key).decode('utf-8')) + "\n"
                    encoded += secret_string
                    fname = re.split(r'[,;\n]',line.rstrip()) #using rstrip to remove the \n
                    
                
                return encoded
            else:
                return False

        
        def create_secret_key(self):
            secret_key = Fernet.generate_key()
            return secret_key

        def write(self,path_txt_file,txt_contents):
            with open(path_txt_file, 'w') as output:
                output.write(txt_contents)

            

        def encode_secret(self,secret_path, path_to_encode, target_path):
            sec = read_secret_file(str(secret_path))
            if sec == False:
                return False
            else:
                #Remove leading b from string
                if sec[0] =='b':
                    sec = sec[1:]
                enc = self.encode_Password_Files(path_to_encode, sec)
                if enc == False:
                    return False
                else:
                    self.write(target_path,enc)
                    return True

        def func_encode(self,secret_path, path_to_encode, target_path):
            a = self.encode_secret(secret_path, path_to_encode, target_path)
            
    
        def configuration_single_file(self):
            #Create config from query fileonfiguration_multiple_files
            self.root.quit
            global update_mode, filepath_selected, filenamefull_selected, filename_selected
            update_mode = "Create Configuration single"
            filepath_selected = self.txt_query_path.get()
            #Only reload pw if button on settings page wasn´t hit
            if use_loaded_encrypted_pw == False:
                read_Password_Files(self.txt_password_path2.get(),None)
            # searching the pattern
            #pattern = '[\w-]+?(?=\.)'
            pattern = '([a-zA-Z0-9\s_\\.\-\(\):\!\{\}])+(.)$'
            a = re.search(pattern, filepath_selected)
            pattern = '[a-zA-Z0-9\s_\\.\-\(\):\!\{\}]+?(?=\.)'
            a = re.search(pattern, a.group())
            print('Creating configuration file for ' + a.group())
            filenamefull_selected = os.path.basename(filepath_selected)
            filename_selected = a.group()
            root.destroy()
            
    
        def configuration_multiple_files(self):
            #Create config from config file
            global update_mode, filepath_selected, filenamefull_selected, filename_selected, file_source_config
            if file_source_config == True:
                file_source_config = True
                filepath_selected = self.txt_config_path2.get()
                # searching the pattern
                #pattern = '[\w-]+?(?=\.)'

                #Extract filename with extension, allows spaces, ., /, \, !, ,{ ,}
                pattern = '([a-zA-Z0-9\s_\\.\-\(\):\!\{\}])+(.)$'
                a = re.search(pattern, filepath_selected)
                #Get filename before '.'
                pattern = '[a-zA-Z0-9\s_\\.\-\(\):\!\{\}]+?(?=\.)'
                a = re.search(pattern, a.group())
                print('Creating configuration file for files in ' + a.group())
                filenamefull_selected = os.path.basename(filepath_selected)
                filename_selected = a.group()
            else:
                file_source_config = False
                filepath_selected = self.txt_dir_path.get()
            
            update_mode = "Create Configuration multiple"
            
            #Only reload pw if button on settings page wasn´t hit
            if use_loaded_encrypted_pw == False:
                read_Password_Files(self.txt_password_path2.get(),None)

            root.destroy()

        def configuration_multiple_files_dir(self):
            global file_source_config
            file_source_config = False
            self.configuration_multiple_files()
            file_source_config = False
   
        def refresh(self):
            global update_mode, filepath_selected, filenamefull_selected, filename_selected
            update_mode = "Refresh"
            filepath_selected = self.txt_config_path.get()
            #Only reload pw if button on settings page wasn´t hit
            if use_loaded_encrypted_pw == False:
                read_Password_Files(self.txt_password_path.get(),None)
            
            # searching the pattern
            #pattern = '[\w-]+?(?=\.)'
            pattern = '([a-zA-Z0-9\s_\\.\-\(\):\!\{\}])+(.)$'
            a = re.search(pattern, filepath_selected)
            pattern = '[a-zA-Z0-9\s_\\.\-\(\):\!\{\}]+?(?=\.)'
            a = re.search(pattern, a.group())
            print('Creating configuration file for ' + a.group())
            filenamefull_selected = os.path.basename(filepath_selected)
            filename_selected = a.group()
            print('Refreshing files in ' + a.group())
            root.destroy()




    root = Tk()
    get_inital_file_paths()
    #Defines size of interface
    window_main = BOA_Interface(root,'BOA Interface',1250,350)
    if update_mode == "No Input":
        sys.exit()
    
    if (update_mode == "Refresh" or update_mode == "Create Configuration multiple") and start_mode != "Scheduled":
        CONFIG_PATH = filepath_selected



def get_logger(logger_name, log_path):
        global logger, location
        # string formatter
        format_string = '%(asctime)s | %(name)s | %(module)s | %(funcName)s | [%(levelname)s] | %(message)s'
        logFormatter = logging.Formatter(format_string, datefmt='%Y%m%d %H:%M')

        # initiate root logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)

        # set up console logger
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.DEBUG)
        console_handler.setFormatter(logFormatter)
        logger.addHandler(console_handler)

        # set up file logger
        log_filename = datetime.now().strftime("%Y%m%d")
        if update_mode == "Refresh" or update_mode == "Create Configuration multiple":
            try:
                location = log_path / f'{log_filename}.log'  # create a path do the log
                #location = os.path.join(log_path,f'{log_filename}.log')
            except:
                location = log_path + f'{log_filename}.log'  # create a path do the log
        else:
            try:    
                location = os.path.join(log_path,f'{log_filename}.log')
            except:
                location = log_path + f'{log_filename}.log'  # create a path do the log

        logfile_handler = TimedRotatingFileHandler(location, when='D', interval=1, backupCount=30)
        logfile_handler.setLevel(logging.INFO)  # do not print DEBUG messages to file
        logfile_handler.setFormatter(logFormatter)
        logger.addHandler(logfile_handler)

        return logger, location



def search_boa_message_window():
    global no_message_window_found, BOA_message_closed, Monitoring_Windows
    CoInitialize() #pythoncom.CoInitialize()
    Monitoring_Windows = True
    BOA_message_closed = False
    no_message_window_found = True
    print('Start checking for BOA message pop-up windows')
    while no_message_window_found and Monitoring_Windows:

        time.sleep(10)
        #Only messages which pop-up as a window are checked, notifications that do not pop-up are not checked
        #Critical errors disconnect the addin completely and supressing the messages via the built-in macro does not work, this function is looking for any window with title "Messages" which the BOA plugin is showing for messages in the tested environment
        try:
            handle = FindWindow(None, r'Messages') #win32gui.FindWindow
            if handle != 0:
                no_message_window_found = False
                print('Message window found')
            else:
                no_message_window_found = True

        except:
            no_message_window_found = True
            print('Error searching BOA Message window')
    
    #Bevore stopping check for message window
    try:
        handle = FindWindow(None, r'Messages')
        if handle != 0:
            #no_message_window_found = False
            print('BOA Message window will be closed')
            close_boa_message_window()
        else:
            pass
            #no_message_window_found = True
    except:
        #no_message_window_found = True
        print('Error searching BOA Message window')
    
    print('Completed checking for BOA message pop-up windows')



def close_boa_message_window():
    #Critical errors disconnect the addin completely and supressing the messages via the built-in macro does not work, this function is looking for any window with title "Messages" which the BOA plugin is showing for messages in the tested environment - then this window is closed
    global BOA_message_closed
    BOA_message_closed =False
    try:
        handle = FindWindow(None, r'Messages')
        if handle != 0:
            PostMessage(handle,WM_CLOSE,0,0) #win32con.WM_CLOSE
            BOA_message_closed = True
            return True
        else:
            return False
    except:
        logger.INFO('Error closing BOA Message window: ' + str(Exception))
        print('Error closing BOA Message window')
        return False

# Split path to filename, filepath and fullpath
def Split_path(pth_to_Split):
    global split_extension, split_nameonly, split_fullname
    split_extension = os.path.splitext(os.path.basename(pth_to_Split))[1]
    split_nameonly = os.path.splitext(os.path.basename(pth_to_Split))[0] #os.path.basename(pth_to_Split).split('.', 1)[0]
    split_fullname = os.path.basename(pth_to_Split)
         
def check_if_Excel_runs():
    global xl_Instance
    try:
        xl_Instance = win32.GetActiveObject("Excel.Application")
        # If there is NO error at this stage, Excel is already running
        return True

    except:
        return False 


def return_running_excel():
    
    time.sleep(3)

    exl_running = find_excel_instances()
    try:
        xl = win32.GetActiveObject("Excel.Application")
        if 'None.Application' in str(xl):
            print('Error activating MS Excel')
            raise AttributeError("Excel not found")
    except:
        # https://gist.github.com/rdapaz/63590adb94a46039ca4a10994dff9dbe?permalink_comment_id=2918299#gistcomment-2918299
        # Corner case dependencies.
        # Remove cache and try again
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py')) 
        xl = win32.GetActiveObject("Excel.Application")
        if 'None.Application' in str(xl) and exl_running == True:
            kill_excel_instances()
            xl = open_excel()
    return xl

def check_if_excel_file_is_open(filename_with_extension):
    if xl_Instance.Workbooks.Count > 0:
        # if none of opened workbooks matches the name, openes my_workbook 

        if not any(i.Name == filename_with_extension for i in xl_Instance.Workbooks): 
            return False
        else:
            return True
    else:
        return False

def close_checked_file(filpath_with_extension):
    if xl_Instance.Workbooks.Count > 0:
        # if none of opened workbooks matches the name, openes my_workbook 
        for i in xl_Instance.Workbooks: 
            Split_path(filpath_with_extension)
            wb_fn = split_fullname
            if i.Name== wb_fn:
                i.Close()


def import_global_configurations(config_location):
    global df_global_configs, path_log, client_config, user_config,password_config, language_config, passfile_config, secret_config,holidays_config,log_config,encoded_config,password,user,client,language, inital_password_path, inital_secret_path, inital_encoded_path, inital_holiday_path, executable_config, executable_path, Send_Mail_Error_config, Send_Mail_Error, Err_Count, settings_sheet_available
    """import all the necessary information so that the script can work well"""
    try:
        #Settings are not madatory for creating configurations
        if update_mode != "Create Configuration multiple":
            df_global_configs = pd.read_excel(config_location, sheet_name='Settings')
        else:
            try:
                df_global_configs = pd.read_excel(config_location, sheet_name='Settings')
            except:
                settings_sheet_available = False
                df_global_configs = pd.DataFrame()

    except PermissionError:
        if start_mode == "Scheduled":

            if check_if_Excel_runs() == True:
                    try:
                        close_checked_file(config_location)
                        time.sleep(2)

                        df_global_configs = pd.read_excel(config_location, sheet_name='Settings')
                    except:
                        print('No permission to open file, file probably opened')
                        Err_Count = Err_Count +1
        else:
            if messagebox.askretrycancel("Error", "Configuration File opened, close and retry or cancel.") :
                import_global_configurations(CONFIG_PATH)  
            else:
                quit()
    except ValueError:
        print('Probably wrong file selected, expected a Configuration file but no Settings was sheet found')
        if start_mode != "Scheduled":
            if messagebox.showerror("Error", "Probably wrong file selected, expected a Configuration file but no Settings was sheet found. Check the file and or select a different file.") :
                 quit()
        Err_Count = Err_Count + 1


    
    #Check which fields are filled
    client_config =True
    user_config =True
    password_config =True
    language_config =True
    passfile_config =True
    secret_config =True
    holidays_config =True
    encoded_config = True
    executable_config = True
    Send_Mail_Error_config = True
    log_config = True


    #Try except prevents failures when older config files are used
    try:
        if (df_global_configs.query('Setting=="logon-client"')['Value'].isnull()).values[0]:
            client_config = False
        else:
            client = df_global_configs.query('Setting=="logon-client"')['Value'].values[0]
    except:
        client_config = False

    try:
        if (df_global_configs.query('Setting=="path-log_directory"')['Value'].isnull()).values[0]:
            log_config = False
            path_log = file_path
        else:
            try:
                try:
                    path_log = pathlib.Path(df_global_configs.query('Setting=="path-log_directory"')['Value'].values[0]).resolve()
                except:
                    path_log = pathlib.Path(df_global_configs.query('Setting=="path-log_directory"')['Value'].values[0])
            except:
                path_log = file_path
    except:
        log_config = False
        path_log = file_path


    try:
        if (df_global_configs.query('Setting=="logon-user"')['Value'].isnull()).values[0]:
            user_config = False
        else:
            user = df_global_configs.query('Setting=="logon-user"')['Value'].values[0]
    except:
        user_config = False

    try:
        if (df_global_configs.query('Setting=="logon-password"')['Value'].isnull()).values[0]:
            password_config = False
        else:
            password = df_global_configs.query('Setting=="logon-password"')['Value'].values[0]
    except:
        password_config = False

    try:
        if (df_global_configs.query('Setting=="logon-language"')['Value'].isnull()).values[0]:
            language_config = False
        else:
            language = df_global_configs.query('Setting=="logon-language"')['Value'].values[0]
    except:
        language_config = False

    try:
        if (df_global_configs.query('Setting=="path-password_file"')['Value'].isnull()).values[0]:
            passfile_config = False
        else:
            pass
    except:
        passfile_config = False

    try:
        if (df_global_configs.query('Setting=="path-secret_file"')['Value'].isnull()).values[0]:
            secret_config = False
        else:
            pass
    except:
        secret_config = False

    try:
        if (df_global_configs.query('Setting=="path-holiday_days_file"')['Value'].isnull()).values[0]:
            holidays_config = False
        else:
            pass
    except:
        holidays_config = False

    try:
        if (df_global_configs.query('Setting=="path-encoded_pw_file"')['Value'].isnull()).values[0]:
            encoded_config = False
        else:
            pass
    except:
        encoded_config = False

    try:
        if (df_global_configs.query('Setting=="Run Script / EXE after refresh"')['Value'].isnull()).values[0]:
            executable_config = False
        else:
            executable_path = df_global_configs.query('Setting=="Run Script / EXE after refresh"')['Value'].values[0]
    except:
        executable_config = False

    try:
        if (df_global_configs.query('Setting=="Send E-Mail to in case of Error"')['Value'].isnull()).values[0]:
            Send_Mail_Error_config = False
        else:
            Send_Mail_Error = df_global_configs.query('Setting=="Send E-Mail to in case of Error"')['Value'].values[0]
    except:
        Send_Mail_Error = ""
        Send_Mail_Error_config = False


    if start_mode == "Scheduled":
        if passfile_config == True:
            inital_password_path = df_global_configs.query('Setting=="path-password_file"')['Value'].values[0]
            read_Password_Files(inital_password_path,None)
        if secret_config == True:
            inital_secret_path = df_global_configs.query('Setting=="path-secret_file"')['Value'].values[0]
            if encoded_config == True:
                inital_encoded_path = df_global_configs.query('Setting=="path-encoded_pw_file"')['Value'].values[0]
                read_Password_Files(inital_encoded_path,inital_secret_path)
        if holidays_config == True:
            inital_holiday_path = df_global_configs.query('Setting=="path-holiday_days_file"')['Value'].values[0]
        #Load from inital paths if config is empty
        if passfile_config == False or (secret_config == False or encoded_config == False):
                get_inital_file_paths()
    return df_global_configs, path_log
    


def Workdays():
    global today_is_workday, current_workday, currentDate, holidays, default_week_mask, week_mask
    currentDate = date.today()
    currentYear = currentDate.year
    previousYear = currentDate.year -1
    nextYear = currentDate.year +1

    startdate = datetime(previousYear, 1,1)
    enddate  = datetime(nextYear,12,31)

    if start_mode == "Scheduled" and holidays_config == False: 
        use_loaded_holidays = False
    else:
        try:
            if use_loaded_holidays == True and loaded_holiday_path != '':
                df_holidays = pd.read_csv(loaded_holiday_path, sep=',', parse_dates=['Holidays'])#, dtype = {'Holidays':datetime})
        except:
            pass
        try:
            #Holiday days configuration
            if holidays_config == True and use_loaded_holidays != True:
                df_holidays = pd.read_csv(df_global_configs.query('Setting=="path-holiday_days_file"')['Value'].values[0], sep=',', parse_dates=['Holidays'])
            else:
                if inital_holiday_path != '':
                    #Read Holiday days configuration from csv
                    df_holidays = pd.read_csv(inital_holiday_path, sep=',', parse_dates=['Holidays'])#, dtype = {'Holidays':datetime})
        except:
            pass
    try:
        print(str(df_holidays['Holidays'].loc[df_holidays.idxmax()]) +' holiday entries found')
    except:
        #add dummy date
        df_holidays = []
        df_holidays.append(
        {
            'Holidays': datetime(2099, 1, 1)
        }
        )
        df_holidays = pd.DataFrame.from_dict(df_holidays)
        print('No holiday entries found')

    df_holidays['Holidays'] = df_holidays['Holidays'].dt.date
    df_holidays['Holidays']= df_holidays['Holidays'].apply(str)

    #Create list of holiday dates
    holidays = df_holidays['Holidays'].values.tolist()

    dates = pd.date_range(start=startdate, end=enddate)
    df_dates = pd.DataFrame({'Date':pd.date_range(start=startdate, end=enddate)})
    df_dates['Month'] = df_dates['Date'].dt.month
    df_dates['Day'] = df_dates['Date'].dt.day
    df_dates['Year'] = df_dates['Date'].dt.year
    df_dates['Date'] = df_dates['Date'].dt.date
    df_dates['Date']= df_dates['Date'].apply(str)


    #Split into separate lists
    dates  = df_dates['Date'].values.tolist()
    months  = df_dates['Month'].values.tolist()
    days  = df_dates['Day'].values.tolist()
    years  = df_dates['Year'].values.tolist()

    # Check dates vs Holidays to determine whether date is workday
    if default_week_mask:
        Is_workday =np.is_busday(dates,
                        holidays=holidays, weekmask=week_mask)
    else:
        Is_workday =np.is_busday(dates,
                        holidays=holidays)

    #Merge lists
    zipped = list(zip(dates, years, months, days, Is_workday))
    #Create dataframe from lists
    df = pd.DataFrame(zipped, columns=['Date','Year', 'Month', 'Day', 'Is_Workday'])

    #Add column Workday to dataframe
    df = df.assign(Workday='')
    


    #Inital year.month
    inital_month = str(df['Year'][0]) + str(df['Month'][0])



    today_is_workday = False
    counter = 0
    #Iterate over all dates, add counter for workdays per month
    for index, row in df.iterrows():
        if str(row["Year"]) + str(row["Month"]) == inital_month:
            if row["Is_Workday"] == True:
                counter = counter +1
                df.loc[index,['Workday']] = counter
        else:
            counter = 0
            inital_month = str(row["Year"]) + str(row["Month"])
            if row["Is_Workday"] == True:
                counter = counter +1
                df.loc[index,['Workday']] = counter
        #Save current workday
        if row["Date"] == str(currentDate):
            current_workday = df.loc[index,['Workday']].item()
            #Save if today is workday
            if row["Is_Workday"] == True:
                today_is_workday = True


    return today_is_workday, current_workday



def import_datasources(config_location):
    global df_datasource_configs
    """import all the necessary information so that the script can work well"""
    df_datasource_configs = pd.read_excel(config_location, sheet_name='Queries')
    
    return df_datasource_configs

def import_files(config_location):
    global df_files_configs
    """import all the necessary information so that the script can work well"""
    df_files_configs = pd.read_excel(config_location, sheet_name='Files')

    #Try to resolve paths
    try:
        df_files_configs_resolved = df_files_configs
        df_files_configs_resolved = df_files_configs_resolved.reset_index()
        for index, row in df_files_configs_resolved.iterrows():
            row['Filepath'] = pathlib.Path(row['Filepath']).resolve()
            row['Fullpath'] = pathlib.Path(row['Fullpath']).resolve()
            #Joining with empty path ensures ending with os.separator
            try:
                row['Filepath'] = os.path.join(row['Filepath'] , '')
            except: 
                pass
        df_files_configs = df_files_configs_resolved
    except:
        pass
    
    return df_files_configs


global global_configs_df


#Variables are only available if existing configuration file is used, therfore paths need to be fixed. Will not be used when configuration is to be created from files in a folder

if update_mode == "Refresh" or update_mode == "Create Configuration multiple" and file_source_config == True:
    global_configs_df, LOG_PATH = import_global_configurations(CONFIG_PATH)  

else:
    #No data from config file
    global_configs_df = []
    client_config =False
    user_config =False
    password_config =False
    language_config =False
    passfile_config =False
    secret_config =False
    holidays_config =False
    encoded_config = False
    filename = filenamefull_selected
    #LOG_PATH = file_path.replace('\\','/')
    try:
        LOG_PATH = pathlib.Path(file_path).resolve()
    except:
        LOG_PATH = file_path
    


#Check if log path exists
if os.path.exists(LOG_PATH):
    try:
        logger, LOG_FILEPATH = get_logger(__name__, LOG_PATH) 
    except:
        #Resolve path e.g. if network drive is used and save in parent directory
        logger, LOG_FILEPATH = get_logger(__name__, pathlib.Path(LOG_PATH).parent.resolve()) 

else:
    print('Could not find log path ' + str(LOG_PATH) + ' --- Log will be saved in: ' +  str(file_path.replace('\\','/')))
    #LOG_PATH = file_path.replace('\\','/')
    LOG_PATH = file_path
    logger, LOG_FILEPATH = get_logger(__name__, LOG_PATH) 


#@retry(reraise=True, wait=wait_fixed(10), before_sleep=before_sleep_log(logger, logging.DEBUG), stop=stop_after_attempt(3))

def get_configurations(config_location):
    global global_configs, data_sources, variables_filters, update_mode
    """get all necessary dataframes to serve the SapRefresh class"""
    if update_mode != "Create Configuration multiple":
        global_configs = pd.read_excel(config_location, sheet_name='Settings')
        data_sources = pd.read_excel(config_location, sheet_name='Queries')
        variables_filters = pd.read_excel(config_location, sheet_name='Variables_Filters', na_filter=False) # nan being replaced with empty string, otherwise values sometimes changed from nan to 1
    else:
        try:
            global_configs = pd.read_excel(config_location, sheet_name='Settings')
        except:
            global_configs =pd.DataFrame()
        try:
            data_sources = pd.read_excel(config_location, sheet_name='Queries')
        except:
            data_sources =pd.DataFrame()
        try:
            variables_filters = pd.read_excel(config_location, sheet_name='Variables_Filters', na_filter=False) # nan being replaced with empty string, otherwise values sometimes changed from nan to 1
        except:
            variables_filters =pd.DataFrame()

    print('Successfully loaded the configurations')
    return global_configs, data_sources, variables_filters






def timeit(func):
    """Print the runtime of the decorated function"""
    @wraps(func)
    def wrapper_timer(*args, **kwargs):
        global run_time
        start_time_wrapper = time.perf_counter()  # 1
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        print('\n', f'Running {func.__name__!r} at: ', now)

        # start waiting spinner
        spinner = Halo(text='Loading' + '\n', spinner='dots')
        spinner.start()

        value = func(*args, **kwargs)

        # stop waiting spinner
        spinner.stop()

        end_time = time.perf_counter()  # 2
        run_time = end_time - start_time_wrapper  # 3
        if run_time < 60:
            print(f"Finished {func.__name__!r} in {run_time:.2f} seconds\n")
        if 60 <= run_time < 3600:
            run_time = run_time / 60  # converts to minutes
            print(f"Finished {func.__name__!r} in {run_time:.1f} minutes\n")
        if run_time >= 3600:
            run_time = run_time / 3600  # converts to minutes
            print(f"Finished {func.__name__!r} in {run_time:.1f} hours\n")
        return value
    return wrapper_timer
    

def get_current_user_processes():
   #Get all processes for current user
   print('Loading running processes from all logged users and search for running Excel instance')
   csv_output = subprocess.check_output(["tasklist","/FI","USERNAME eq {}".format(os.getenv("USERNAME")),"/FO","CSV"]).decode("ascii","ignore")
   cr = csv.reader(csv_output.splitlines())
   next(cr)   # skip title lines
   return {int(row[1]):row[0] for row in cr}

def kill_excel_instances():

    d = get_current_user_processes()
    #Only check processes of current user - necessary if working on a shared machine with several users logged, otherwise processes running on other users cant be killed and cause an error
    processes = [proc for proc in process_iter() if proc.pid in d] #psutil.process_iter()

    for proc in processes:
        if proc.name().lower() == "excel.exe":
            print("A running Excel instance was found. The script is going to kill it as a sanity check procedure.")
            proc.kill()




def find_excel_instances():
    d = get_current_user_processes()
    #Only check processes of current user - necessary if working on a shared machine with several users logged, otherwise processes running on other users cant be killed and cause an error
    processes = [proc for proc in process_iter() if proc.pid in d] #psutil.process_iter()
    Excel_found = False
    for proc in processes:
        if proc.name().lower() == "excel.exe":
            print("A running Excel instance was found.")
            Excel_found = True

    if Excel_found:
        return True
    else:
        return False           


    

def open_excel():
    """Start a instance of Excel application"""
    global xl_Instance
    #xl_Instance = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        xl_Instance = win32.gencache.EnsureDispatch('Excel.Application')
    except:
        #deletes gen_py folder in case of Attribute error
        xl_Instance = dispatch('Excel.Application')
    return xl_Instance


class Excel:
    global xl_Instance
    def __init__(self):

        if check_if_Excel_runs() == False:
            self.xlApp  = xl_Instance = open_excel()
        else:
            try:
                if 'None' in str(self.xlApp):
                    self.xlApp = xl_Instance = return_running_excel()
            except:
                self.xlApp = xl_Instance = return_running_excel()
        try:
            if self.xlApp.Visible == False:
                self.xlApp.Visible = True

            if self.xlApp.DisplayAlerts == True:
                self.xlApp.DisplayAlerts = False
        except:
            pass


def get_crosstab_size(xl_instance, filename, crosstab):

    try:
        named_range = filename + '!SAP' + crosstab
        named_range_rows = xl_Instance.Range(named_range).Rows.Count
        named_range_columns  = xl_Instance.Range(named_range).Columns.Count
        crosstab_row  = xl_Instance.Range(named_range).Row
        crosstab_column = xl_Instance.Range(named_range).Column

        if named_range_rows == 1 and named_range_columns == 1:
            named_range_value = xl_Instance.Range(named_range).Value
            return False, named_range_value, named_range_rows, named_range_columns, crosstab_row, crosstab_column

        return True, None, named_range_rows, named_range_columns, crosstab_row, crosstab_column
    except:
        return False, None, None, None, None, None


@retry(reraise=True, wait=wait_fixed(10), before_sleep=before_sleep_log(logger, logging.DEBUG), stop=stop_after_attempt(3))
@timeit
def open_workbook(xl_Instance, path):
    global wb

    """
    Open the file in the new Excel instance,
        The 1st false: don't update the links
        The 2nd false: and don't open it read-only
    """

    try:
        wb = xl_Instance.Workbooks.Open(path, False, False)
    except:
        #2147023170 The remote procedure call failed
        #2147023174 The RPC server is unavailable
        #if e.args[0] == -2147352567:
        #if Exception.args[0] == -2147023174:
         #   print("RPC server Error")
        #Try to resolve path in case of network drive
        #resolved_path = pathlib.Path(filepath).resolve()
        logger.info("Could not open file, closing and restarting Excel")
        print("Could not open file, closing and restarting Excel")
        kill_excel_instances()
        xl_Instance = open_excel()
        wb = xl_Instance.Workbooks.Open(path, False, False)


    try:
        #Turn off autosave, not compatible with BOA
        if wb.AutoSaveOn:
            wb.AutoSaveOn = False
    except:
        pass

    return wb, xl_Instance 

#Disabled Excel addins can only be enabled by deleting the Key from the registry and restarting Excel
def enable_addins(xl_Ins):
    app_version = xl_Ins.Application.Version
    strPath = 'Software\\Microsoft\\Office\\' + app_version + '\\Excel\\Resiliency\\DisabledItems\\'
    try:
        aReg = ConnectRegistry(None, HKEY_LOCAL_MACHINE)
        key = OpenKey(HKEY_CURRENT_USER, strPath)
        DeleteKey(HKEY_CURRENT_USER,strPath)
    except FileNotFoundError:
        #No disabled plugins found
        return True
    except:
        print('Enabling disabled addins, failed ')
        print(str(Exception))
        return False

    return True

def check_addin_enabled(xl_Ins):
    try:
        """Force the plugin to be enabled in the instance of Excel"""
        for addin in xl_Ins.Application.COMAddIns:
            if addin.ProgId == 'SapExcelAddIn': #  progID used before, raised AttributeError
                if not addin.Connect:
                    return False
                elif addin.Connect:
                    return True
        return False
    except:
        return False


@timeit
def ensure_addin(xl_Ins):
    try:
        """In order to kill Excel popup we can hit enter"""
        shell = win32.Dispatch("WScript.Shell")
        shell.SendKeys("{ENTER}")
    except:
        pass
    """Force the plugin to be enabled in the instance of Excel"""
    for addin in xl_Ins.Application.COMAddIns:
        if addin.ProgId == 'SapExcelAddIn': #  progID used before, raised AttributeError
            if not addin.Connect:
                addin.Connect = True
            elif addin.Connect:
                addin.Connect = False
                addin.Connect = True

    #print('\n', 'Is SapExcelAddIn Enabled?', xl_Ins.Application.COMAddIns['SapExcelAddIn'].Connect)


def ensure_wb_active(xl_Ins, filename):
    global wb_name
    """Check if WorkBook is active, otherwise Logon fails if another wb is selected"""
    wb_name = xl_Ins.Application.ActiveWorkbook.Name
    print('Current workbook active is:', wb_name)
    if wb_name != filename:
        print(f'The desired wb ({filename}) is not active. Forcing it...', end='')
        xl_Ins.Application.Windows(filename).Activate()
        new_wb = xl_Ins.Application.ActiveWorkbook.Name
        print('Done! Workbook active is:', new_wb)


def optimize_instance(xl_Instance, action):
    """deals with excel calculation optimization"""
    if action == 'start':
        xl_Instance.Visible = True
        xl_Instance.DisplayAlerts = False
        xl_Instance.ScreenUpdating = False
        # xl_Instance.EnableEvents = False  # todo: check in reference code if this statement cause negative behavior in the script before uncomment it
    elif action == 'stop':
        xl_Instance.DisplayAlerts = True
        xl_Instance.ScreenUpdating = True
        # xl_Instance.EnableEvents = True
        try:
            xl_Instance.Application.Cursor = cst.xlDefault
        except:
            pass
        xl_Instance.Application.StatusBar = ''  # equivalent to vbNullString

def screenupdating_instance(xl_Instance, action):
    """deals with excel calculation optimization"""
    if action == 'start':
        xl_Instance.ScreenUpdating = True
    elif action == 'stop':
        xl_Instance.ScreenUpdating = False


def calculation_state(xl_Ins, action, state=None):
    """set the Calculation State to Manual in Excel"""
    #-4135 = Automatic
    #-4105 = Manual
    calcstate = ''
    try:
        calcstate = cst.xlCalculationManual
    except:
        calcstate = -4105
    if action == 'start':
        state = xl_Ins.Application.Calculation

        if xl_Ins.Application.Calculation != calcstate: # cst.xlCalculationManual: #
            xl_Ins.Application.Calculation = calcstate # cst.xlCalculationManual #
    elif action == 'stop' and state is not None:
        xl_Ins.Application.Calculation = state

    return state



def close_workbook(wb_Instance):

    #global refresh_running, Monitoring_Windows

    #Start Thread to monitor for BOA pop-up
    #refresh_running = True
    #thread_monitor_windows = Thread(target=search_boa_message_window)
    #thread_monitor_windows.start()


    """Save the file and close it if in refresh mode"""
    if update_mode == "Create Configuration single" or update_mode == "Create Configuration multiple":
        pass
    else:
        wb_Instance.Save()
    wb_Instance.Close()

    #Stop thread after refresh
   #Monitoring_Windows = False
   #thread_monitor_windows.join()
   #refresh_running = False


@timeit
def Save_workbook_input(workbook):
    workbook.Save()

@timeit
def Save_copy_as_input(workbook, target):
    workbook.SaveCopyAs(target)


class SpinnerCursor(object):
    def __init__(self, text, spinner):
        """Create a spinner to show execution while waiting for processes"""
        self.spinner = Halo(text=text, spinner=spinner)

    def start(self):
        """start the animation of cursor"""
        self.spinner.start()

    def stop(self):
        """stop the animation of cursor"""
        self.spinner.succeed('End!')


def get_time_intelligence():
    """get all the time intelligence references variables to the application"""
    values = dict()
    values['current_period'] = date.today()
    delta = relativedelta(months=-1)
    values['previous_period'] = values['current_period'] + delta
    values['year_current_period'] = values['current_period'].year
    values['year_previous_period'] = values['previous_period'].year
    values['range_current_month'] = '{} - {}'.format(values['current_period'].month, values['current_period'].month)
    values['range_previous_month'] = '{} - {}'.format(values['previous_period'].month, values['previous_period'].month)
    values['key_date'] = values['current_period'].strftime("%d.%m.%Y")
    return values



def append_values(dict_target,var,name,val):
        dict_target.append(
        {
            'Variable': var,
            'Name': name,
            'Value': val
        }
    )


def append_month_values(dict_target,offset,MMYYYY,MMMYYYY,FirstDate,LastDate):
        dict_target.append(
        {
            'Offset': offset,
            'MM.YYYY': MMYYYY,
            'MMM.YYYY': MMMYYYY,
            'FirstDate':FirstDate,
            'LastDate':LastDate
        }
    )

def time_intelligence():
    global df_month_values, df_time_values, default_week_mask, week_mask

    """get all the time intelligence references variables to the application"""
    #values = dict()
    values = []
    currentDate = date.today()
    #Current 
    currentDay = currentDate.day
    currentMonth = currentDate.month
    currentYear= currentDate.year

    First_Day_CM = datetime(currentYear, currentMonth, 1)

    CM_MM_YYYY = str(currentMonth).zfill(2) + '.' + str(currentYear)
    CM_MMM_YYYY = str(currentMonth).zfill(3) + '.' + str(currentYear)
    
    append_values(values,'Today', 'Current Date Today MM.DD.YYYY',currentDate.strftime("%d.%m.%Y"))


    #Last Workay
    Is_workday_check = False
    count = 1

    try:
        while Is_workday_check == False:
            dt = datetime(currentYear,currentMonth,currentDay) - relativedelta(days=count)
            count = count +1

            # Check dates vs Holidays to determine whether date is workday
            if default_week_mask:
                Is_workday_check =np.is_busday(dt.date(),
                            holidays=holidays, weekmask=week_mask)
            else:
                Is_workday_check =np.is_busday(dt.date(),
                            holidays=holidays)

            if Is_workday_check ==True:
                LastWorkday = dt
                break
    except:
        try:
            while Is_workday_check == False:
                dt = datetime(currentYear,currentMonth,currentDay) - relativedelta(days=count)
                count = count +1
                
                # Check dates vs Holidays to determine whether date is workday
                if default_week_mask:
                    Is_workday_check =np.is_busday(dt.date(),weekmask=week_mask)
                else:
                    Is_workday_check =np.is_busday(dt.date())

                if Is_workday_check ==True:
                    LastWorkday = dt
                    break
        except:
            LastWorkday = datetime(currentYear,currentMonth,currentDay) - relativedelta(days=1)

    append_values(values,'LastWD', 'LastWorkday MM.DD.YYYY',LastWorkday.strftime("%d.%m.%Y"))
    append_values(values,'CM_MM_YYYY', 'Current Month MM.YYYY' ,CM_MM_YYYY)
    append_values(values,'CM_MMM_YYYY','Current Month MMM.YYYY',CM_MMM_YYYY)

    #Last Month
    Last_Date_LM = ((datetime((date.today()).year, (date.today() - relativedelta(months=0)).month, 1)) - relativedelta(days=1))
    First_Day_LM = datetime((date.today() - relativedelta(months=1)).year, (date.today() - relativedelta(months=1)).month,1)
    
    LM_MM_YYYY = str(Last_Date_LM.month).zfill(2) + '.' + str(Last_Date_LM.year)
    LM_MMM_YYYY = str(Last_Date_LM.month).zfill(3) + '.' + str(Last_Date_LM.year)

    append_values(values,'LM_MM_YYYY', 'Last Month MM.YYYY' ,LM_MM_YYYY)
    append_values(values,'LM_MMM_YYYY','Last Month MMM.YYYY',LM_MMM_YYYY)




    #First day of the year
    firstdayCY = datetime(currentYear, 1, 1)

    #First day of the past years
    firstdayPY = datetime(currentYear-1, 1, 1)
    firstdayPYPY = datetime(currentYear-2, 1, 1)

    firstdayPYPYPY = datetime(currentYear-3, 1, 1)

    #Last day of the year
    lastdayCY = (datetime(currentYear+1, 1, 1) - relativedelta(days=1))

    #Last day of the past years
    lastdayPY = (datetime(currentYear, 1, 1) - relativedelta(days=1))
    lastdayPYPY = (datetime(currentYear-1, 1, 1) - relativedelta(days=1))
    lastdayPYPYPY = (datetime(currentYear-2, 1, 1) - relativedelta(days=1))

    #PYTD

    
    PY_MM_DD_YYYY = str(firstdayPY.strftime("%d.%m.%Y")) + ' - ' + str(lastdayPY.strftime("%d.%m.%Y"))
    PYTD_MM_DD_YYYY = str(firstdayPY.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    
    PYPY_MM_DD_YYYY = str(firstdayPYPY.strftime("%d.%m.%Y")) + ' - ' + str(lastdayPYPY.strftime("%d.%m.%Y"))
    PYPYPY_MM_DD_YYYY = str(firstdayPYPYPY.strftime("%d.%m.%Y")) + ' - ' + str(lastdayPYPYPY.strftime("%d.%m.%Y"))

    PYTD_MM_YYYY = str(firstdayPY.month).zfill(2) +  '.' + str(firstdayPY.year) + ' - '  + LM_MM_YYYY
    PYPYTD_MM_YYYY = str(firstdayPYPY.month).zfill(2) +  '.' + str(firstdayPYPY.year) + ' - '  + LM_MM_YYYY
    PYPYPYTD_MM_YYYY = str(firstdayPYPYPY.month).zfill(2) +  '.' + str(firstdayPYPYPY.year) + ' - '  + LM_MMM_YYYY

    PYTDLM_MM_YYYY = str(firstdayPY.month).zfill(2) +  '.' + str(firstdayPY.year) + ' - '  + CM_MM_YYYY
    PYPYTDLM_MM_YYYY = str(firstdayPYPY.month).zfill(2) +  '.' + str(firstdayPYPY.year) + ' - '  + CM_MM_YYYY
    PYPYPYTDLM_MM_YYYY = str(firstdayPYPYPY.month).zfill(2) +  '.' + str(firstdayPYPYPY.year) + ' - '  + CM_MM_YYYY





    PY_MM_YYYY = str(firstdayPY.month).zfill(2) +  '.' + str(firstdayPY.year) + ' - '  + str(lastdayPY.month).zfill(2) +  '.' + str(lastdayPY.year)
    PYPY_MM_YYYY = str(firstdayPYPY.month).zfill(2) +  '.' + str(firstdayPYPY.year) + ' - '  + str(lastdayPYPY.month).zfill(2) +  '.' + str(lastdayPYPY.year)
    PYPYPY_MM_YYYY = str(firstdayPYPYPY.month).zfill(2) +  '.' + str(firstdayPYPYPY.year) + ' - '  + str(lastdayPYPYPY.month).zfill(2) +  '.' + str(lastdayPYPYPY.year)


    PYTD_MMM_YYYY = str(firstdayPY.month).zfill(3) +  '.' + str(firstdayPY.year) + ' - '  + CM_MMM_YYYY
    PYPYTD_MMM_YYYY = str(firstdayPYPY.month).zfill(3) +  '.' + str(firstdayPYPY.year) + ' - '  + CM_MMM_YYYY
    PYPYPYTD_MMM_YYYY = str(firstdayPYPYPY.month).zfill(3) +  '.' + str(firstdayPYPYPY.year) + ' - '  + CM_MMM_YYYY
    PY_MMM_YYYY = str(firstdayPY.month).zfill(3) +  '.' + str(firstdayPY.year) + ' - '  + str(lastdayPY.month).zfill(3) +  '.' + str(lastdayPY.year)
    PYPY_MMM_YYYY = str(firstdayPYPY.month).zfill(3) +  '.' + str(firstdayPYPY.year) + ' - '  + str(lastdayPYPY.month).zfill(3) +  '.' + str(lastdayPYPY.year)


    PYPYPY_MMM_YYYY = str(firstdayPYPYPY.month).zfill(3) +  '.' + str(firstdayPYPYPY.year) + ' - '  + str(lastdayPYPYPY.month).zfill(3) +  '.' + str(lastdayPYPYPY.year)
    
    CY_JAN_MM_DD_YYYY = firstdayCY.strftime("%d.%m.%Y")
    CY_JAN_MM_YYYY = str(firstdayCY.month).zfill(2) +  '.' + str(firstdayCY.year)
    CY_JAN_MMM_YYYY = str(firstdayCY.month).zfill(3) +  '.' + str(firstdayCY.year)

    PY_JAN_MM_DD_YYYY = firstdayPY.strftime("%d.%m.%Y")
    PY_JAN_MM_YYYY = str(firstdayPY.month).zfill(2) +  '.' + str(firstdayPY.year)
    PY_JAN_MMM_YYYY = str(firstdayPY.month).zfill(3) +  '.' + str(firstdayPY.year)

    PYPY_JAN_MM_DD_YYYY = firstdayPYPY.strftime("%d.%m.%Y")
    PYPY_JAN_MM_YYYY = str(firstdayPYPY.month).zfill(2) +  '.' + str(firstdayPYPY.year)
    PYPY_JAN_MMM_YYYY = str(firstdayPYPY.month).zfill(3) +  '.' + str(firstdayPYPY.year)

    PYPYPY_JAN_MM_DD_YYYY = firstdayPYPYPY.strftime("%d.%m.%Y")
    PYPYPY_JAN_MM_YYYY = str(firstdayPYPYPY.month).zfill(2) +  '.' + str(firstdayPYPYPY.year)
    PYPYPY_JAN_MMM_YYYY = str(firstdayPYPYPY.month).zfill(3) +  '.' + str(firstdayPYPYPY.year)


    Key_date = (datetime(currentYear, currentMonth, 1) - relativedelta(days=1)).strftime("%d.%m.%Y")
    Key_date_PY = (datetime(lastdayPY.year, lastdayPY.month, lastdayPY.day)).strftime("%d.%m.%Y")
    PYTDLM_MM_DD_YYYY = str(firstdayPY.strftime("%d.%m.%Y")) + ' - ' + str(Key_date)


    append_values(values,'Key_Date_Last_Month', 'Key Date Last Month MM.DD.YYYY', Key_date)
    append_values(values,'Key_Date_Last_Year_End', 'Key Date Last Year End MM.DD.YYYY', Key_date_PY)
    append_values(values,'PY_MM_DD_YYYY', 'Past Year full Date Range MM.DD.YYYY', PY_MM_DD_YYYY)
    append_values(values,'PYTD_MM_DD_YYYY', 'Past Year to current Date Range MM.DD.YYYY', PYTD_MM_DD_YYYY)
    append_values(values,'PYTDLM_MM_DD_YYYY', 'Past Year to last month Range MM.DD.YYYY', PYTDLM_MM_DD_YYYY)
    append_values(values,'PYPY_MM_DD_YYYY', 'Year-2 full Date Range MM.DD.YYYY', PYPY_MM_DD_YYYY)
    append_values(values,'PYPYPY_MM_DD_YYYY', 'Year-3 full Date Range MM.DD.YYYY', PYPYPY_MM_DD_YYYY)
    append_values(values,'PYTD_MM_YYYY', 'Past Year to current Month Range MM.YYYY', PYTD_MM_YYYY)
    append_values(values,'PYPYTD_MM_YYYY', 'Past Year to current Month Range MM.YYYY', PYPYTD_MM_YYYY)
    append_values(values,'PYPYPYTD_MM_YYYY', 'Past Year to current Month Range MM.YYYY', PYPYPYTD_MM_YYYY)
    append_values(values,'PYTDLM_MM_YYYY', 'Past Year to last Month Range MM.YYYY', PYTDLM_MM_YYYY)
    append_values(values,'PYPYTDLM_MM_YYYY', 'Past Year to last Month Range MM.YYYY', PYPYTDLM_MM_YYYY)
    append_values(values,'PYPYPYTDLM_MM_YYYY', 'Past Year to last Month Range MM.YYYY', PYPYPYTDLM_MM_YYYY)
    append_values(values,'PY_MM_YYYY', 'Past Year full Month Range MM.YYYY', PY_MM_YYYY)
    append_values(values,'PYPY_MM_YYYY', 'Year-2 full Month Range MM.YYYY', PYPY_MM_YYYY)
    append_values(values,'PYPYPY_MM_YYYY', 'Year-3 full Month Range MM.YYYY', PYPYPY_MM_YYYY)
    append_values(values,'PYTD_MMM_YYYY', 'Past Year to current Month Range MMM.YYYY', PYTD_MMM_YYYY)
    append_values(values,'PYPYTD_MMM_YYYY', 'Past Year to current Month Range MMM.YYYY', PYPYTD_MMM_YYYY)
    append_values(values,'PYPYPYTD_MMM_YYYY', 'Past Year to current Month Range MMM.YYYY', PYPYPYTD_MMM_YYYY)
    append_values(values,'PY_MMM_YYYY', 'Past Year full Month Range MMM.YYYY', PY_MMM_YYYY)
    append_values(values,'PYPY_MMM_YYYY', 'Year-2 full Month Range MMM.YYYY', PYPY_MMM_YYYY)
    append_values(values,'PYPYPY_MMM_YYYY', 'Year-3 full Month Range MMM.YYYY', PYPYPY_MMM_YYYY)
  

    #YTD
    YTD_MM_DD_YYYY = str(firstdayCY.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    YTLM_MM_DD_YYYY = str(firstdayCY.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))

    CYTD_MM_YYYY = str(firstdayCY.month).zfill(2) +  '.' + str(firstdayCY.year) + ' - '  + CM_MM_YYYY
    CYTLM_MM_YYYY = str(firstdayCY.month).zfill(2) +  '.' + str(firstdayCY.year) + ' - '  + LM_MM_YYYY
    CY_MM_YYYY = str(firstdayCY.month).zfill(2) +  '.' + str(firstdayCY.year) + ' - '  + str(lastdayCY.month).zfill(2) +  '.' + str(lastdayCY.year)

    CYTD_MMM_YYYY = str(firstdayCY.month).zfill(3) +  '.' + str(firstdayCY.year) + ' - '  + CM_MMM_YYYY
    CY_MMM_YYYY = str(firstdayCY.month).zfill(3) +  '.' + str(firstdayCY.year) + ' - '  + str(lastdayCY.month).zfill(3) +  '.' + str(lastdayCY.year)
    CYTLM_MMM_YYYY = str(firstdayCY.month).zfill(3) +  '.' + str(firstdayCY.year) + ' - '  + LM_MMM_YYYY

    append_values(values,'YTD_MM_DD_YYYY', 'Current Year to today Range MM.DD.YYYY', YTD_MM_DD_YYYY)
    append_values(values,'YTLM_MM_DD_YYYY', 'Current Year to last month date Range MM.DD.YYYY', YTLM_MM_DD_YYYY)
    append_values(values,'CYTD_MM_YYYY', 'Current Year to current month Month Range MM.YYYY', CYTD_MM_YYYY)
    append_values(values,'CYTLM_MM_YYYY', 'Current Year to last month Month Range MM.YYYY', CYTLM_MM_YYYY)
    append_values(values,'CY_MM_YYYY', 'Current Year full Month Range MM.YYYY', CY_MM_YYYY)
    append_values(values,'CYTD_MMM_YYYY', 'Current Year to current month Month Range MMM.YYYY', CYTD_MMM_YYYY)
    append_values(values,'CYTLM_MMM_YYYY', 'Current Year to last month Month Range MMM.YYYY', CYTLM_MMM_YYYY)
    append_values(values,'CY_MMM_YYYY', 'Current Year full Month Range MMM.YYYY', CY_MMM_YYYY)

    
    append_values(values,'CY_JAN_MM_DD_YYYY', 'Current Year first day MM.DD.YYYY', CY_JAN_MM_DD_YYYY)
    append_values(values,'CY_JAN_MM_YYYY', 'Current Year January MM.YYYY', CY_JAN_MM_YYYY)
    append_values(values,'CY_JAN_MMM_YYYY', 'Current Year January MMM.YYYY', CY_JAN_MMM_YYYY)

    append_values(values,'PY_JAN_MM_DD_YYYY', 'Past Year first day MM.DD.YYYY', PY_JAN_MM_DD_YYYY)
    append_values(values,'PY_JAN_MM_YYYY', 'Past Year January MM.YYYY', PY_JAN_MM_YYYY)
    append_values(values,'PY_JAN_MMM_YYYY', 'Past Year January MMM.YYYY', PY_JAN_MMM_YYYY)

    append_values(values,'PYPY_JAN_MM_DD_YYYY', 'Year-2 first day MM.DD.YYYY', PYPY_JAN_MM_DD_YYYY)
    append_values(values,'PYPY_JAN_MM_YYYY', 'Year-2 January MM.YYYY', PYPY_JAN_MM_YYYY)
    append_values(values,'PYPY_JAN_MMM_YYYY', 'Year-2 January MMM.YYYY', PYPY_JAN_MMM_YYYY)

    append_values(values,'PYPYPY_JAN_MM_DD_YYYY', 'Year-3 first day MM.DD.YYYY', PYPYPY_JAN_MM_DD_YYYY)
    append_values(values,'PYPYPY_JAN_MM_YYYY', 'Year-3 January MM.YYYY', PYPYPY_JAN_MM_YYYY)
    append_values(values,'PYPYPY_JAN_MMM_YYYY', 'Year-3 January MMM.YYYY', PYPYPY_JAN_MMM_YYYY)


    #M-1
    Month_1 = (date.today() - relativedelta(months=1)).month
    Year_1 = (date.today() - relativedelta(months=1)).year
    Date_1  = datetime(Year_1, Month_1, 1)
    Last_Date_1 = ((datetime((date.today() - relativedelta(months=0)).year, (date.today() - relativedelta(months=0)).month, 1)) - relativedelta(days=1)).strftime("%d.%m.%Y")



    MM_YYYY_1 = str(Month_1).zfill(2) + '.' + str(Year_1)
    MMM_YYYY_1 = str(Month_1).zfill(3) + '.' + str(Year_1)
    

        

    #append_values(values,'MM_YYYY_1', 'Month -1 MM.YYYY', MM_YYYY_1)
    #append_values(values,'MMM_YYYY_1', 'Month -1 MMM.YYYY', MMM_YYYY_1)
    #append_values(values,'Date_1', 'Month -1 MM.DD.YYYY', Date_1)
    #append_values(values,'Date_1', 'Month -1 MM.DD.YYYY', Date_1)


    #M-4
    Month_4 = (date.today() - relativedelta(months=4)).month
    Year_4 = (date.today() - relativedelta(months=4)).year
    Date_4  = datetime(Year_4, Month_4, 1)
    Last_Date_4 = ((datetime((date.today() - relativedelta(months=3)).year, (date.today() - relativedelta(months=3)).month, 1)) - relativedelta(days=1)).strftime("%d.%m.%Y")


    MM_YYYY_4 = str(Month_4).zfill(2) + '.' + str(Year_4)
    MMM_YYYY_4= str(Month_4).zfill(3) + '.' + str(Year_4)


    #R4M
    R4M_today_MM_DD_YYYY = str(Date_4.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R4M_MM_DD_YYYY = str(Date_4.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R4M_MM_YYYY = str(Date_4.month).zfill(2) +  '.' + str(Date_4.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R4M_MMM_YYYY = str(Date_4.month).zfill(3) +  '.' + str(Date_4.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)


    #M-6
    Month_6 = (date.today() - relativedelta(months=6)).month
    Year_6 = (date.today() - relativedelta(months=6)).year
    Date_6  = datetime(Year_6, Month_6, 1)
    Last_Date_6 = ((datetime((date.today() - relativedelta(months=5)).year, (date.today() - relativedelta(months=5)).month, 1)) - relativedelta(days=1)).strftime("%d.%m.%Y")


    MM_YYYY_6 = str(Month_6).zfill(2) + '.' + str(Year_6)
    MMM_YYYY_6 = str(Month_6).zfill(3) + '.' + str(Year_6)


    #R6M
    R6M_today_MM_DD_YYYY = str(Date_6.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R6M_MM_DD_YYYY = str(Date_6.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R6M_MM_YYYY = str(Date_6.month).zfill(2) +  '.' + str(Date_6.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R6M_MMM_YYYY = str(Date_6.month).zfill(3) +  '.' + str(Date_6.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)

    R6M_CM_MM_YYYY = str(Date_6.month).zfill(2) +  '.' + str(Date_6.year) + ' - '  + str(currentMonth).zfill(2) +  '.' + str(currentYear)
    R6M_CM_MMM_YYYY = str(Date_6.month).zfill(3) +  '.' + str(Date_6.year) + ' - '  + str(currentMonth).zfill(3) +  '.' + str(currentYear)

    #M-12
    Month_12 = (date.today() - relativedelta(months=12)).month
    Year_12 = (date.today() - relativedelta(months=12)).year
    Date_12  = datetime(Year_12, Month_12, 1)
    Last_Date_12 = ((datetime((date.today() - relativedelta(months=11)).year, (date.today() - relativedelta(months=11)).month, 1)) - relativedelta(days=1))


    MM_YYYY_12 = str(Month_12).zfill(2) + '.' + str(Year_12)
    MMM_YYYY_12 = str(Month_12).zfill(3) + '.' + str(Year_12)

    #R12M
    R12M_today_MM_DD_YYYY = str(Date_12.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R12M_MM_DD_YYYY = str(Date_12.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R12M_MM_YYYY = str(Date_12.month).zfill(2) +  '.' + str(Date_12.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R12M_MMM_YYYY = str(Date_12.month).zfill(3) +  '.' + str(Date_12.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)



    #M-13
    Month_13 = (date.today() - relativedelta(months=13)).month
    Year_13 = (date.today() - relativedelta(months=13)).year
    Date_13  = datetime(Year_13, Month_13, 1)
    Last_Date_13 = ((datetime((date.today() - relativedelta(months=12)).year, (date.today() - relativedelta(months=12)).month, 1)) - relativedelta(days=1))


    MM_YYYY_13 = str(Month_13).zfill(2) + '.' + str(Year_13)
    MMM_YYYY_13 = str(Month_13).zfill(3) + '.' + str(Year_13)

    #R13M
    R13M_today_MM_DD_YYYY = str(Date_13.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R13M_MM_DD_YYYY = str(Date_13.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R13M_MM_YYYY = str(Date_13.month).zfill(2) +  '.' + str(Date_13.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R13M_MMM_YYYY = str(Date_13.month).zfill(3) +  '.' + str(Date_13.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)



    #M-18
    Month_18 = (date.today() - relativedelta(months=18)).month
    Year_18 = (date.today() - relativedelta(months=18)).year
    Date_18  = datetime(Year_18, Month_18, 1)
    Last_Date_18 = ((datetime((date.today() - relativedelta(months=17)).year, (date.today() - relativedelta(months=17)).month, 1)) - relativedelta(days=1))


    MM_YYYY_18 = str(Month_18).zfill(2) + '.' + str(Year_18)
    MMM_YYYY_18 = str(Month_18).zfill(3) + '.' + str(Year_18)

    #R18M
    R18M_today_MM_DD_YYYY = str(Date_18.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R18M_MM_DD_YYYY = str(Date_18.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R18M_MM_YYYY = str(Date_18.month).zfill(2) +  '.' + str(Date_18.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R18M_MMM_YYYY = str(Date_18.month).zfill(3) +  '.' + str(Date_18.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)



    #M-19
    Month_19 = (date.today() - relativedelta(months=19)).month
    Year_19 = (date.today() - relativedelta(months=19)).year
    Date_19  = datetime(Year_19, Month_19, 1)
    Last_Date_19 = ((datetime((date.today() - relativedelta(months=18)).year, (date.today() - relativedelta(months=18)).month, 1)) - relativedelta(days=1))


    MM_YYYY_19 = str(Month_19).zfill(2) + '.' + str(Year_19)
    MMM_YYYY_19 = str(Month_19).zfill(3) + '.' + str(Year_19)


    #R19M
    R19M_today_MM_DD_YYYY = str(Date_19.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R19M_MM_DD_YYYY = str(Date_19.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R19M_MM_YYYY = str(Date_19.month).zfill(2) +  '.' + str(Date_19.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R19M_MMM_YYYY = str(Date_19.month).zfill(3) +  '.' + str(Date_19.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)


    #M-24
    Month_24 = (date.today() - relativedelta(months=24)).month
    Year_24 = (date.today() - relativedelta(months=24)).year
    Date_24  = datetime(Year_24, Month_24, 1)
    Last_Date_24 = ((datetime((date.today() - relativedelta(months=23)).year, (date.today() - relativedelta(months=23)).month, 1)) - relativedelta(days=1))


    MM_YYYY_24 = str(Month_24).zfill(2) + '.' + str(Year_24)
    MMM_YYYY_24 = str(Month_24).zfill(3) + '.' + str(Year_24)

    #R24M
    R24M_today_MM_DD_YYYY = str(Date_24.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R24M_MM_DD_YYYY = str(Date_24.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R24M_MM_YYYY = str(Date_24.month).zfill(2) +  '.' + str(Date_24.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R24M_MMM_YYYY = str(Date_24.month).zfill(3) +  '.' + str(Date_24.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)


    #M-29
    Month_29 = (date.today() - relativedelta(months=29)).month
    Year_29 = (date.today() - relativedelta(months=29)).year
    Date_29  = datetime(Year_29, Month_29, 1)
    Last_Date_29 = ((datetime((date.today() - relativedelta(months=28)).year, (date.today() - relativedelta(months=28)).month, 1)) - relativedelta(days=1))


    MM_YYYY_29 = str(Month_29).zfill(2) + '.' + str(Year_29)
    MMM_YYYY_29 = str(Month_29).zfill(3) + '.' + str(Year_29)

    #R29M
    R29M_today_MM_DD_YYYY = str(Date_29.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R29M_MM_DD_YYYY = str(Date_29.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R29M_MM_YYYY = str(Date_29.month).zfill(2) +  '.' + str(Date_29.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R29M_MMM_YYYY = str(Date_29.month).zfill(3) +  '.' + str(Date_29.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)



    #M-36
    Month_36 = (date.today() - relativedelta(months=36)).month
    Year_36 = (date.today() - relativedelta(months=36)).year
    Date_36  = datetime(Year_36, Month_36, 1)
    Last_Date_36 = ((datetime((date.today() - relativedelta(months=35)).year, (date.today() - relativedelta(months=35)).month, 1)) - relativedelta(days=1))


    MM_YYYY_36 = str(Month_36).zfill(2) + '.' + str(Year_36)
    MMM_YYYY_36 = str(Month_36).zfill(3) + '.' + str(Year_36)

    #R36M
    R36M_today_MM_DD_YYYY = str(Date_36.strftime("%d.%m.%Y")) + ' - ' + str(currentDate.strftime("%d.%m.%Y"))
    R36M_MM_DD_YYYY = str(Date_36.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    R36M_MM_YYYY = str(Date_36.month).zfill(2) +  '.' + str(Date_36.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    R36M_MMM_YYYY = str(Date_36.month).zfill(3) +  '.' + str(Date_36.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)

    append_values(values,'R4M_today_MM_DD_YYYY', 'Rolling 4 months until today date range MM.DD.YYYY', R4M_today_MM_DD_YYYY)
    append_values(values,'R4M_MM_DD_YYYY', 'Rolling 4 months until last month date range MM.DD.YYYY', R4M_MM_DD_YYYY)
    append_values(values,'R4M_MM_YYYY', 'Rolling 4 months until last month month range MM.YYYY', R4M_MM_YYYY)
    append_values(values,'R4M_MMM_YYYY', 'Rolling 4 months until today date range MMM.YYYY', R4M_MMM_YYYY)

    append_values(values,'R6M_today_MM_DD_YYYY', 'Rolling 6 months until today date range MM.DD.YYYY', R6M_today_MM_DD_YYYY)
    append_values(values,'R6M_MM_DD_YYYY', 'Rolling 6 months until last month date range MM.DD.YYYY', R6M_MM_DD_YYYY)
    append_values(values,'R6M_MM_YYYY', 'Rolling 6 months until last month month range MM.YYYY', R6M_MM_YYYY)
    append_values(values,'R6M_MMM_YYYY', 'Rolling 6 months until today date range MMM.YYYY', R6M_MMM_YYYY)

    append_values(values,'R12M_today_MM_DD_YYYY', 'Rolling 12 months until today date range MM.DD.YYYY', R12M_today_MM_DD_YYYY)
    append_values(values,'R12M_MM_DD_YYYY', 'Rolling 12 months until last month date range MM.DD.YYYY', R12M_MM_DD_YYYY)
    append_values(values,'R12M_MM_YYYY', 'Rolling 12 months until last month month range MM.YYYY', R12M_MM_YYYY)
    append_values(values,'R12M_MMM_YYYY', 'Rolling 12 months until today date range MMM.YYYY', R12M_MMM_YYYY)

    append_values(values,'R13M_today_MM_DD_YYYY', 'Rolling 13 months until today date range MM.DD.YYYY', R13M_today_MM_DD_YYYY)
    append_values(values,'R13M_MM_DD_YYYY', 'Rolling 13 months until last month date range MM.DD.YYYY', R13M_MM_DD_YYYY)
    append_values(values,'R13M_MM_YYYY', 'Rolling 13 months until last month month range MM.YYYY', R13M_MM_YYYY)
    append_values(values,'R13M_MMM_YYYY', 'Rolling 13 months until today date range MMM.YYYY', R13M_MMM_YYYY)

    append_values(values,'R18M_today_MM_DD_YYYY', 'Rolling 18 months until today date range MM.DD.YYYY', R18M_today_MM_DD_YYYY)
    append_values(values,'R18M_MM_DD_YYYY', 'Rolling 18 months until last month date range MM.DD.YYYY', R18M_MM_DD_YYYY)
    append_values(values,'R18M_MM_YYYY', 'Rolling 18 months until last month month range MM.YYYY', R18M_MM_YYYY)
    append_values(values,'R18M_MMM_YYYY', 'Rolling 18 months until today date range MMM.YYYY', R18M_MMM_YYYY)

    append_values(values,'R19M_today_MM_DD_YYYY', 'Rolling 19 months until today date range MM.DD.YYYY', R19M_today_MM_DD_YYYY)
    append_values(values,'R19M_MM_DD_YYYY', 'Rolling 19 months until last month date range MM.DD.YYYY', R19M_MM_DD_YYYY)
    append_values(values,'R19M_MM_YYYY', 'Rolling 19 months until last month month range MM.YYYY', R19M_MM_YYYY)
    append_values(values,'R19M_MMM_YYYY', 'Rolling 19 months until today date range MMM.YYYY', R19M_MMM_YYYY)


    append_values(values,'R24M_today_MM_DD_YYYY', 'Rolling 24 months until today date range MM.DD.YYYY', R24M_today_MM_DD_YYYY)
    append_values(values,'R24M_MM_DD_YYYY', 'Rolling 24 months until last month date range MM.DD.YYYY', R24M_MM_DD_YYYY)
    append_values(values,'R24M_MM_YYYY', 'Rolling 24 months until last month month range MM.YYYY', R24M_MM_YYYY)
    append_values(values,'R24M_MMM_YYYY', 'Rolling 24 months until today date range MMM.YYYY', R24M_MMM_YYYY)

    append_values(values,'R29M_today_MM_DD_YYYY', 'Rolling 29 months until today date range MM.DD.YYYY', R29M_today_MM_DD_YYYY)
    append_values(values,'R29M_MM_DD_YYYY', 'Rolling 29 months until last month date range MM.DD.YYYY', R29M_MM_DD_YYYY)
    append_values(values,'R29M_MM_YYYY', 'Rolling 29 months until last month month range MM.YYYY', R29M_MM_YYYY)
    append_values(values,'R29M_MMM_YYYY', 'Rolling 29 months until today date range MMM.YYYY', R29M_MMM_YYYY)

    append_values(values,'R36M_today_MM_DD_YYYY', 'Rolling 36 months until today date range MM.DD.YYYY', R36M_today_MM_DD_YYYY)
    append_values(values,'R36M_MM_DD_YYYY', 'Rolling 36 months until last month date range MM.DD.YYYY', R36M_MM_DD_YYYY)
    append_values(values,'R36M_MM_YYYY', 'Rolling 36 months until last month month range MM.YYYY', R36M_MM_YYYY)
    append_values(values,'R36M_MMM_YYYY', 'Rolling 36 months until today date range MMM.YYYY', R36M_MMM_YYYY)
    #4,6,12,13,18,19,24,29,36

    #M+6
    Month_plus_6 = (date.today() + relativedelta(months=6)).month
    Year_plus_6 = (date.today() + relativedelta(months=6)).year
    Date_plus_6  = datetime(Year_plus_6, Month_plus_6, 1)
    Last_Date_plus_6 = ((datetime((date.today() + relativedelta(months=7)).year, (date.today() + relativedelta(months=7)).month, 1)) - relativedelta(days=1))

    MM_YYYY_plus_6= str(Month_plus_6).zfill(2) + '.' + str(Year_plus_6)
    MMM_YYYY__plus_6 = str(Month_plus_6).zfill(3) + '.' + str(Year_plus_6)

    #R+6M
    M6_MM_DD_YYYY = str(First_Day_CM.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_plus_6.strftime("%d.%m.%Y"))
    M6_MM_YYYY = str(First_Day_CM.month).zfill(2) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_6.month).zfill(2) +  '.' + str(Last_Date_plus_6.year)
    M6_MMM_YYYY = str(First_Day_CM.month).zfill(3) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_6.month).zfill(3) +  '.' + str(Last_Date_plus_6.year)



    #M+12
    Month_plus_12 = (date.today() + relativedelta(months=12)).month
    Year_plus_12 = (date.today() + relativedelta(months=12)).year
    Date_plus_12  = datetime(Year_plus_12, Month_plus_12, 1)
    Last_Date_plus_12 = ((datetime((date.today() + relativedelta(months=13)).year, (date.today() + relativedelta(months=13)).month, 1)) - relativedelta(days=1))



    MM_YYYY_plus_12 = str(Month_plus_12).zfill(2) + '.' + str(Year_plus_12)
    MMM_YYYY_plus_12 = str(Month_plus_12).zfill(3) + '.' + str(Year_plus_12)

    #R+12M
    M12_MM_DD_YYYY = str(First_Day_CM.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_plus_12.strftime("%d.%m.%Y"))
    M12_MM_YYYY = str(First_Day_CM.month).zfill(2) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_12.month).zfill(2) +  '.' + str(Last_Date_plus_12.year)
    M12_MMM_YYYY = str(First_Day_CM.month).zfill(3) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_12.month).zfill(3) +  '.' + str(Last_Date_plus_12.year)


    #M+13
    Month_plus_13 = (date.today() + relativedelta(months=13)).month
    Year_plus_13 = (date.today() + relativedelta(months=13)).year
    Date_plus_13  = datetime(Year_plus_13, Month_plus_13, 1)
    Last_Date_plus_13 = ((datetime((date.today() + relativedelta(months=14)).year, (date.today() + relativedelta(months=14)).month, 1)) - relativedelta(days=1))


    MM_YYYY_plus_13 = str(Month_plus_13).zfill(2) + '.' + str(Year_plus_13)
    MMM_YYYY_plus_13 = str(Month_plus_13).zfill(3) + '.' + str(Year_plus_13)

    #R+13M
    M13_MM_DD_YYYY = str(First_Day_CM.strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_plus_13.strftime("%d.%m.%Y"))
    M13_MM_YYYY = str(First_Day_CM.month).zfill(2) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_13.month).zfill(2) +  '.' + str(Last_Date_plus_13.year)
    M13_MMM_YYYY = str(First_Day_CM.month).zfill(3) +  '.' + str(First_Day_CM.year) + ' - '  + str(Last_Date_plus_13.month).zfill(3) +  '.' + str(Last_Date_plus_13.year)


    append_values(values,'M6_MM_DD_YYYY', 'Current month + 6 date range MM.DD.YYYY', M6_MM_DD_YYYY)
    append_values(values,'M6_MM_YYYY', 'Current month + 6 month range MM.YYYY', M6_MM_YYYY)
    append_values(values,'M6_MMM_YYYY', 'Current month + 6 month range MMM.YYYY', M6_MMM_YYYY)

    append_values(values,'M12_MM_DD_YYYY', 'Current month + 12 date range MM.DD.YYYY', M12_MM_DD_YYYY)
    append_values(values,'M12_MM_YYYY', 'Current month + 12 month range MM.YYYY', M12_MM_YYYY)
    append_values(values,'M12_MMM_YYYY', 'Current month + 12 month range MMM.YYYY', M12_MMM_YYYY)

    append_values(values,'M13_MM_DD_YYYY', 'Current month + 13 date range MM.DD.YYYY', M13_MM_DD_YYYY)
    append_values(values,'M13_MM_YYYY', 'Current month + 13 month range MM.YYYY', M13_MM_YYYY)
    append_values(values,'M13_MMM_YYYY', 'Current month + 13 month range MMM.YYYY', M13_MMM_YYYY)


    #Special Formats

    #08PYPYPY
    PYPYPY08_MM_DD_YYYY = str(datetime(firstdayPYPYPY.year,8,1).strftime("%d.%m.%Y")) + ' - ' + str(Last_Date_LM.strftime("%d.%m.%Y"))
    PYPYPY08_MM_YYYY = str(8).zfill(2) +  '.' + str(firstdayPYPYPY.year) + ' - '  + str(Last_Date_LM.month).zfill(2) +  '.' + str(Last_Date_LM.year)
    PYPYPY08_MMM_YYYY = str(8).zfill(3) +  '.' + str(firstdayPYPYPY.year) + ' - '  + str(Last_Date_LM.month).zfill(3) +  '.' + str(Last_Date_LM.year)

    #R29M from end of last year

    PY_Month_29 = (lastdayPY - relativedelta(months=28)).month
    PY_Year_29 = (lastdayPY - relativedelta(months=28)).year
    PY_Date_29  = datetime(PY_Year_29, PY_Month_29, 1)
    PY_Last_Date_29 = ((datetime((date.today() - relativedelta(months=28)).year, (date.today() - relativedelta(months=28)).month, 1)) - relativedelta(days=1))

    PY_R29M_MM_DD_YYYY = str(datetime(PY_Year_29,PY_Month_29,1).strftime("%d.%m.%Y")) + ' - ' + str(lastdayPY.strftime("%d.%m.%Y"))
    PY_R29M_MM_YYYY = str(PY_Month_29).zfill(2) +  '.' + str(PY_Year_29) + ' - '  + str(lastdayPY.month).zfill(2) +  '.' + str(lastdayPY.year)
    PY_R29M_MMM_YYYY = str(PY_Month_29).zfill(3) +  '.' + str(PY_Year_29) + ' - '  + str(lastdayPY.month).zfill(3) +  '.' + str(lastdayPY.year)



    append_values(values,'PYPYPY08_MM_DD_YYYY', 'Year - 3 (August) until last month date range MM.DD.YYYY', PYPYPY08_MM_DD_YYYY)
    append_values(values,'PYPYPY08_MM_YYYY', 'Year - 3 (August) until last month month range MM.YYYY', PYPYPY08_MM_YYYY)
    append_values(values,'PYPYPY08_MMM_YYYY', 'Year - 3 (August) until last month month range MMM.YYYY', PYPYPY08_MMM_YYYY)


    append_values(values,'PY_R29M_MM_DD_YYYY', 'Rolling 9 months from end of last year date range MM.DD.YYYY', PY_R29M_MM_DD_YYYY)
    append_values(values,'PY_R29M_MM_YYYY', 'Rolling 9 months from end of last year month range MM.YYYY', PY_R29M_MM_YYYY)
    append_values(values,'PY_R29M_MMM_YYYY', 'Rolling 9 months from end of last year month range MMM.YYYY', PY_R29M_MMM_YYYY)

    currentMonth_first_day =datetime(currentYear,currentMonth,1)

    Y_1 = (date.today().year) +1
    Y_2 = (date.today().year) +2

    NY_M01 = datetime(Y_1, 1, 1)
    NYNY_CM = datetime(Y_2,currentMonth,1)
    
    NY01_NYNYCM_MM_DD_YYYY = str(NY_M01.strftime("%d.%m.%Y")) + ' - ' + str(NYNY_CM.strftime("%d.%m.%Y"))
    NY01_NYNYCM_MM_YYYY = str(1).zfill(2) +  '.' + str(Y_1) + ' - '  + str(currentMonth).zfill(2) +  '.' + str(Y_2) 
    NY01_NYNYCM_MMM_YYYY = str(1).zfill(3) +  '.' + str(Y_1) + ' - '  + str(currentMonth).zfill(3) +  '.' + str(Y_2) 

    append_values(values,'NY01_NYNYCM_MM_DD_YYYY', 'Year + 1 (January) until current month Y+2 date range MM.DD.YYYY', NY01_NYNYCM_MM_DD_YYYY)
    append_values(values,'NY01_NYNYCM_MM_YYYY', 'Year + 1 (January) until current month Y+2 month range MM.YYYY', NY01_NYNYCM_MM_YYYY)
    append_values(values,'NY01_NYNYCM_MMM_YYYY', 'Year + 1 (January) until current month Y+2 month range MMM.YYYY', NY01_NYNYCM_MMM_YYYY)
    
    #Set formats
    keydate = (currentMonth_first_day - relativedelta(days=1)).strftime("%d.%m.%Y")
    currentDate = currentDate.strftime("%d.%m.%Y")
    currentMonth_first_day = currentMonth_first_day.strftime("%d.%m.%Y")



    #month_values = dict()
    month_values =[]


    for i in range(-48, 49):
            Month_loop = (date.today() + relativedelta(months=i)).month
            Year_loop  = (date.today() + relativedelta(months=i)).year

            MM_YYYY_loop = str(Month_loop).zfill(2) + '.' + str(Year_loop)
            MMM_YYYY_loop = str(Month_loop).zfill(3) + '.' + str(Year_loop)
            Date_loop   = datetime(Year_loop , Month_loop , 1)


            Next_date = (Date_loop + relativedelta(months=1))
            Last_Date_loop  = Next_date - relativedelta(days=1) #((datetime((Date_loop.date()).year , (Date_loop.date() + relativedelta(months=1)).month, 1)) - relativedelta(days=1))
            append_month_values(month_values,'M'+str(i),MM_YYYY_loop,MMM_YYYY_loop,Date_loop.strftime("%d.%m.%Y"),Last_Date_loop.strftime("%d.%m.%Y"))
            if i < 0:
                prefix = '_minus_'+str(abs(i))
            else:
                prefix = '_plus_'+str(abs(i))
            append_values(values,'M'+prefix +'_MM_YYYY','M'+str(i) +' MM.YYYY',MM_YYYY_loop)
            append_values(values,'M'+prefix +'_MMM_YYYY','M'+str(i) +' MMM.YYYY',MMM_YYYY_loop)
            append_values(values,'M'+prefix +'_FirstDate','M'+str(i) +' FirstDate',Date_loop.strftime("%d.%m.%Y"))
            append_values(values,'M'+prefix +'_LastDate','M'+str(i) +' LastDate',Last_Date_loop.strftime("%d.%m.%Y"))





    #values['PY_R29M_MM_DD_YYYY'] = PY_R29M_MM_DD_YYYY
    #values['current_period'] = date.today()
    #delta = relativedelta(months=-1)
    #values['previous_period'] = values['current_period'] + delta
    #values['year_current_period'] = values['current_period'].year
    #values['year_previous_period'] = values['previous_period'].year
    #values['range_current_month'] = '{} - {}'.format(values['current_period'].month, values['current_period'].month)
    #values['range_previous_month'] = '{} - {}'.format(values['previous_period'].month, values['previous_period'].month)
    #values['key_date'] = Key_date
    #values=[]

    #values.append(
    #    {
    #        'Name': 'SAPSetVariable',
    #        'Value': PY_R29M_MM_DD_YYYY
    #    }
    #)
 
    
    df_month_values = pd.DataFrame.from_dict(month_values)
    df_time_values = pd.DataFrame.from_dict(values)
    return df_month_values

def _get_wb_path(filename):
    """Path of excel file to import"""
    workbook_filepath = Path.cwd().joinpath('Workbooks/'+filename)
    return workbook_filepath



def search_directory(data_directory):
    """search the directory for excel files to be refreshed"""
    onlyfiles = [f for f in listdir(data_directory) if isfile(join(data_directory, f))]
    list_files = []
    for filename in onlyfiles:
        if filename[-4:].lower() == 'xlsx':
            list_files.append(filename)
    return list_files





@retry(reraise=True, wait=wait_fixed(10), before_sleep=before_sleep_log(logger, logging.DEBUG), stop=(stop_after_delay(300) | stop_after_attempt(3)))
@timeit
def sap_logon(xl_Instance, source, client, user, password, language):
    """API method to trigger a logon to a system for a specified data source"""
    result = xl_Instance.Application.Run("SAPLogon", source, client, user, password, language)
    if result == 1:
        print('\nSuccessfully logged in SAP AfO')
    else:
        raise ConnectionError("Couldn't login in SAP AfO")

    return result

def sap_logon_check(xl_Instance, source, client, user, password, language):
    """API method to trigger a logon to a system for a specified data source"""
    result = xl_Instance.Application.Run("SAPLogon", source, client, user, password, language)
    if result == 1:
        print('\nSuccessfully logged in SAP AfO')
    return result


@timeit
def sap_refresh(xl_Instance):
    """
    Do there initial refresh of data in the workbook.
    All data sources and planning objects will be refreshed.
    If you execute this command for a data source which is already refreshed, all corresponding crosstabs are redrawn.
    """
    result = xl_Instance.Application.Run("SAPExecuteCommand", "Refresh")
    if result == 1:
        print('\nSuccessfully refreshed the workbook')
    else:
        raise ConnectionError("Couldn't refresh the SAP AfO")  # todo: need to handle errors differently to not stop the runtime
    return result


@timeit
def sap_refresh_data(xl_Instance, source):
    """
    Refresh the transaction data for all or defined data sources in the workbook.
    The corresponding transaction data is updated from the server and the crosstabs are redrawn.
    """
    print("Start Refresh " + source)
    result = xl_Instance.Application.Run("SAPExecuteCommand", "RefreshData", source)
    if result == 1:
        print(f'\nSuccessfully refreshed the source: {source}')
    else:
        raise ConnectionError(f"Couldn't refresh the the source: {source}")

    print("Refreshed " + source)
    return result




def sap_message_display(xl_Instance):
    global messageList, curr_boa_message_id, BOA_Err_Message, current_source, boa_message_severity, boa_message_text, curr_boa_message_id, file_target, prev_msg

    #SAPListOfMessages will return tuples containing error informations, if no message occurs an empty string is returned. 
    # It is possible that multiple messages are returned, in this case first element of the tuple is another tuple

    messageList = xl_Instance.Application.Run("SAPListOfMessages",'',"True")
    boa_message_count = 0

    try:
        #If result is no tuple then no message was returned
        if not isinstance(messageList, tuple):
            return 0,0

        #Check for multiple messages - if first element is tuple
        if isinstance(messageList[0], tuple) and messageList != "":
            number_of_boa_messages = len(messageList)
            #Loop messages
            for i in range(0,number_of_boa_messages):
                curr_msg = messageList[i]
                #Only proceed if error is new (Message ID > then previous ID or Message ID is 0 and messages are not equal)
                if int(curr_msg[0]) > int(curr_boa_message_id) or (curr_msg[0] == 0 and prev_msg != curr_msg):
                    boa_message_count = boa_message_count +1
                    prev_msg = curr_msg
                    curr_boa_message_id = curr_msg[0]
                    boa_message_text = curr_msg[1]
                    boa_message_severity = curr_msg[4]
                    #Only take into account errors & critical warnings, rest is logged
                    if boa_message_severity == "ERROR" or boa_message_severity == "CRITICAL":
                        BOA_Err_Message = True
                        logger.error('BOA Error for ' + str(current_source) + ' in ' + str(file_target) + '  '  + str(boa_message_severity) + ' : ' +str(boa_message_text))
                        boa_message_err_text = str(boa_message_severity) + ': ' + str(boa_message_text)
                        #Concat messages if more than one error message is returned 
                        if boa_message_count > 1:
                            boa_message_text = str(boa_message_err_text) + ' ' + str(boa_message_text)
                        else:
                            boa_message_text = str(boa_message_severity) + ': ' + str(boa_message_text)  

                    else:
                        logger.info('BOA Message for ' + str(current_source) + ' in ' + str(file_target) + '  ' + str(boa_message_severity) + ' : ' +str(boa_message_text)) 
                else:
                    return 0, 0
        # 1 Message returned
        else:
            if messageList != "":
                #Only proceed if error is new
                if int(messageList[0]) > int(curr_boa_message_id):
                    curr_boa_message_id = messageList[0]
                    boa_message_text = messageList[1]
                    boa_message_severity = messageList[4]
                    #Only take into account errors & critical warnings, rest is logged
                    if boa_message_severity == "ERROR" or boa_message_severity == "CRITICAL":
                        BOA_Err_Message = True
                        logger.error('BOA Error for ' + str(current_source) + ' in ' + str(file_target) + '  ' + str(boa_message_severity) + ' : ' +str(boa_message_text))

                        boa_message_text = str(boa_message_severity) + ': ' + str(boa_message_text)  
                    else:
                        logger.info('BOA Message for ' + str(current_source) +  ' in ' + str(file_target) + '  ' + str(boa_message_severity) + ' : ' +str(boa_message_text))
                else:
                    return 0, 0
            else:
                return 0, 0


    except Exception as e:
        print('Error reading SAP messages: ' +str(e))
        return 0,0
    
    
    return boa_message_severity, boa_message_text


def sap_get_more_info(xl_Instance, data_values):
    """get other information about SAP data source"""
    data_values['DataSourceName'] = xl_Instance.Application.Run("SapGetSourceInfo", data_values['DS'], "DataSourceName")
    data_values['Query'] = xl_Instance.Application.Run("SapGetSourceInfo", data_values['DS'], "QueryTechName")
    data_values['System'] = xl_Instance.Application.Run("SapGetSourceInfo", data_values['DS'], "System")
    return data_values


def sap_get_variables(xl_Instance, source):
    """Get the list of variables of a data source"""
    variables_list = xl_Instance.Application.Run("SAPListOfVariables", source, "INPUT_STRING", "ALL")
    return variables_list


def sap_get_technical_name(xl_Instance, source, variable_name):
    """returns the value of the technical name for a specific variable"""
    var_tech_name = xl_Instance.Application.Run("SAPGetVariable", source, variable_name, "TECHNICALNAME")
    return var_tech_name


def sap_get_filters(xl_Instance, source):
    """Get the list of filters (measures) of a data source"""
    filters_list = xl_Instance.Application.Run("SAPListOfDynamicFilters", source, "INPUT_STRING")
    return filters_list


def sap_get_dimensions(xl_Instance, source):
    """Get the list of dimensions (fields) of a data source"""
    dimensions_list = xl_Instance.Application.Run("SAPListOfDimensions", source)
    return dimensions_list


def sap_is_ds_active(xl_Instance, source):
    """check whether a data source is active"""
    state_data_source = xl_Instance.Application.Run("SAPGetProperty", "IsDataSourceActive", source)
    return state_data_source


def sap_last_error(xl_Instance):
    """return last errot"""
    boa_last_err = xl_Instance.Application.Run("SAPGetProperty", "LastError","Number")
    return boa_last_err

def sap_is_connected(xl_Instance, source):
    """check whether a data source is already connected"""
    state_connection = xl_Instance.Application.Run("SAPGetProperty", "IsConnected", source)
    return state_connection


def last_update_input(xl_Instance, source):
    global last_update
    last_update = xl_Instance.Application.Run("SapGetSourceInfo", source, "LastDataUpdate")
    return last_update

@timeit
def logoff(xl_Instance):


    global refresh_running, Monitoring_Windows
    try:
        #Start Thread to monitor for BOA pop-up
        refresh_running = True
        thread_monitor_windows = Thread(target=search_boa_message_window)
        thread_monitor_windows.start()

        xl_Instance.Application.Run("SAPLogOff", False)
    except:
        pass
    finally:
        #Stop thread after logoff
        Monitoring_Windows = False
        thread_monitor_windows.join()
        refresh_running = False


def logon_input(xl_Instance, source):

    global client,user, password, language
    """
    Logon into the SAP AfO System. The logon is datasource dependent. 
    """
    # assign variables
    client = global_configs.query('Setting=="logon-client"')['Value'].values[0]
    if password_found == False:
        user = global_configs.query('Setting=="logon-user"')['Value'].values[0]
        password = global_configs.query('Setting=="logon-password"')['Value'].values[0]
    language = global_configs.query('Setting=="logon-language"')['Value'].values[0]

    if source is not None:
        source = source
    # execute the logon method
    is_logged = sap_logon(xl_Instance, source, client, user, password, language)



def cleanup_mei(time_threshold = 86400): # Default setting: Remove after 1 day, time_threshold in seconds
    """
    Rudimentary workaround for https://github.com/pyinstaller/pyinstaller/issues/2379
    """

    mei_bundle = getattr(sys, "_MEIPASS", False)
    if mei_bundle:
        dir_mei, current_mei = mei_bundle.split("_MEI")
        for file in os.listdir(dir_mei):
            if file.startswith("_MEI") and not file.endswith(current_mei):
                if (time.time()-os.path.getctime(file)) > time_threshold:
                    try:
                        rmtree(os.path.join(dir_mei, file))
                    except PermissionError:  # mainly to allow simultaneous pyinstaller instances
                        pass
    else:
        #Clear all files from temp older than a day
        dir_tmp = tempfile.gettempdir() + os.sep
        for file in os.listdir(dir_tmp):
            if file.startswith("_MEI"):
                if (time.time()-os.path.getctime(os.path.join(dir_tmp, file))) > time_threshold:
                    try:
                        rmtree(os.path.join(dir_tmp, file))
                    except PermissionError:  # mainly to allow simultaneous pyinstaller instances
                        pass

def dispatch(app_name:str):
    try:
        app = win32.gencache.EnsureDispatch(app_name)
    except AttributeError:
        
        # https://gist.github.com/rdapaz/63590adb94a46039ca4a10994dff9dbe?permalink_comment_id=2918299#gistcomment-2918299
        # Corner case dependencies.
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        
        app = win32.gencache.EnsureDispatch(app_name)
    return app


def check_if_DS_exists(ds_to_check, instance):
    global ds_exists, clear_excel_sessions, SapReport, Err_Count
    ds_exists = False
    count = 0
    addin_is_enabled = False

    try:
        ds_list_check = instance.Application.Run("SAPListOf", "CROSSTABS")
    except:
        #Main reason for errors is that the analysis plugin is not running
        logger.info(f"Listing Crosstabs failed, Excel restarted " +  str(file_target))
        kill_excel_instances()
        
        SapReport = SapRefresh()
        SapReport.ExcelInstance = open_excel()
        # open SAP AfO report

        try:
            SapReport.open_report(file_target)
            addin_is_enabled = SapReport.enable_boa_addin(file_target)
            ds_list_check = instance.Application.Run("SAPListOf", "CROSSTABS")
        except:
            clear_excel_sessions_init = clear_excel_sessions
            try:
                clear_excel_sessions = True   
                clear_excel_sessions = clear_excel_sessions_init
                logger.info(f"Listing Crosstabs failed, Re-initiating Report " +  str(file_target))
                del SapReport
                SapReport = SapRefresh()
                SapReport.ExcelInstance = open_excel()
                SapReport.open_report(file_target)
                addin_is_enabled = SapReport.enable_boa_addin(file_target)
                ds_list_check = instance.Application.Run("SAPListOf", "CROSSTABS")
                #logger.info(f"Report opened " +  str(file_target))
            except:
                logger.info(f"Re-initiating Report failed " +  str(file_target))
                Err_Count = Err_Count +1
                return 2 #could not open report
        finally:
            if addin_is_enabled == False:
                Err_Count = Err_Count +1
                logger.critical(f"Could not activate Analysis plugin in " +  str(FilePath))
                return 1 #could not enable analysis
    
    #Check if only 1 query is in the file (if more queries present, multiple tuples will be returned)
    for item in ds_list_check:
        if isinstance(item, tuple):
            pass
        else:
            if count == 0: 
                act_ct = item
            
            elif count == 1: 
                act_ctt = item
            elif count == 2:    
                act_ds = item  
                if str(act_ds) == str(ds_to_check):
                    ds_exists = True
           
        count = count + 1

    for ds in ds_list_check:
        if isinstance(ds, tuple):
            act_ds = ds[2]
            act_ct = ds[0]
            act_ctt = ds[1]
            if str(act_ds) == str(ds_to_check):
                ds_exists = True
    
    return ds_exists 



class SapRefresh:
    """Class that handles all the events related to the SapRefresh application"""
    global ExcelInstance

    def __init__(self, config_df=global_configs_df):
        # global configuration dataframe
        self.global_configs = config_df
        # file related parameter. Need to be set only when a workbook is loaded
        self.data_source = None  # dictionary with all infos about the data source
        self.source = None  # the name of the data source that was loaded (usually DS_1)
        self.filepath = None  # the filepath of the SAP AfO file
        self.IsSharePointPath = None
        self.IsWebPath = None
        self.SavedTempPath = None
        self.SavedLocally = False
        # Classes of Excel API
        self.ExcelInstance = None
        self.WorkbookSAP = None
        # parameters set when workbook is opened
        self.calc_state_init = None
        # state parameters
        self.is_logged = None
        self.is_refreshed = None
        self.is_refreshed_data = None

        self.state_refresh_behavior = None
        self.state_variable_submit = None
        # List of all variables and filters in the data source
        self.variables_filters = None


    def connect_opened_report(self):

        global xl_Instance
        
        if check_if_Excel_runs() == False:
            self.ExcelInstance  = xl_Instance = open_excel()
        else:
            if 'None' in str(self.ExcelInstance):
                self.ExcelInstance = xl_Instance = return_running_excel()
           


    def open_report(self, filepath):
        global open_wb, ExcelApp
        #"""
        #Initiate Excel instance. Configure it to optimize the execution. Then open the target workbook.
        #Capture the initial calculation state. Activate workbook.
        #"""      
        
        # kill every excel instance so that no further erros can happen, try first without excel to be closed
        if clear_excel_sessions == True:
            try:
                self.ExcelInstance = open_excel()
            except:
                kill_excel_instances()
                self.ExcelInstance = open_excel()
            
        else:
            if check_if_Excel_runs() == False:
                self.ExcelInstance = open_excel()
            else:
                if ExcelApp.xlApp == None or 'None' in str(ExcelApp.xlApp):
                    del ExcelApp
                    ExcelApp = Excel()
                self.ExcelInstance = xl_Instance = ExcelApp.xlApp #= return_running_excel()
        
        # Make it visible otherwise it doesn’t work
        if self.ExcelInstance.Visible == False:
            self.ExcelInstance.Visible = True

        if self.ExcelInstance.DisplayAlerts == True:
            self.ExcelInstance.DisplayAlerts = False
        

        #Check if file is already open
        Split_path(filepath)
        if check_if_excel_file_is_open(split_fullname):
            try:
                self.WorkbookSAP, self.ExcelInstance   = self.ExcelInstance.Workbooks(split_fullname)
            except:
                #Try to resolve path in case of network drive
                Split_path(pathlib.Path(filepath).resolve())
                self.WorkbookSAP, self.ExcelInstance   = self.ExcelInstance.Workbooks(split_fullname)
        else:
            self.WorkbookSAP, self.ExcelInstance = open_workbook(self.ExcelInstance, filepath) #xl_Instance

        open_wb = self.WorkbookSAP

        
        self.calc_state_init = calculation_state(self.ExcelInstance, 'start')
        
        #Activate Workbook
        #If configuration file is to be created (else) use path of selected file
        if update_mode == 'Refresh':
            # assign the filepath of the SAP AfO file
            self.filepath = filepath
            if type(filepath) is pathlib.WindowsPath:
                ensure_wb_active(self.ExcelInstance, filepath.name)
            else:
                ensure_wb_active(self.ExcelInstance, current_filename) #xl_Instance
        else:
            # assign the filepath of the SAP AfO file
            self.filepath = current_filepath
            if type(filepath) is pathlib.WindowsPath:
                ensure_wb_active(self.ExcelInstance, filepath.name)
            else:
                ensure_wb_active(self.ExcelInstance, str(split_nameonly + split_extension)) #xl_Instance
        
        #Check if path is SharePoint
        if 'sharepoint' in str(self.filepath).lower():
            self.IsSharePointPath = True
        else:
            self.IsSharePointPath = False

        #Check if path is other Web Path
        if 'http' in str(self.filepath).lower():
            self.IsWebPath = True
        else:
            self.IsWebPath = False

        print('The report is loaded')




    def enable_boa_addin(self, target_file):
        global Err_Count
        
        try:

            #check if addin_enabled, by calling last error from BOA plugin. Error is thrown from sap_last_error when the plugin is enabled but not working (e.g. not visible in Excel tool pane but showing as enabled)
            try:
                check_addin = sap_last_error(self.ExcelInstance)
                if check_addin == -2146826252:
                    print('Addin not running, trying to enable Addin')
                    ensure_addin(self.ExcelInstance) 
                    return True
                else:
                    print('Addin is running')
                    return True
            except:
                #De-activate and re-activate adding
                ensure_addin(self.ExcelInstance) 
                return True
        except Exception as e:
            print('Addin not running, closing Excel and re-opening report. Error: '+str(e))
            kill_excel_instances()
            try:
                # try to check if plugin is disabled, on error try to enable addin via registry in enable_addins()
                result_addin = enable_addins(self.ExcelInstance)
                if result_addin:
                    print("Re-enabling Analysis Addin succeeded or no deactivated Add-ins were found")
                else:
                    logger.info(f"Re-enabling Analysis Addin failed")

            except Exception as e1:
                logger.info(f"Possible error due to disabled Analysis Plugin: " + str(e1))
                
                
            #Re-open report
            self.open_report(target_file)
            
            try:
                #check if addin_enabled, by calling last error from BOA plugin. Error is thrown from sap_last_error when the plugin is enabled but not working (e.g. not visible in Excel tool pane but showing as enabled)
                try:
                    check_addin = sap_last_error(self.ExcelInstance)
                    if check_addin == -2146826252:
                        print('Addin not running, trying to enable')
                        ensure_addin(self.ExcelInstance) 
                        return True
                    else:
                        print('Addin is running')
                        return True
                except:
                    #De-activate and re-activate adding
                    ensure_addin(self.ExcelInstance) 
                    check_addin = sap_last_error(self.ExcelInstance)
                    if check_addin == -2146826252:
                        print('Addin not running, trying to enable')
                        return False
                    else:
                        print('Addin is running')
                        return True
            except Exception as e2:
                Err_Count = Err_Count +1
                logger.error(f"Error activating Analysis Plugin: " + str(e2))
                return False

    def calculate(self):
            """Calculate workbook to refresh values"""
            self.ExcelInstance.Application.Calculate()



    def logon(self, source): #(self, source=None):
        global client,user, password, language, Err_Count, default_system
        """
        Logon into the SAP AfO System. The logon is file dependent. That's because you need to refer the
        data source to connect to SAP. It uses the source of the get_data_source method.
        """
        #Logon - if empty show pop-up window
        #If configuration file is to be created (else) no logon data is saved - prompt will appear to login.  Will not be used when configuration is to be created from files in a folder

        if update_mode == 'Refresh' or update_mode == "Create Configuration multiple" and file_source_config == True:
            get_sys_pw(current_sys)
            #if system found in df_pws by function get_sys_pw
            if password_found == False:
                if user_config == True:
                    user = self.global_configs.query('Setting=="logon-user"')['Value'].values[0]
                else:
                    user = ''
                if password_config == True:
                    password = self.global_configs.query('Setting=="logon-password"')['Value'].values[0]
                else: 
                    password = ''
            
                if client_config == True:
                    client = self.global_configs.query('Setting=="logon-client"')['Value'].values[0]
                else:
                    client = 100
                if language_config == True:
                    language = self.global_configs.query('Setting=="logon-language"')['Value'].values[0]
                    #language must be 2 characters
                    if len(language)>2:
                        language = 'en'
                else:
                    language = 'en'
                if (user_config == False or password_config == False) and start_mode == "Scheduled":
                    Err_Count = Err_Count +1
                    raise CredentialsException('Logon credentials were not found, refresh stopped. Make sure password files are availabe and/or settings in Configuration File are maintained correctly.')
                    #logger.critical('Logon credentials were not found, refresh stopped. Make sure password files are availabe and/or settings in Configuration File are maintained correctly.')

        else:
            client = 100
            get_sys_pw(default_system)
            if password_found == False:
                user = ''
                password = ''
            else:
                user = user
                password = password
            language = 'en'
        

        
        # assign variables
        if source is not None:
            self.source = source
        # execute the logon method
        self.is_logged = sap_logon(self.ExcelInstance, self.source, client, user, password, language)


    def refresh(self):
        """Do there initial refresh of data in the workbook."""
        self.is_refreshed = sap_refresh(self.ExcelInstance)

    def refresh_data(self, source):
        """Refresh the transaction data for all data sources in the workbook."""
        self.is_refreshed_data = sap_refresh_data(self.ExcelInstance, source)

    def additional_source_info(self):
        """
        Query more information and append it to the data source dictionary.
        Attention. This method is dependent of Data Source initiation.
        """
        self.data_source = sap_get_more_info(self.ExcelInstance, self.data_source)
        print('Additional data source information retrieved', '\n', self.data_source)
    
    @timeit
    def close_workbook(self):
        """Save the file and close it"""


        #self.WorkbookSAP.Save()
        time.sleep(5)
        self.WorkbookSAP.Close()


    def refesh_all(self):
        global Refresh_All_Waiting_Time
        if Run_Refresh_All_Excel:
            try:
                self.WorkbookSAP.RefreshAll()
                time.sleep(Refresh_All_Waiting_Time)
            except:
                print('Could not run Refresh All in current Excel Workbook')
            #self.WorkbookSAP.CalculateUntilAsyncQueriesDone()
    
    
    @timeit
    def Save_workbook(self):
        self.WorkbookSAP.Save()
        self.SavedLocally = False
        #Sharepoint files are saved locally in temp folder - otherwise the save as CSV which is performed after all queries refreshed would not work 
        if self.IsSharePointPath == True or self.IsWebPath == True:
            save_name = str(self.WorkbookSAP.Name)
            save_pth = os.path.join(path_tmp,save_name)
            self.SavedTempPath = save_pth
            try:
                self.WorkbookSAP.SaveAs(Filename=save_pth)
                time.sleep(5)
                self.WorkbookSAP  = self.ExcelInstance.Workbooks(str(self.WorkbookSAP.Name))
                self.SavedLocally = True
                temp_paths.append(
                    {
                        'Filename': str(self.WorkbookSAP.Name),
                        'TempPath': path_tmp,
                        'TempFilepath': save_pth
                    }
                )
                time.sleep(60)
            except:
                self.SavedLocally = False
                logger.warning(f"Couldn't save Web/SharePoint file locally")

    @timeit
    def Save_copy_as(self, target):
        self.WorkbookSAP.SaveCopyAs(target)



    def close(self):
        #"""
        #Make all the necessary procedures to terminate the excel instance.
        #    - put the  calculation back to the original state
        #    - terminate the workbook instance
        #    - Configure the Excel Instance to the original state
        #    - Close the Excel Instance
        #"""
        calculation_state(self.ExcelInstance, 'stop', self.calc_state_init)
        self.close_workbook(self.WorkbookSAP)
        optimize_instance(self.ExcelInstance, 'stop')

        self.ExcelInstance.Application.Quit()
        self.WorkbookSAP = None
        self.ExcelInstance = None
        print('The application was Successfully closed')

    def get_variables_list(self):
        """Return a dictionary of the variables that exists in the data source"""
        variables_list = sap_get_variables(self.ExcelInstance, self.source)
        return variables_list

    def variables_filters_list(self):
        """Return a dataframe with all the variables and filters inside the datasource"""
        # get the list of variables
        variables_list = sap_get_variables(self.ExcelInstance, self.source)
        # get technical name of the variables and append to Restrictions list
        restrictions = []
        for variable in variables_list:
            restrictions.append(
                {
                    'command': 'SAPSetVariable',
                    'field': sap_get_technical_name(self.ExcelInstance, self.source, variable[0]),
                    'field_name': variable[0],
                    'value': variable[1]
                }
            )
        # get the list of filters (measures)
        filters_list = sap_get_filters(self.ExcelInstance, self.source)
        # get list of dimensions (fields)
        dimensions_list = sap_get_dimensions(self.ExcelInstance, self.source)
        # search in dimensions the technical name of each filter then append values to Restrictions list
        for filter_ in filters_list:
            if filter_[0] != 'Measures':
                values = dict()
                values['command'] = 'SAPSetFilter'
                for dimension in dimensions_list:
                    if dimension[1] == filter_[0]:
                        values['field'] = dimension[0]  # get the technical name
                values['field_name'] = filter_[0]
                values['value'] = filter_[1]
                restrictions.append(values)
        # create the dataframe with filters and variables
        # noinspection PyTypeChecker
        variables_filters = pd.DataFrame.from_dict(restrictions)
        variables_filters['data_source'] = self.source
        variables_filters['reference_type'] = 'value'
        variables_filters['data_source_name'] = self.data_source['DataSourceName']
        variables_filters['data_source_sheet'] = self.data_source['Sheet']
        # assign values to properties
        self.variables_filters = variables_filters
        return variables_filters

    def data_source_list(self):
        """Return a Df with data source information"""
        data_source = pd.DataFrame(list(self.data_source.items()), columns=['Key', 'Value'])
        return data_source


   
    def export_variables_filters(self):
        """export to an Excel file the data source information and variables and filters values"""
        # load information from class' properties
        filepath = self.filepath
        path_data_info = pathlib.Path(self.global_configs.query('Setting=="path-data_info"')['value'].values[0])
        # create the pathname
        name_file = filepath.name[:len(filepath.suffix) * -1]
        complement = '__information'
        file_extension = filepath.suffix
        new_name = pathlib.Path(name_file + complement + file_extension)
        new_filepath = path_data_info / new_name
        # assign dataframes
        data_source_info = self.data_source_list()
        variables_filters_info = self.variables_filters_list()
        # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pd.ExcelWriter(new_filepath, engine='xlsxwriter')
        # write the dataframes
        data_source_info.to_excel(writer, sheet_name='data_source_info')
        variables_filters_info.to_excel(writer, sheet_name='variables_filters_info')
        # Close the Pandas Excel writer and output the Excel file.
        writer.close()

    def is_ds_active(self):
        """check whether a data source is active"""
        state_data_source = sap_is_ds_active(self.ExcelInstance, self.source)
        return state_data_source

    def is_connected(self):
        """check whether a data source is already connected"""
        state_connection = sap_is_connected(self.ExcelInstance, self.source)
        return state_connection
    

    def set_refresh_variables(self, variables_list):
        print('Starting to set the variables:')
        allow_zero_numeric_values = False
        self.state_refresh_behavior = self.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
        self.state_variable_submit = self.ExcelInstance.Application.Run("SAPExecuteCommand", "PauseVariableSubmit", "On")
        for index, variable in variables_list.iterrows():
            print('\n\t', f'Trying to set [{variable.Field}] to [{variable.Value}] ', end='')
            # Zero numeric values might negatively impact system performance, therefore zero values are excluded by default
            if variable.Value == 0 or variable.Value == '0':
                if allow_zero_numeric_values:
                    self.ExcelInstance.Application.Run(
                    variable.Command,
                    variable.Field,
                    variable.Value,
                    "INPUT_STRING",

                    variable.Datasource)
                    print('Ok! Zero value set!')
                else:
                    print('Warning! Zero value not set!')      
            else:
                self.ExcelInstance.Application.Run(
                    variable.Command,
                    variable.Field,
                    variable.Value,
                    "INPUT_STRING",
                    variable.Datasource)
                print('Ok!')

        # start waiting spinner
        spinner = Halo(text='Loading', spinner='dots')
        spinner.start()
        
        # refresh data with new variable values
        self.state_variable_submit = self.ExcelInstance.Application.Run("SAPExecuteCommand", "PauseVariableSubmit", "Off")
        #self.state_refresh_behavior = self.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
        # stop waiting spinner
        spinner.succeed('End!')
        print('The variables were set properly')


    def set_refresh_filters(self, df_filters):
        print('Starting to set the filters:')
        self.state_refresh_behavior = self.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
        for index, filter_item in df_filters.iterrows():
            print('\n\t', f'Trying to set [{filter_item.Field}] to [{filter_item.Value}] ', end='')
            self.ExcelInstance.Application.Run(
                filter_item.Command,
                filter_item.Datasource,
                filter_item.Field,
                filter_item.Value,
                "INPUT_STRING"
            )

            print('Ok!')
        # start waiting spinner
        spinner = Halo(text='Loading', spinner='dots')
        spinner.start()
        
        # refresh data with new variable values
        #self.state_refresh_behavior = self.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
        # stop waiting spinner
        spinner.succeed('End!')
        print('The filters were set properly')

@timeit
def open_workbook_target(xl_Instance, path):
    global wb_target

    """
    Open the file in the new Excel instance,

        The 1st false: don't update the links
        The 2nd false: and don't open it read-only
    """
    wb_target = xl_Instance.Workbooks.Open(path, False, False)
    time.sleep(2)

    return wb_target

def implement_BOA_macro_to_Excel(wb_BOA):
    #insert macro

    #wb_BOA = xl.Workbooks.Add()

    global Callback_BeforeMessageDisplay_Exists,Workbook_SAP_Initialize_Exists, Func_GetAsTwoDimArray_Exists, tmp_Module_Exists, inital_content_thiswb, inital_lines
    Callback_BeforeMessageDisplay_Exists= False
    Workbook_SAP_Initialize_Exists= False
    Func_GetAsTwoDimArray_Exists= False
    tmp_Module_Exists = False


    #Get ThisWorkbook Codemodule - not accessed by name as it can be different due to languagues being used
    ThisWorkbook_CodeModule = wb_BOA.VBProject.VBComponents(wb_BOA.CodeName).CodeModule
    ThisWorkbook_CodeModule_Name = wb_BOA.VBProject.VBComponents(wb_BOA.CodeName).Name

    #WorkbookName
    FileName_CodeModule = wb_BOA.Name

    #Look for already exising temp module
    for i in wb_BOA.VBProject.VBComponents:
        if i.Name == "tmp_module_BOA":
            tmp_Module_Exists = True

    if tmp_Module_Exists == False:
        #Add new module tmp_module_BOA
        xlmodule = wb_BOA.VBProject.VBComponents.Add(1)
        xlmodule.Name = 'tmp_module_BOA'

    #Check if macros already availabe
    for i in wb_BOA.VBProject.VBComponents:
        if i.Name == "ThisWorkbook" or i.Name == "tmp_module_BOA":
            num_lines = i.CodeModule.CountOfLines
            for j in range(1, num_lines+1):
                if 'Sub' in i.CodeModule.Lines(j, 1) and not 'End Sub' in i.CodeModule.Lines(j, 1):
                    if i.CodeModule.Lines(j, 1) == "Public Sub Callback_BeforeMessageDisplay()":
                        Callback_BeforeMessageDisplay_Exists = True
                    if i.CodeModule.Lines(j, 1) == "Public Sub Workbook_SAP_Initialize()":
                        Workbook_SAP_Initialize_Exists = True
                if 'Function' in i.CodeModule.Lines(j, 1) and not 'End Function' in i.CodeModule.Lines(j, 1):
                    if i.CodeModule.Lines(j, 1) == "Function GetAsTwoDimArray(value As Variant) As Variant":
                        Func_GetAsTwoDimArray_Exists = True


    #Select ThisWorkbook component
    VBModule = ThisWorkbook_CodeModule #wb_BOA.VBProject.VBComponents.Item("ThisWorkbook").CodeModule


    #Get inital contents of ThisWorkbook Module
    inital_lines = VBModule.CountOfLines
    # If empty skip
    if inital_lines == 0:
        inital_content_thiswb =""
    else:
        inital_content_thiswb = VBModule.Lines(1,inital_lines)


    #Save Callback macro to ThisWorkbook
    code = '''Public Sub Workbook_SAP_Initialize()
    ' register callbacks
    Call Application.Run("SAPExecuteCommand", "RegisterCallback", "BeforeMessageDisplay", "Callback_BeforeMessageDisplay")
    End Sub'''
    if Callback_BeforeMessageDisplay_Exists == False:
        VBModule.AddFromString(code)



    #Select tmp_module_BOA component
    VBModule = wb_BOA.VBProject.VBComponents.Item("tmp_module_BOA").CodeModule

    #Save Before Message display macro to tmp_module_BOA
    code = '''Public Sub Callback_BeforeMessageDisplay()
    Dim messageList As Variant
    Dim messages As Variant
    Dim lRet As Variant
    Dim messageCount As Variant
    Dim i As Integer

    'Declare variables/objects first
    Dim fso As Object, oFile As Object
    Dim oShell As Object, FileName As String

    'Turn "%TEMP%" into a real file path
    On Error Resume Next
    Set oShell = CreateObject("WScript.Shell")
    FileName = oShell.ExpandEnvironmentStrings("%Temp%\BOA_Messages.txt")
    Set oShell = Nothing 'Tidy up the Objects we no longer need



    'This bit creates the file
    Set fso = CreateObject("Scripting.FileSystemObject")
        If VBA.Dir(FileName) = "" Then
            Set oFile = fso.CreateTextFile(FileName)
            File_Exists = True
        Else
            Open FileName For Append As #1
            File_Exists = False
        End If

    On Error GoTo 0

    
    messageList = Application.Run("SAPListOfMessages", , "True")
    messages = GetAsTwoDimArray(messageList) ' see "Using Analysis functions"
    messageCount = UBound(messages, 1)
    
    For i = 1 To messageCount
    'If messages(i, 5) = "INFORMATION" Then
        lRet = Application.Run("SAPSuppressMessage", messages(i, 1))
        'End If
        'Write Message to txt File
        On Error Resume Next
        If File_Exists = True Then
            oFile.WriteLine messages(i, 5) & " - " & messages(i, 2) & " - " & ThisWorkbook.Name & " - " & Now()
            oFile.Close
        Else
            Write #1, messages(i, 5) & " - " & messages(i, 2) & " - " & ThisWorkbook.Name & " - " & Now()
            Close #1
        End If
        On Error GoTo 0
    Next i
    On Error Resume Next
    Set oFile = Nothing 'Tidy up the Objects we no longer need
    Set fso = Nothing 'Tidy up the Objects we no longer need
    On Error GoTo 0
    End Sub

    Function GetAsTwoDimArray(value As Variant) As Variant
    'value is error return the error
    If IsError(value) Then
    GetAsTwoDimArray = value

    'value is array
    ElseIf IsArray(value) Then
    'first check if the array is two-dimensional
    'by requesting the upper bound of the 2nd dimension.
    'if this is not the case an error occurs (Err.Number <> 0).

    'ignore errors, handled locally
    On Error Resume Next
    Dim lIndex As Integer
    Dim lErrorCode As Integer
    lIndex = UBound(value, 2)
    lErrorCode = Err.Number
    'set error handling back to default
    On Error GoTo 0

    If lErrorCode = 0 Then
    'no error: array is two-dimensional
    GetAsTwoDimArray = value
    Else
    'copy one-dimensional array into a two-dimensional one
    Dim i As Integer
    Dim lArray() As Variant
    ReDim lArray(1 To 1, 1 To UBound(value))
    For i = 1 To UBound(lArray, 2)
    lArray(1, i) = value(i)
    Next
    GetAsTwoDimArray = lArray
    End If

    Else
    'return Empty
    GetAsTwoDimArray = Empty

    End If

    End Function

    Sub DeleteModule_tmp()
        'Dim VBProj As VBIDE.VBProject
        'Dim VBComp As VBIDE.VBComponent
        On Error Resume Next
        Set VBProj = ActiveWorkbook.VBProject
        Set VBComp = VBProj.VBComponents("tmp_module_BOA")
        VBProj.VBComponents.Remove VBComp
        On Error Goto 0
    End Sub
    '''
    if Workbook_SAP_Initialize_Exists == False and Func_GetAsTwoDimArray_Exists == False:
        VBModule.AddFromString(code)
    else:
        print("tmp_module_BOA not created, exists already")

    try:
        macro =  FileName_CodeModule + '!'+ str(ThisWorkbook_CodeModule_Name + '.Workbook_SAP_Initialize')
        wb_BOA.Application.Run(macro)
    except:
        print("Couldnt initialize BOA Macro for Message suppression")
    print("BOA Macro for Message suppression initialized")


def delete_BOA_macro_from_Excel(wb_BOA):
    #Delete contents from ThisWorkbook or restore inital contents if 
    #Select ThisWorkbook component
    #Get ThisWorkbook Codemodule - not accessed by name as it can be different due to languagues being used
    ThisWorkbook_CodeModule = wb_BOA.VBProject.VBComponents(wb_BOA.CodeName).CodeModule
    ThisWorkbook_CodeModule_Name = wb_BOA.VBProject.VBComponents(wb_BOA.CodeName).Name
    VBModule = wb_BOA.VBProject.VBComponents.Item(ThisWorkbook_CodeModule_Name).CodeModule


    #Get inital contents of ThisWorkbook Module
    lines = VBModule.CountOfLines

    # If empty skip
    if lines == 0:
        pass
    else:
        VBModule.DeleteLines(1,lines)
    # If empty skip
    if inital_lines == 0:
        pass
    else:
        VBModule.AddFromString(inital_content_thiswb)


    try:
        wb_BOA.Application.Run('tmp_module_BOA.DeleteModule_tmp')

    except:
        print("tmp_module_BOA not deleted, probably already deleted")
        pass


def get_list_of_DS():

    global act_ds, act_ct,act_ctt, reset_lists,user,password, filepath_selected, filenamefull_selected, Err_Count, settings_sheet_available, default_system, default_client
    query_list = []
    reset_lists = True
    err_get_crosstabs = False
    
#try:
    try:
        ds_list = SapReportCollect.ExcelInstance.Application.Run("SAPListOf", "CROSSTABS")
    except Exception as e:
        #Try to re-initate report
        try:
            print('Could not run macro SAPListof Crosstabs, re-initiating workbook to refresh')
            initiate_report_to_refresh(filepath_selected, filenamefull_selected)
            time.sleep(2)
            ds_list = SapReportCollect.ExcelInstance.Application.Run("SAPListOf", "CROSSTABS")
        except Exception as e:
            logger.error('Could not run macro SAPListof Crosstabs : '+ str(e) )
            err_get_crosstabs = True
            #Dummy ds_list to continue
            ds_list = ['ErrorListingCrosstabs', 'ErrorListingCrosstabs', 'DS_X']


    #if err_get_crosstabs == False:
    count = 0
    single_query = False

    #Check if only 1 query is in the file (if more queries present, multiple tuples will be returned)
    for item in ds_list:
        #Tuple = multiple queries
        if isinstance(item, tuple):
            pass
        else:
            if count == 0: 
                act_ct = item
            elif count == 1: 
                act_ctt = item
            elif count == 2:    
                act_ds = item  
            single_query = True
        count = count + 1


    # assign variables
    #If configuration file is to be created (else) no logon data is saved - prompt will appear to login.  Will not be used when configuration is to be created from files in a folder
    if update_mode == 'Refresh' or update_mode == "Create Configuration multiple" and file_source_config == True and settings_sheet_available == True:
        client = global_configs.query('Setting=="logon-client"')['Value'].values[0]
        if user_config:
            user = global_configs.query('Setting=="logon-user"')['Value'].values[0]
        else:
            get_sys_pw(default_system)
        if password_config:
            password = global_configs.query('Setting=="logon-password"')['Value'].values[0]
        else:
            get_sys_pw(default_system)
        language = global_configs.query('Setting=="logon-language"')['Value'].values[0]
    else:
        client = default_client
        get_sys_pw(default_system)

        #Reset user / pw if not found, prompt will open
        if password_found == False:
            user = ''
            password = ''
        else:
            user = user
            password = password
        language = 'en'
    count = 0
    
    

    print('Datasources found:  '+str(ds_list))
    #Loop all datasources in file
    for ds in ds_list:
        count = count +1
        if isinstance(ds, tuple):
            act_ds = ds[2]
            act_ct = ds[0]
            act_ctt = ds[1]

        #Assign current DS to SapReportCollect, needed in is_ds_active()
        SapReportCollect.source = act_ds
        #If DS is not active, logon - skip if error with inital crosstab extract. 
        if SapReportCollect.is_ds_active() == False and err_get_crosstabs == False:
            # Will not be used when configuration is to be created from files in a folder
            if update_mode == 'Refresh' or update_mode == "Create Configuration multiple" and file_source_config == True:
                sap_logon(SapReportCollect.ExcelInstance, act_ds, client, user, password, language)
            else:
                if inital_password_path != "":
                    for syst, usr, pw in zip(df_passw['System'], df_passw['User'], df_passw['Password']):
                        #Variable prompt will open if user / pw not given
                        logon_check = sap_logon_check(SapReportCollect.ExcelInstance, act_ds, client, usr, pw, language)
                        #Will always return 0 if logged in manually, only continues if login is completed
                        if logon_check == 1:
                            break
                    if logon_check != 1:
                       sap_logon(SapReportCollect.ExcelInstance, act_ds, client, user, password, language) 
                else:
                    sap_logon(SapReportCollect.ExcelInstance, act_ds, client, user, password, language)
    
            src = act_ds

            #Remove Filename from Filepath
            if str(wb_name) in str(FilePath):
                FilePath_str = str(FilePath).replace(str(wb_name),'')
            else:
                FilePath_str = str(FilePath)
            
            #Refresh all DS in file
            time.sleep(0.5)        
            try:
                #Do not refresh with inital errors on crosstab extraction
                if err_get_crosstabs == False:
                    result = SapReportCollect.ExcelInstance.Application.Run("SAPExecuteCommand", "Refresh", SapReportCollect.source)
                else:
                    result = 0
            except Exception as e:
                logger.error('Could not run macro Refresh Queries : '+ str(e) )
                result = 0
            
            time.sleep(0.5)   
            if result == 1:

                print('\nSuccessfully refreshed '+ str(SapReportCollect.source))
            else:
                print('Resetting Datasource')
                #logoff - logon - refresh to reset datasource
                logger.info(f"Resetting Datasource " + str(act_ds))
                #result = SapReportCollect.ExcelInstance.Application.Run("SAPExecuteCommand", "Restart", SapReportCollect.source)
                err_logoff = False
                err_logon = False

                try:
                    logoff(xl_Instance)
                except:
                    err_logoff = True
                try:
                    sap_logon(SapReportCollect.ExcelInstance, act_ds, client, user, password, language)
                except:
                    err_logon = True
                    logger.error('Could not logon to : '+ str(act_ds) )
                try:
                    result = SapReportCollect.ExcelInstance.Application.Run("SAPExecuteCommand", "Refresh", SapReportCollect.source)
                except:
                    logger.CRITICAL("Couldn´t load variables & filters for " + str(act_ds) + " remove and replace values for this query in configuration file. Error: " + str(Exception))
                    Err_Count = Err_Count +1

                    # Create entries manually, error refreshing query, this query will be skipped, entries must be corrected manually. Ensures that configuration file can be completed when errors occur
                    query_list_values = dict()
                    query_list_values['Filename'] = wb_name
                    query_list_values['Filepath'] = FilePath_str
                    query_list_values['Fullpath'] = FilePath
                    query_list_values['Sheet'] = ''
                    query_list_values['Datasource'] = src
                    if err_logon:
                        query_list_values['Query name'] = 'Error logging on, remove and replace values for this query'
                    else:
                        query_list_values['Query name'] = 'Error refreshing queries, remove and replace values for this query'
                    if err_get_crosstabs:
                        query_list_values['Query name'] = 'Error refreshing queries, file has been skipped'
                    query_list_values['Query technical name'] = ''
                    query_list_values['System'] = ''
                    query_list_values['Crosstab'] = act_ct
                    query_list_values['Crosstab name'] = act_ctt
                    query_list_values['Refresh'] = ''
                    query_list_values['Refresh on Workday'] = ''
                    query_list_values['Refresh on Local Time'] = ''
                    query_list_values['Last refreshed'] = ''
                    query_list_values['Save as CSV Fullpath'] = ''
                    query_list.append(query_list_values)
                    df_query_list = pd.DataFrame.from_dict(query_list)
                    restrictions= [{'Command': 'ErrorLoadingVariablesFilters', 'Field': '', 'Field name': 'Error loading variables & filters, remove and replace values for this query', 'Value': ''}]
                    variables_filters = pd.DataFrame.from_dict(restrictions)
                    variables_filters['Filename'] = wb_name
                    variables_filters['Datasource'] = src
                    variables_filters['Query name'] = ''
                    variables_filters['Query technical name'] = ''
                    variables_filters['System'] = ''
                    variables_filters['Crosstab'] = act_ct
                    variables_filters['Crosstab name'] = act_ctt
                    variables_filters['Filepath'] = FilePath_str
                    variables_filters['Fullpath'] = FilePath   
                    variables_filters['Sheet'] = ''
                    variables_filters['Variable to use as Value'] = ''
                    variables_filters = variables_filters[['Filename', 'Filepath', 'Fullpath', 'Sheet','Datasource','Command','Field','Field name','Value','Query name','Query technical name', 'System','Crosstab','Crosstab name','Variable to use as Value']]
                    
                    if count > 1:
                        df_variables_filters = pd.concat([variables_filters, df_variables_filters])
                    elif count <= 1:
                        df_variables_filters = variables_filters.copy()
                    if single_query == True:
                        print('Exit loop - single query')
                        break    
                    continue
            



            #Create Query List input
            query_list_values = dict()
            
            query_list_values['Filename'] = wb_name
            query_list_values['Filepath'] = FilePath_str
            query_list_values['Fullpath'] = FilePath
            query_list_values['Sheet'] = SapReportCollect.ExcelInstance.ActiveWorkbook.Names("SAP" + act_ct).RefersToRange.Parent.Name
            query_list_values['Datasource'] = src
            query_list_values['Query name'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "DataSourceName")
            query_list_values['Query technical name'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "QueryTechName")
            query_list_values['System'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "System")
            query_list_values['Crosstab'] = act_ct
            query_list_values['Crosstab name'] = act_ctt
            query_list_values['Refresh'] = ''
            query_list_values['Refresh on Workday'] = 99
            query_list_values['Refresh on Local Time'] = ''
            query_list_values['Last refreshed'] = ''
            query_list_values['Save as CSV Fullpath'] = FilePath_str + str(SapReportCollect.ExcelInstance.ActiveWorkbook.Names("SAP" + act_ct).RefersToRange.Parent.Name) + '_' + str(SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "QueryTechName")) + '.csv'
            

            # create the dataframe with query list
            query_list.append(query_list_values)
            queries = pd.DataFrame.from_dict(query_list)
            


            df_query_list = queries.copy()


            # get the list of variables
            try:
                variables_list = SapReportCollect.get_variables_list()
            except:
                variables_list = 0
            """Return a dataframe with all the variables and filters inside the datasource"""
            # get technical name of the variables and append to Restrictions list
            restrictions = []

            #If variable list returns integer, error occured - usually SAP Analysis Plugin fails
            if type(variables_list) == int:
                print('Resetting Datasource')
                #logoff - logon - refresh to reset datasource
                logger.info(f"Resetting Datasource " + str(act_ds))
                #result = SapReportCollect.ExcelInstance.Application.Run("SAPExecuteCommand", "Restart", SapReportCollect.source)
                logoff(xl_Instance)
                sap_logon(SapReportCollect.ExcelInstance, act_ds, client, user, password, language)
                result = SapReportCollect.ExcelInstance.Application.Run("SAPExecuteCommand", "Refresh", SapReportCollect.source)
                try:
                    variables_list = SapReportCollect.get_variables_list()
                except:
                    variables_list = 0
                

            if type(variables_list) != int:
                for variable in variables_list:
                    restrictions.append(
                        {
                            'Command': 'SAPSetVariable',
                            'Field': sap_get_technical_name(SapReportCollect.ExcelInstance, SapReportCollect.source, variable[0]),
                            'Field name': variable[0],
                            'Value': variable[1]

                        }
                    )

            else:
                    #Create empty entries on error
                    restrictions.append(
                        {
                            'Command': 'NoVariable',
                            'Field': '',
                            'Field name': '',
                            'Value': ''

                        }
                    )
            try:
                err_filter_dimensions = False
                # get the list of filters (measures)
                filters_list = sap_get_filters(SapReportCollect.ExcelInstance, SapReportCollect.source)
                # get list of dimensions (fields)
                dimensions_list = sap_get_dimensions(SapReportCollect.ExcelInstance, SapReportCollect.source)
            except:
                err_filter_dimensions = True
            
            #Create  entries only when no error
            if err_filter_dimensions == False:
            

                for filter_ in filters_list:

                    #Initialize values
                    values = dict()
                    values['Command'] = 'SAPSetFilter' 

                    #Check if only 1 filter is applied (if more filters are applied, tuples will be returned for each filter)
                    if isinstance(filter_, tuple):  
                        values['Value'] = filter_[1]
                        filter_ = filter_[0]
                    else:
                        values['Value'] = filter_

                    #Get dimension description and technical name (=Filter) from dimension list 
                    for dimension in dimensions_list:
                        if dimension[1] == filter_:
                            techn = dimension[0]
                            name= dimension[1]#filter_[0]

                    values['Field'] =techn  # get the technical name
                    values['Field name'] = name
                    
                    #Only append if value does not equal fieldname otherwise measures will create 2 rows
                    if values['Value'] != values['Field name']:
                        restrictions.append(values)

            # create the dataframe with filters and variables
            variables_filters = pd.DataFrame.from_dict(restrictions)

            #Remove Filename from Filepath
            if str(wb_name) in str(FilePath):
                FilePath_str = str(FilePath).replace(str(wb_name),'')
            else:
                FilePath_str = str(FilePath)

            variables_filters['Filename'] = wb_name
            variables_filters['Datasource'] = src
            #Error indicates issues with gen_py folder, using dispatch Excel to delete
            try:
                variables_filters['Query name'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "DataSourceName") #src['DataSourceName']
                variables_filters['Query technical name'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "QueryTechName")
                variables_filters['System'] = SapReportCollect.ExcelInstance.Application.Run("SapGetSourceInfo", src, "System")
            except:
                variables_filters['Query name'] = ''
                variables_filters['Query technical name'] = ''
                variables_filters['System'] = ''           
            variables_filters['Crosstab'] = act_ct
            variables_filters['Crosstab name'] = act_ctt
            variables_filters['Filepath'] = FilePath_str
            variables_filters['Fullpath'] = FilePath   
            try:
                variables_filters['Sheet'] = SapReportCollect.ExcelInstance.ActiveWorkbook.Names("SAP" + act_ct).RefersToRange.Parent.Name
            except:
                variables_filters['Sheet'] = ''
            variables_filters['Variable to use as Value'] = ''

            #Concat if more than 1 query

            if count > 1:
                #df_query_list = pd.concat([queries, df_query_list])
                df_variables_filters = pd.concat([variables_filters, df_variables_filters])
            elif count <= 1:
                #df_query_list = queries.copy()
                df_variables_filters = variables_filters.copy()


        if single_query == True:
            print('Exit loop - single query')
            break            
    
   

    #Adjust order 

    #query_list = query_list[['Filename', 'Filepath', 'Sheet','Datasource','Query name','Query technical name', 'System','Crosstab','Crosstab name', 'Refresh','Refresh on Workday' ,'Refresh on Local Time','Last refreshed']]
    df_variables_filters = df_variables_filters[['Filename', 'Filepath', 'Fullpath', 'Sheet','Datasource','Command','Field','Field name','Value','Query name','Query technical name', 'System','Crosstab','Crosstab name','Variable to use as Value']]
#except:
#    print('Err')
#    pass
    

    


    return df_query_list, df_variables_filters


def initiate_report_to_refresh(fullpath, filename):

    global wb_to_collect, FilePath, SapReportCollect, current_filepath, ExcelApp, clear_excel_sessions, Err_Count
    # initiate workbook
    SapReportCollect = SapRefresh()
    

    FilePath = fullpath
    current_filepath = FilePath
    addin_is_enabled = False
    
        

    #Check if file is already open
    if check_if_Excel_runs() == False:
        SapReportCollect.ExcelInstance = open_excel()
    else:
        if ExcelApp.xlApp == None or 'None' in str(ExcelApp.xlApp):
            del ExcelApp
            ExcelApp = Excel()
        SapReportCollect.ExcelInstance = xl_Instance = ExcelApp.xlApp 

    Split_path(FilePath)

    
    if check_if_excel_file_is_open(split_fullname):
        #SapReportCollect.ExcelInstance = xl_Instance
        SapReportCollect.WorkbookSAP = SapReportCollect.ExcelInstance.Workbooks(split_fullname)
        SapReportCollect.calc_state_init = calculation_state(SapReportCollect.ExcelInstance, 'start')
    else:
        #SapReportCollect.open_report(FilePath)
        # open SAP AfO report
        try:
            SapReportCollect.open_report(FilePath)
            addin_is_enabled = SapReportCollect.enable_boa_addin(FilePath)
        except:
            clear_excel_sessions_init = clear_excel_sessions
            try:
                clear_excel_sessions = True   
                clear_excel_sessions = clear_excel_sessions_init
                logger.info(f"Opening report for inital refresh failed, Re-initiating Report " +  str(FilePath))
                del SapReportCollect
                SapReportCollect = SapRefresh()
                SapReportCollect.open_report(FilePath)
                addin_is_enabled = SapReportCollect.enable_boa_addin(FilePath)
            except:
                Err_Count = Err_Count +1
                return 2 #could not open report
        finally:
            if addin_is_enabled == False:
                Err_Count = Err_Count +1
                logger.critical(f"Could not activate Analysis plugin in " +  str(FilePath))
                return 1 #could not enable analysis


    # Make it visible otherwise it doesn’t work

    #if SapReportCollect.ExcelInstance.Visible == False:
    #    SapReportCollect.ExcelInstance.Visible = True
    #if SapReportCollect.ExcelInstance.DisplayAlerts == True:
    #    SapReportCollect.ExcelInstance.DisplayAlerts = False
    #Start Analysis Addin - already done in open_report
    #ensure_addin(SapReportCollect.ExcelInstance)



    try:
        calc_state_init = calculation_state(SapReportCollect.ExcelInstance, 'start')
    except:
        pass
        
    if type(FilePath) is pathlib.WindowsPath:
        ensure_wb_active(SapReportCollect.ExcelInstance, FilePath.name)
    else:
        ensure_wb_active(SapReportCollect.ExcelInstance, filename)

    return 0

def refresh_report_err_handler(wb, save_file):
    try:
        #Make sure macro is removed
        if Supress_All_BOA_Messages:
            if BOA_macro_implemented:
                if BOA_macro_removed == False:
                    delete_BOA_macro_from_Excel(wb)
    except:
        pass
    
    try:
        #Save file
        if save_file:
            Save_workbook_input(wb)
    except:
        pass

def capture_runtime(filepath,sheet,ds,cs,query,type,start,end):
    global run_times
    try:
        run_times.append(
            {
                'Filename': str(filepath),
                'Sheet': str(sheet),
                'Datasource': str(ds),
                'Crosstab': str(cs),
                'Query': str(query),
                'Type': str(type),
                'Start time': str(start),
                'End time': str(end),
                'Run time': str(end-start)

            }
            )
    except:
        print('Error capturing runtimes')

@timeit
def refresh_report(filename, data_sources, variables_filters):
    """execute the flow necessary to refresh de desired report"""
    global SapReport, xl_Instance, clear_excel_sessions, file_target,refresh_running, BOA_macro_removed, BOA_macro_implemented, no_message_window_found, Monitoring_Windows, Err_Count, BOA_Err_Message, current_source, run_times, run_time_start, run_time_end, capture_runtimes, current_sheet, current_crosstab, current_query_technical_name, result_val
    #variables_filters.fillna('', inplace=True)
    BOA_macro_implemented =False
    BOA_macro_removed = False
    #BOA_message_closed = False
    addin_is_enabled = False
    #Set true if a error message is found in sap_message_display()
    BOA_Err_Message = False
    #Query result if only one cell is returned
    result_val = ''

    # configure path - if error and entry doesnt exist use file path
    #If configuration file is to be created (else) use path of selected file    

    if update_mode == 'Refresh':
        try:
            file_target = current_filepath
        except:
            file_target = current_filepath
    else:
        file_target = current_filepath#filepath_selected
    
    # collect data sources
    current_source = data_sources
    # replace dynamic values in the parameters
    dict_time_values = get_time_intelligence()
    print(current_source)

    #Check if file has to be opened / closed - defined in refresh_auto_reports()
    if Open_file_Check == True:
        # initiate workbook
        SapReport = SapRefresh()

        # open SAP AfO report

        try:
            SapReport.open_report(file_target)
            addin_is_enabled = SapReport.enable_boa_addin(file_target)
            if Supress_All_BOA_Messages:
                try:
                    implement_BOA_macro_to_Excel(SapReport.WorkbookSAP)
                    BOA_macro_implemented = True
                except:
                    logger.info('Could not implement macros to supress BOA messages: '+ str(Exception) )
                finally:
                    pass
        except e as Exception:
            clear_excel_sessions_init = clear_excel_sessions
            try:
                clear_excel_sessions = True   
                clear_excel_sessions = clear_excel_sessions_init
                logger.info(f"Opening report failed, Re-initiating Report " +  str(file_target))
                print(e)
                logger.info(str(e))
                del SapReport
                SapReport = SapRefresh()
                SapReport.open_report(file_target)
                addin_is_enabled = SapReport.enable_boa_addin(file_target)
                if Supress_All_BOA_Messages:
                    try:
                        implement_BOA_macro_to_Excel(SapReport.WorkbookSAP)
                        BOA_macro_implemented = True
                    except:
                        logger.info('Could not implement macros to supress BOA messages: '+ str(Exception) )
                    finally:
                        pass
            except:
                Err_Count = Err_Count +1
                refresh_report_err_handler(SapReport.WorkbookSAP,True)  
                return 2 #could not open report
            
        finally:
            if addin_is_enabled == False:
                Err_Count = Err_Count +1
                logger.critical(f"Could not activate Analysis plugin in " +  str(FilePath))
                refresh_report_err_handler(SapReport.WorkbookSAP,True)  
                return 1 #could not enable analysis

        
        print(str(filename) + ' opened')
        #Check if DS exists
        optimize_instance(SapReport.ExcelInstance, 'start')

        if check_if_DS_exists(str(current_source),SapReport.ExcelInstance) == True:
            try:
                SapReport.WorkbookSAP.Activate()
            except:
                pass
            try:
                SapReport.calculate()
            except:
                pass
            # logging on
            SapReport.logon(current_source)
            #if SapReport.is_logged == False
            # do initial data refresh
            try:
                
                if Force_Close_BOA_Messages_Window:
                    #Start Thread to monitor for BOA pop-up
                    refresh_running = True
                    thread_monitor_windows = Thread(target=search_boa_message_window)
                    thread_monitor_windows.start()
                    #Inital datasource refresh
                    run_time_start = datetime.now()
                    result = xl_Instance.Application.Run("SAPExecuteCommand", "RefreshData", current_source)
                    SapReport.is_refreshed_data = result
                    run_time_end = datetime.now()
                    #Capture Run times
                    if capture_runtimes:
                        capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Inital refresh',run_time_start,run_time_end)
                    #Stop thread after refresh
                    Monitoring_Windows = False
                    thread_monitor_windows.join()
                    refresh_running = False

                    sap_message_display(xl_Instance)

                    if BOA_message_closed== True:
                        if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                            logger.ERROR('Error during inital refresh of ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                            Err_Count = Err_Count +1 
                            refresh_report_err_handler(SapReport.WorkbookSAP,True)              
                            return 6
                        else:
                            print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                            logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')

                    if result == 1:
                        print(f'\nSuccessfully refreshed the source: {current_source}')
                    else:
                        raise ConnectionError(f"Couldn't refresh the the source: {current_source}")

                    print("Refreshed " + current_source)


                else:
                    run_time_start = datetime.now()
                    SapReport.refresh_data(current_source)
                    run_time_end = datetime.now()
                    #Capture Run times
                    if capture_runtimes:
                        capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Inital refresh',run_time_start,run_time_end)


            except:
                Err_Count = Err_Count +1
                refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                return 6 # Refresh data
            
            # start to deal with filters
            df_filters = variables_filters.query(f'Filename=="{filename}" and Datasource=="{data_sources}" and Command=="SAPSetFilter"').replace({"Value": dict_time_values})
            # start to deal with variables
            df_variables = variables_filters.query(f'Filename=="{filename}" and Datasource=="{data_sources}" and Command=="SAPSetVariable"').replace({"Value": dict_time_values})
            
            if not df_filters.empty:
                # Refresh after setting filters only if no variables must be set
                try:
                    SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
                    SapReport.set_refresh_filters(df_filters)
                    #No variables to set -> Refresh data
                    if df_variables.empty:
                        print('Start to refresh the data with the new filter restrictions')
                        if Force_Close_BOA_Messages_Window:
                            #Start Thread to monitor for BOA pop-up
                            refresh_running = True
                            thread_monitor_windows = Thread(target=search_boa_message_window)
                            thread_monitor_windows.start()

                            #Enable Refresh
                            run_time_start = datetime.now()
                            SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                            run_time_end = datetime.now()
                            #Capture Run times
                            if capture_runtimes:
                                capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)

                            #Stop thread after refresh
                            Monitoring_Windows = False
                            thread_monitor_windows.join()
                            refresh_running = False
                            sap_message_display(xl_Instance)
                            #BOA message Popup handling
                            if BOA_message_closed== True:
                                print(str(current_source)+ 'active ' + str(sap_is_ds_active(SapReport.ExcelInstance,current_source)))
                                if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                                    logger.ERROR('Error setting filters for ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                                    Err_Count = Err_Count +1   
                                    refresh_report_err_handler(SapReport.WorkbookSAP,True)             
                                    return 3
                                else:
                                    print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                                    logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')
                            else:
                                #Set variable to False to stop thread running search_boa_message_window()
                                no_message_window_found = False
                                Monitoring_Windows = False
                        else:
                            run_time_start = datetime.now()
                            SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                            run_time_end = datetime.now()
                            #Capture Run times
                            if capture_runtimes:
                                capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)

                except:
                    Err_Count = Err_Count +1
                    refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                    return 7 # Error setting filters

            if not df_variables.empty:
                #logger.info(f"Set filters  -empty")
                try:
                    SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
                    SapReport.set_refresh_variables(df_variables)
                    print('Start to refresh the data with the new variable restrictions')

                    if Force_Close_BOA_Messages_Window:
                        #Start Thread to monitor for BOA pop-up
                        
                        refresh_running = True
                        thread_monitor_windows = Thread(target=search_boa_message_window)
                        thread_monitor_windows.start()
                        #Enable Refresh
                        run_time_start = datetime.now()
                        SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                        run_time_end = datetime.now()
                        #Capture Run times
                        if capture_runtimes:
                            capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)
                        #Stop thread after refresh
                        Monitoring_Windows = False
                        thread_monitor_windows.join()
                        refresh_running = False
                        sap_message_display(xl_Instance)
                        #BOA message Popup handling
                        if BOA_message_closed== True:
                            print(str(current_source)+ 'active ' + str(sap_is_ds_active(SapReport.ExcelInstance,current_source)))
                            if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                                logger.ERROR('Error setting variables for ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                                Err_Count = Err_Count +1
                                refresh_report_err_handler(SapReport.WorkbookSAP,True)                
                                return 4
                            else:
                                print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                                logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')
                        else:
                            #Set variable to False to stop thread running search_boa_message_window()
                            no_message_window_found = False
                            Monitoring_Windows = False
                    else:
                        run_time_start = datetime.now()
                        SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                        run_time_end = datetime.now()
                        #Capture Run times
                        if capture_runtimes:
                            capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)
                except:
                    Err_Count = Err_Count +1
                    refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                    return 5 # Error setting variables
        



        crosstab_result_available, result_val, result_rows, result_columns, ct_row, ct_col  = get_crosstab_size(xl_Instance, split_fullname, current_crosstab)

        if Close_file_Check == True:
            # save and close the report
            if update_mode == "Create Configuration single" or update_mode == "Create Configuration multiple":
                pass
            else:
                if Supress_All_BOA_Messages:
                    try:
                        delete_BOA_macro_from_Excel(SapReport.WorkbookSAP)
                        BOA_macro_removed= True
                    except:
                        logger.info('Could not implement macros to supress BOA messages: '+ str(Exception) )
                    finally:
                        pass
                
                try:
                    SapReport.refesh_all()
                except:
                    logger.info('Could not run Refresh All in File: '+ str(Exception) )             


                try:
                    SapReport.Save_workbook()
                except:
                    logger.info('Could not save file: '+ str(Exception) )

            try:
                logoff(xl_Instance)
            except:
                pass
            time.sleep(5)
            SapReport.close_workbook()
            print(str(filename) + ' closed')
            

    else:
        # activate SAP AfO report
        SapReport.connect_opened_report()
        optimize_instance(SapReport.ExcelInstance, 'start')
        screenupdating_instance(SapReport.ExcelInstance, 'start')
        ensure_wb_active(SapReport.ExcelInstance, filename)
        try:
            SapReport.WorkbookSAP.Activate()
        except:
            pass        
        try:
            SapReport.ExcelInstance.Application.Calculate()
        except:
            pass

        #Check if DS exists
        if check_if_DS_exists(str(current_source),SapReport.ExcelInstance) == True:
            # logging on
            SapReport.logon(current_source)

            # do initial data refresh
            try:

                if Force_Close_BOA_Messages_Window:
                    #Start Thread to monitor for BOA pop-up
                    refresh_running = True
                    thread_monitor_windows = Thread(target=search_boa_message_window)
                    thread_monitor_windows.start()
                    #Inital datasource refresh
                    run_time_start = datetime.now()
                    result = xl_Instance.Application.Run("SAPExecuteCommand", "RefreshData", current_source)
                    SapReport.is_refreshed_data = result
                    run_time_end = datetime.now()
                    #Capture Run times
                    if capture_runtimes:
                        capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Inital refresh',run_time_start,run_time_end)
                    SapReport.is_refreshed_data = result

                    #Stop thread after refresh
                    Monitoring_Windows = False
                    thread_monitor_windows.join()
                    refresh_running = False
                    sap_message_display(xl_Instance)
                    if BOA_message_closed== True:
                        if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                            logger.ERROR('Error during inital refresh of ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                            Err_Count = Err_Count +1  
                            refresh_report_err_handler(SapReport.WorkbookSAP,True)              
                            return 6
                        else:
                            print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                            logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')

                    if result == 1:
                        print(f'\nSuccessfully refreshed the source: {current_source}')
                    else:
                        raise ConnectionError(f"Couldn't refresh the the source: {current_source}")

                    print("Refreshed " + current_source)
                else:
                    run_time_start = datetime.now()
                    SapReport.refresh_data(current_source) 
                    run_time_end = datetime.now()
                    #Capture Run times
                    if capture_runtimes:
                        capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Inital refresh',run_time_start,run_time_end)
            except:
                Err_Count = Err_Count +1
                refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                return 6 # Refresh data
            
            # start to deal with filters
            df_filters = variables_filters.query(f'Filename=="{filename}" and Datasource=="{data_sources}" and Command=="SAPSetFilter"').replace({"Value": dict_time_values})
            # start to deal with variables
            df_variables = variables_filters.query(f'Filename=="{filename}" and Datasource=="{data_sources}" and Command=="SAPSetVariable"').replace({"Value": dict_time_values})          
            
                    

            if not df_filters.empty:
                #logger.info(f"Set filters  -not empty")
                #Refresh after setting filters only if no variables must be set
                try:
                    SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
                    SapReport.set_refresh_filters(df_filters)
                    #No variables to set -> Refresh data
                    if df_variables.empty:
                        print('Start to refresh the data with the new filter restrictions')
                        if Force_Close_BOA_Messages_Window:
                            #Start Thread to monitor for BOA pop-up
                            refresh_running = True
                            
                            thread_monitor_windows = Thread(target=search_boa_message_window)
                            thread_monitor_windows.start()
                            #Enable Refresh
                            run_time_start = datetime.now()
                            SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                            run_time_end = datetime.now()
                            #Capture Run times
                            if capture_runtimes:
                                capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)
                            #Stop thread after refresh
                            Monitoring_Windows = False
                            thread_monitor_windows.join()
                            sap_message_display(xl_Instance)
                            refresh_running = False
                            #BOA message Popup handling
                            if BOA_message_closed== True:
                                if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                                    logger.ERROR('Error setting filters for ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                                    Err_Count = Err_Count +1      
                                    refresh_report_err_handler(SapReport.WorkbookSAP,True)          
                                    return 3
                                else:
                                    print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                                    logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')
                            else:
                                #Set variable to False to stop thread running search_boa_message_window()
                                no_message_window_found = False
                                Monitoring_Windows = False

                        else:
                            run_time_start = datetime.now()
                            SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                            run_time_end = datetime.now()
                            #Capture Run times
                            if capture_runtimes:
                                capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)
                except:
                    Err_Count = Err_Count +1
                    refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                    return 7 # Error setting filters

            if not df_variables.empty:
                #logger.info(f"Set filters  -empty")
                try:
                    SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "Off")
                    SapReport.set_refresh_variables(df_variables)
                    print('Start to refresh the data with the new variable restrictions')
                    if Force_Close_BOA_Messages_Window:
                        #Start Thread to monitor for BOA pop-up
                        refresh_running = True
                        
                        thread_monitor_windows = Thread(target=search_boa_message_window)
                        thread_monitor_windows.start()
                        #Enable Refresh

                        run_time_start = datetime.now()
                        SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                        run_time_end = datetime.now()
                        #Capture Run times
                        if capture_runtimes:
                            capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)
                        #Stop thread after refresh
                        Monitoring_Windows = False
                        thread_monitor_windows.join()
                        refresh_running = False
                        sap_message_display(xl_Instance)
                        #BOA message Popup handling
                        if BOA_message_closed== True:
                            if sap_is_ds_active(SapReport.ExcelInstance,current_source) == False:
                                logger.ERROR('Error setting variables for ' +str(current_source)  + ' in ' + str(filename) + ' due to BOA message pop-up, disconnecting Addin. Query has been skipped.')
                                Err_Count = Err_Count +1  
                                refresh_report_err_handler(SapReport.WorkbookSAP,True)              
                                return 4
                            else:
                                print('BOA message pop-up for ' +str(current_source)  + ' closed. Continuing as DS is still active.')
                                logger.info('BOA message pop-up for ' +str(current_source)  + ' closed in ' + str(filename) + '  Continuing as DS is still active.')
                        else:
                            #Set variable to False to stop thread running search_boa_message_window()
                            no_message_window_found = False
                            Monitoring_Windows = False
                    else:
                        run_time_start = datetime.now()
                        SapReport.state_refresh_behavior = SapReport.ExcelInstance.Application.Run("SAPSetRefreshBehaviour", "On")
                        run_time_end = datetime.now()
                        #Capture Run times
                        if capture_runtimes:
                            capture_runtime(file_target,current_sheet,current_source,current_crosstab,current_query_technical_name,'Refresh',run_time_start,run_time_end)

                except:
                    Err_Count = Err_Count +1
                    refresh_report_err_handler(SapReport.WorkbookSAP,True) 
                    return 5 # Error setting variables



        crosstab_result_available, result_val, result_rows, result_columns, ct_row, ct_col  = get_crosstab_size(xl_Instance, split_fullname, current_crosstab)

        if Close_file_Check == True:
            # save and close the report
            if update_mode == "Create Configuration single" or update_mode == "Create Configuration multiple":
                pass
            else:
                if Supress_All_BOA_Messages:
                    try:
                        delete_BOA_macro_from_Excel(SapReport.WorkbookSAP)
                        BOA_macro_removed = True
                    except:
                        logger.info('Could not delete macros to supress BOA messages: '+ str(Exception) )
                    finally:
                        pass
                
                try:
                    SapReport.refesh_all()
                except:
                    logger.info('Could not run Refresh All in File: '+ str(Exception) )       

                try:
                    SapReport.Save_workbook()
                except:
                    logger.info('Could not save file: '+ str(Exception) )
            logoff(xl_Instance)
            time.sleep(5)
            #open_wb.Close()
            SapReport.close_workbook()
            print(str(filename) + ' closed')
            optimize_instance(xl_Instance, 'stop')

    #Check size of resulting crosstabs, if only 1 row and 1 column have been returned an error is likely or there might be no data or the size limit has been exceeded
    try:
        if crosstab_result_available == False:
            if result_rows == 1 and result_columns ==1:
                Err_Count = Err_Count +1
                return 8
    except:
        pass


    return 0


def Check_Queries_to_refresh():
    global config_wb, config_ws, queries_lastrow, ExcelApp, always_refresh_error_queries

    if check_if_Excel_runs() == False:
        open_excel()
    else:
        if ExcelApp.xlApp == None or 'None' in str(ExcelApp.xlApp):
            del ExcelApp
            ExcelApp = Excel()
        xl_Instance = ExcelApp.xlApp 
        
    print(CONFIG_PATH)
    open_workbook(xl_Instance, CONFIG_PATH)
    
    optimize_instance(xl_Instance, 'start')
    screenupdating_instance(xl_Instance, 'start')
    Workdays()
    time_intelligence()
    
    config_wb = wb
    config_ws = config_wb.Worksheets("Queries")

    #Number Format Text to ensure that values are pasted correctly to variables 
    config_wb.Worksheets("Variables_Filters").Activate()
    config_wb.Worksheets("Variables_Filters").Columns("I").NumberFormat = '@'
    config_wb.Worksheets("Queries").Activate()


    #Determine last row
    xlUp = -4162
    queries_lastrow = config_ws.Cells(config_ws.Rows.Count, "A").End(xlUp).Row + 1
    queries_to_refresh_count = 0
    queries_workday_entered_count = 0
    queries_workday_to_refresh_count = 0
    #Check which queries to refresh (either Workday = current workday or =99)
    # LOOP sheet queries 
    for i in range(2,queries_lastrow):
        
        if isinstance(config_ws.Cells(i,12).Value,int):
            if config_ws.Cells(i,12).Value <= 31 and config_ws.Cells(i,12).Value > 0 and config_ws.Cells(i,12).Value != '':
                queries_workday_entered_count = queries_workday_entered_count + 1
        #If current workday equals target workday update to 99
        if config_ws.Cells(i,12).Value == current_workday:
            config_ws.Cells(i,11).Value = 99
            queries_to_refresh_count = queries_to_refresh_count +1
            queries_workday_to_refresh_count = queries_workday_to_refresh_count +1
        else:
            #If current workday does not equal target workday clear value
            if config_ws.Cells(i,12).Value != 99:
               config_ws.Cells(i,11).Value = ''
        #If workday entered is 99 always refresh
        if config_ws.Cells(i,12).Value == 99:
                config_ws.Cells(i,11).Value = 99
                queries_to_refresh_count = queries_to_refresh_count +1
        #If setting to always refresh queries with errors activated update value to 99 if last refresh cell starts with "Error"
        try:
            if always_refresh_error_queries:
                if str(config_ws.Cells(i,14).Value)[:5] == "Error":
                    config_ws.Cells(i,11).Value = 99
                    queries_to_refresh_count = queries_to_refresh_count +1
        except:
            pass


    if queries_to_refresh_count == 0:
        if queries_workday_entered_count > 0:
            print(f"No queries scheduled for refresh on current workday " + str(current_workday) + f" - if query should be refreshed please check the sheet Queries in the configuration file and change the setting in column Refresh on workday if necessary")
            logger.info(f"No queries scheduled for refresh on current workday " + str(current_workday) + f" - if query should be refreshed please check the sheet Queries in the configuration file and change the setting in column Refresh on workday if necessary")
            print('Refresh will be stopped')
            time.sleep(10)
            config_wb.Close()
            try:
                xl_Instance.Quit()
            except:
                pass
            sys.exit()

        else:
            print(f"No queries scheduled for refresh - if query should be refreshed please check the sheet Queries in the configuration file and change the setting in column Refresh on workday if necessary")
            logger.info(f"No queries scheduled for refresh - if query should be refreshed please check the sheet Queries in the configuration file and change the setting in column Refresh on workday if necessary")
            print('Refresh will be stopped')
            time.sleep(10)
            config_wb.Close()
            try:
                xl_Instance.Quit()
            except:
                pass
            sys.exit()


    #If Time Variables are entered convert to Values and Replace Time Values
    vars_ws = config_wb.Worksheets("Variables_Filters")
    vars_lastrow = vars_ws.Cells(vars_ws.Rows.Count, "A").End(xlUp).Row + 1
    #Determine last row
    xlUp = -4162
    #Check if variable is entered, if yes update with value
    for i in range(2,vars_lastrow):
        value_new = str(vars_ws.Cells(i,15).Value)
        
        if value_new != None and value_new != '' and value_new != 0 and value_new != 'None':
            #Split Value
            value_split = re.split(' +', str(value_new))
            for x in value_split:

                #Remove search operators from string
                for character in '!=><[]*;':
                    x = str(x).replace(character,'')
                    
                # LOOP time variables, replace value in config file with new calculated value 
                for var_time, val_time in zip(df_time_values['Variable'],df_time_values['Value']):
                    if str(var_time) == x and str(vars_ws.Cells(i,15).Value)!='':
                        value_new = value_new.replace(var_time,val_time)
            vars_ws.Cells(i,9).Value = value_new

    
 
    
    config_wb.Save() 
    time.sleep(5)

    print('Closing configuration file')

    config_wb.Close()
    

def Write_to_config_file_queries(fpath, DS, CT):
    for i in range(2,queries_lastrow):
        # LOOP CELLS 
        if config_ws.Cells(i,1).Value == fpath and config_ws.Cells(i,5).Value == DS and config_ws.Cells(i,9).Value == CT:
            #last refresh
            config_ws.Cells(i,14).Value = datetime.now()

def get_sys_pw(system):

    global user, password, password_found,df_pws
    password_found = False
    # if pw was not found from PW files
    if df_pws.equals(pd.DataFrame({0,0,0})):
        if user_config == False and password_config == False:
            logger.warning("Passwords could not be extracted from PW files, files are either missing or corrupted/wrong syntax - Make sure Secret.txt and PW_Encoded.txt are available and correct or use unenconed file Passwords.txt or consider saving user and password in the configuration file")
        elif user_config == True and password_config == True:
            #Use data from config file if no PW files are available
            try:
                user = df_global_configs.query('Setting=="logon-user"')['Value'].values[0]
                password = df_global_configs.query('Setting=="logon-password"')['Value'].values[0]
                password_found = True
            except:
                password_found = False


    else:
        for index, row in df_pws.iterrows():
            if  str(row["System"]) == system:
                user =str(row["User"]) 
                password =str(row["Password"]) 
                password_found = True
        

def refresh_auto_reports():
    global Open_file_Check, Close_file_Check, current_filepath, current_filename, refresh_status, current_sys, Err_Count, boa_message_severity, boa_message_text, current_sheet, current_crosstab, current_query_technical_name, Restart_Excel_after_refresh_Error
    """function to automate the refresh of reports based on the parameters set in config file"""
    _, data_sources, variables_filters = get_configurations(CONFIG_PATH)
    
    # change the dynamic days in the sources
    today = date.today().day


    data_sources['Refresh'] = data_sources['Refresh'].apply(lambda x: 'Y' if x >= today or x == 99 else 'N')

    files = data_sources.query('Refresh=="Y"').drop_duplicates(subset=['Filename'])

    #Number of sources to refresh
    data_sources = data_sources.query('Refresh=="Y"').reset_index(drop=True)

    #Sort by filename ensures files are processed subsequently
    #data_sources.sort_values(by='Datasource', ascending=True, inplace=True)
    data_sources.sort_values(by='Filename', ascending=True, inplace=True)
    #Reset index otherwise the check for opening and closing files produces wrong sequences
    data_sources.reset_index(inplace = True, drop = True)
    #Ensure index matches after sorting, otherwise the check for opening and closing files fails
    data_sources.index.names =['index']

    # number of datasources used to determine wheter to open / close files
    idx_max = data_sources.index.max()
    ds_index = 0

    # start to refresh the reports

    for index, row in data_sources.iterrows():  # execute command only to valid rows

            #Check whether to close or open files - same file remains open and is not closed, only in first iteration
            if ds_index == 0:
                Open_file_Check = True
                Close_file_Check = False
                if index+1 > idx_max:
                    #Only 1 file, therefore current file to be closed
                    Close_file_Check = True
                if index+1 <= idx_max:
                    #More than 1 file
                    if data_sources['Filename'].loc[data_sources.index[index]] != data_sources['Filename'].loc[data_sources.index[index+1]]:
                     Close_file_Check = True   
            else:
                #more datasources to refresh
                
                if index+1 <= idx_max:

                    #File does not need to be closed, as next datasource is in same file
                    if data_sources['Filename'].loc[data_sources.index[index]] == data_sources['Filename'].loc[data_sources.index[index+1]]:
                    
                        Open_file_Check = True
                        Close_file_Check = False    
                    #File is closed, as next datasource is in different file
                    else:
                        Close_file_Check = True  
                #last file to refresh, close after refresh
                else:
                    #Last file, close
                    Close_file_Check = True

                #Check if file needs to be opened, except for first iteration
                if index-1 >= 0:
                    #Same Filename, file does not need to be opened - will overwrite inital Open_file_Check
                    if data_sources['Filename'].loc[data_sources.index[index]] == data_sources['Filename'].loc[data_sources.index[index-1]]:
                        Open_file_Check = False
                    #Different Filename, previous file is NOT same file, File must be opened
                    else:
      
                        Open_file_Check = True
                #First iteration, open file
                else:
                    Open_file_Check = True

            ds_index = ds_index +1


            print(f'Starting to refresh the reports: ' + row['Filename'])
            print(f'Starting to refresh : ' + str(row['Datasource']))
            current_sys  = row['System']
            current_filename = row['Filename']
            current_filepath = row['Fullpath'] #+ row['Filename']
            current_sheet = row['Sheet']
            current_crosstab = row['Crosstab']
            current_query_technical_name = row['Query technical name']

            try:
                result_refresh = refresh_report(row['Filename'], row['Datasource'], variables_filters)
            except:
                result_refresh = 11

            # returns > 2 on refresh_report error
            if result_refresh > 2:
                print('Error during refresh, re-initalizing Excel')
                #If setting is true, close Excel and restart file
                if Restart_Excel_after_refresh_Error:
                    try:
                        Open_file_Check = True
                        time.sleep(10)
                        kill_excel_instances()
                        time.sleep(10)
                        open_excel()
                        result_refresh = refresh_report(row['Filename'], row['Datasource'], variables_filters)

                    except:
                        result_refresh = 10
                else:
                    if check_if_Excel_runs() == False:
                        open_excel()
                        try:
                            result_refresh = refresh_report(row['Filename'], row['Datasource'], variables_filters)
                        except:
                            result_refresh = 9  


            #in function refresh_report check is performed if the DS exists in the current file - stored in variable ds_exists
            # df_refresh_status is used to write status of refresh back to configuration file
            if ds_exists == True:
                try:
                    if result_refresh == 0:
                        #Refresh completed without error but error message appeared
                        if BOA_Err_Message:
                            Err_Count = Err_Count +1
                            refresh_status.append(
                            {
                                'Filename': str(row['Filename']),
                                'Datasource': str(row['Datasource']),
                                'Crosstab': str(row['Crosstab']),
                                'Error': True,
                                'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                                'DS exists': True,
                                'Error Description': str(boa_message_text)
                            }
                            )

                        else:
                            refresh_status.append(
                            {
                                'Filename': str(row['Filename']),
                                'Datasource': str(row['Datasource']),
                                'Crosstab': str(row['Crosstab']),
                                'Error': False,
                                'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                                'DS exists': True,
                                'Error Description': ''

                            }
                            )

                    elif result_refresh == 1:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Could not activate Analysis plugin'
                            
                        }
                        )

                    elif result_refresh == 2:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Could not open/access file'
                            
                        }
                        )
                    elif result_refresh == 3:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Error setting filters for due to BOA message pop-up, disconnecting Addin. Query has been skipped.'

                        }
                        )
                    elif result_refresh == 4:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Error setting variables for due to BOA message pop-up, disconnecting Addin. Query has been skipped.'

                        }
                        )
                    elif result_refresh == 5:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Error setting variables for due to BOA message pop-up, disconnecting Addin. Query has been skipped.'

                        }
                        )
                    elif result_refresh == 6:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Error during inital datasource refresh. Query has been skipped.'

                        }
                        )
                    elif result_refresh == 7:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Error setting filters for due to BOA message pop-up, disconnecting Addin. Query has been skipped.'

                        }
                        )    
                    elif result_refresh == 8:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Query did not return results. ' + str(result_val)
                             

                        }
                        )  

                    elif result_refresh == 9:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Uncaught error during second attempt of query refresh' 

                        }
                        )  
                    elif result_refresh == 10:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Uncaught error during second attempt of query refresh after restarting Excel'

                        }
                        )                

                    elif result_refresh > 10:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Uncaught error during query refresh. Query has been skipped.'

                        }
                        )
                        logger.critical(f"Datasource " + str(row['Datasource']) +  ' in ' +str(row['Filename']) + ' was not refreshed - Could not open/access file.')

                    else:
                        refresh_status.append(
                        {
                            'Filename': str(row['Filename']),
                            'Datasource': str(row['Datasource']),
                            'Crosstab': str(row['Crosstab']),
                            'Error': True,
                            'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            'DS exists': True,
                            'Error Description': 'Error: Uncaught error during query refresh. Query has been skipped.'

                        }
                        )

                except:
                    logger.info(f"Couldn't update last refresh data for " + str(row['Datasource']) +  ' in ' +str(row['Filename']))
            else:
                logger.warning(f"Datasource " + str(row['Datasource']) +  'in ' +str(row['Filename']) + ' does not exist and was skipped.')
                try:
                    refresh_status.append(
                    {
                        'Filename': str(row['Filename']),
                        'Datasource': str(row['Datasource']),
                        'Crosstab': str(row['Crosstab']),
                        'Error': True,
                        'Last refresh': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                        'DS exists': False

                    }
                    )
                except:
                    logger.info(f"Couldn't update info for missing " + str(row['Datasource']) +  ' in ' +str(row['Filename']))

 
    print(f'Finished the refreshing of: ' + row['Filename'])

def Main_Create_Configuration(): #Main_Collect():
    """function to automatically create config file"""
    global multiple_mode, SapReportCollect, current_filename, current_filepath, filenamefull_selected, filepath_selected, file_source_config

    
    current_filename = ''# row['Filename']
    current_filepath = ''# row['Fullpath'] 
    #Configuration file will not be used when configuration is to be created from files in a folder
    if update_mode == "Refresh" or update_mode == "Create Configuration multiple" and file_source_config == True:
        _, data_sources, variables_filters = get_configurations(CONFIG_PATH)
    #Define Layout of setting table - same naming has to be used as required from other functions  
    settings_list = {'Setting':['logon-client','logon-user','logon-password','logon-language','path-log_directory','path-password_file','path-secret_file','path-encoded_pw_file','path-holiday_days_file','Run Script / EXE after refresh','Send E-Mail to in case of Error'],'Value':[100,'','','en',file_path,'','','','','','']}#,'is-used':['','','','','','','','','',''] }
    #settings_list = {'Setting':['logon-client','logon-user','logon-password','logon-language','path-log_directory','path-password_file','path-secret_file','path-encoded_pw_file','path-holiday_days_file'],'Value':[100,'','','en',file_path, file_path,'','',''] }
     
    df_settings = pd.DataFrame(settings_list)
    #Create File list
    file_list = []
    


    # import files - if config file is to be created (else) set selected filename and path
    if update_mode == "Create Configuration multiple":

        multiple_mode = True
        
        #Check wheter to take files from config file or from files in a folder
        if file_source_config:
            import_files(CONFIG_PATH)
            files = df_files_configs.drop_duplicates(subset=['Fullpath'])
        #Create dataframe with all Excel files, except config File from the provided path
        else:
            print('Creating configuration file in folder '+ str(filepath_selected))
            folder = filepath_selected
            folder = pathlib.Path(folder).resolve()
            data = []

            #List all files in folder and save as dataframe 
            for file in sorted(os.listdir(folder)):
                fullpath = os.path.join(folder,file)

                Split_path(fullpath)
                #Only take into account Excel files and exclude config file
                if ".xl" in split_extension and str(file).lower().find("config") <0:
                    data.append((file, folder, fullpath))

            files = pd.DataFrame(data, columns=['Filename', 'Filepath', 'Fullpath'])

        # Collect variables from files
        for index, row in files.iterrows():  
            if file_source_config:
                filepath_selected = row['Fullpath'] 
            current_filename =  row['Filename']
            current_filepath =  row['Fullpath'] 
            try:
                result = initiate_report_to_refresh(row['Fullpath'], row['Filename'])
            except:
                logger.critical("Failed to create configuration for " +  str(row['Filename']))
                result = 3
            if result == 0:
                df_query, df_var = get_list_of_DS()
                
                if str(row['Filename']) in str(row['Filepath']):
                    filepath_selected_str = str(row['Filepath']).replace(str(row['Filename']),'')
                    file_list.append(
                        {
                            'Filename': row['Filename'],
                            'Filepath': filepath_selected_str,
                            'Fullpath': filepath_selected
                        }
                    )
                
                else:
                    
                    file_list.append(
                        {
                            'Filename': row['Filename'],
                            'Filepath': row['Filepath'],
                            'Fullpath': row['Fullpath']
                        }
                    )
                #append if index > 0 
                
                if index > 0:
                    df_query_list = pd.concat([df_query_list, df_query])
                    df_variable_list = pd.concat([df_variable_list, df_var])
                else:
                    df_query_list = df_query
                    df_variable_list = df_var

                try:
                    # Close Workbook after refresh
                    if xl_Instance.DisplayAlerts == True:
                        xl_Instance.DisplayAlerts = False

                    SapReportCollect.close_workbook()

                except:
                    logger.info("Failed to close file after refresh " +  str(row['Filename']))
                    pass
            else:
                logger.info("Failed to open file  " +  str(row['Filename']))

        try:
            SapReportCollect.WorkbookSAP = None
            SapReportCollect.ExcelInstance = None
            SapReportCollect.source = None
            del SapReportCollect
        except:
            pass
        
        
        df_query = df_query_list
        df_var = df_variable_list

    elif update_mode == "Create Configuration single":
        current_filename =  filenamefull_selected
        #Split - get path only
        #filepath_selected =  os.path.split(filepath_selected)[0]
        current_filepath =  filepath_selected
        #Try to resolve paths
        try:
            current_filepath_resolved = pathlib.Path(filepath_selected).resolve()
            filepath_selected = current_filepath_resolved
            current_filepath = current_filepath_resolved
        except:
            pass

        result = initiate_report_to_refresh(filepath_selected, filenamefull_selected)
        if result == 0:
            df_query, df_var = get_list_of_DS()
            if str(filenamefull_selected) in str(filepath_selected):
                filepath_selected_str = str(filepath_selected).replace(str(filenamefull_selected),'')
                file_list.append(
                    {
                        'Filename': str(filenamefull_selected),
                        'Filepath': filepath_selected_str,
                        'Fullpath': str(filepath_selected)
                    }
                )
            else:
                file_list.append(
                    {
                        'Filename': str(filenamefull_selected),
                        'Filepath': str(filepath_selected),
                        'Fullpath': str(filepath_selected) + str(filenamefull_selected)
                    }
                )
        else:
            logger.info("Failed to open file  " +  str(filenamefull_selected))

    df_files = pd.DataFrame.from_dict(file_list)

    coniguration_path = str(file_path) + r'/Configuration.xlsx'

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(coniguration_path, engine='xlsxwriter')
    # write the dataframes

    #Write Variables and filters
    df_var.to_excel(writer, sheet_name='Variables_Filters', startrow=1, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Variables_Filters']
    workbook = writer.book

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_var.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_var.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)
    #worksheet.set_column('A:A', 15, "dd.mm.yyyy")
    


    #Write Query List
    df_query.to_excel(writer, sheet_name='Queries', startrow=1, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Queries']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_query.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_query.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)



    #Write File List
    df_files.to_excel(writer, sheet_name='Files', startrow=1, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Files']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_files.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_files.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)




    #Write Settings List
    df_settings.to_excel(writer, sheet_name='Settings', startrow=1, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Settings']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_settings.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_settings.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)



    time_intelligence()

    df_time_values.to_excel(writer, sheet_name='Time Values', startrow=1, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Time Values']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_time_values.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_time_values.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.

    worksheet.set_column(0, max_col - 1, 12)


    #Create named ranges for time variables

    #for i in range(2, max_row+1):
    #     var_name = df_time_values.iloc[i-2,0]
    #     workbook.define_name(str(var_name), "='Time Values'!$A$"+str(i)) #""='Time Values'!$A$"+
        

    df_month_values.to_excel(writer, sheet_name='Time Values', startrow=1, startcol=5, header=False, index = False)
    
    # Get the xlsxwriter workbook and worksheet objects.
    worksheet = writer.sheets['Time Values']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df_month_values.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df_month_values.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 5, max_row, max_col - 1 + 5, {'columns': column_settings})

    
    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1 + 5, 12)



    


    # Close the Pandas Excel writer and output the Excel file.
    writer.close()

    time.sleep(1)

   #Add comments with openpyxl
    try:
        workbook = openpyxl.load_workbook(filename= coniguration_path)

        worksheet = workbook["Queries"]
        

        #Add comments Queries
        comment = Comment("Filename with extension, must be Excel File", "fs090")
        workbook["Queries"].cell(row=1,column=1).comment = comment
        workbook["Files"].cell(row=1,column=1).comment = comment
        workbook["Variables_Filters"].cell(row=1,column=1).comment = comment

        comment = Comment("Filepath, can be on SharePoint", "fs090")
        workbook["Queries"].cell(row=1,column=2).comment = comment
        workbook["Files"].cell(row=1,column=2).comment = comment
        workbook["Variables_Filters"].cell(row=1,column=2).comment = comment

        comment = Comment("Filepath and Filename with Extension, can be on SharePoint", "fs090")
        workbook["Queries"].cell(row=1,column=3).comment = comment
        workbook["Files"].cell(row=1,column=3).comment = comment
        workbook["Variables_Filters"].cell(row=1,column=3).comment = comment

        
        comment = Comment("Do not enter any value, processed automatically", "fs090")
        worksheet.cell(row=1,column=11).comment = comment
        comment = Comment("Enter 99 if the query should always be refreshed, or enter the working day number e.g. 2 if the query should be refreshed on the second working day in the month", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=1,column=12).comment = comment
        comment = Comment("Enter only a valid local path if you want to create a CSV out of the query result, leave empty if not required. Must be local path, SharePoint not possible.", "fs090")
        worksheet.cell(row=1,column=15).comment = comment
        comment.width = 120
        comment.height = 80
        
        comment = Comment("Curently not implemented, without function", "fs090")
        worksheet.cell(row=1,column=13).comment = comment       
        

        #Add comments Settings
        worksheet = workbook["Settings"]
        comment = Comment("Must be Local/Netdrive file, not on SharePoint.  .txt File with Syntax System;User;Password", "fs090")
        worksheet.cell(row=7,column=1).comment = comment

        comment = Comment("Must be Local/Netdrive file, not on SharePoint. Secret File to Decode encoded Passwords, can be created on the interface on Settings tab with Create Secret File", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=8,column=1).comment = comment

        comment = Comment("Must be Local/Netdrive file, not on SharePoint. Encoded Passwords, can be created on the interface on Settings tab with Encode Password File - make sure to use correct Secret File", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=9,column=1).comment = comment

        comment = Comment("Must be Local/Netdrive file, not on SharePoint. Must be .csv file with one Column and Header Holidays, use date format 4/18/2022 m/d/yyyy", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=10,column=1).comment = comment

        comment = Comment("Must be an EXE or VBS file, other types might not be supported. Enther the full path of the file.", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=11,column=1).comment = comment


        comment = Comment("Enter E-Mail, for multiple recipients use correct separator between recipients to prevent errors.", "fs090")
        comment.width = 120
        comment.height = 80
        worksheet.cell(row=12,column=1).comment = comment

        #Add comments Variables_Filters

        
        worksheet = workbook["Variables_Filters"]
        comment = Comment("Value in this column will be used to refresh Query, make sure to use correct format to avoid errors", "fs090")
        worksheet.cell(row=1,column=9).comment = comment

        comment = Comment("Variables on Time Values Sheet can be used, enter Variable name from Column A, it will be replaced with the current value of the variable. You can also use more variables in one cell e.g. LM_MM_YYYY - LM_MM_YYYY", "fs090")
        worksheet.cell(row=1,column=15).comment = comment



        # save the workbook file
        workbook.save(coniguration_path)

    except:
        pass



def update_time_values(configuration_path):
    ## TODO - update time values in Excel configuration file
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(configuration_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    time_intelligence()
    print(df_time_values.head())
    
    df_time_values.to_excel(writer, sheet_name='Time Values', startrow=2, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    worksheet = writer.sheets['Time Values']

    # Get the dimensions of the dataframe.
    #(max_row, max_col) = df_time_values.shape

    # Create a list of column headers, to use in add_table().
    #column_settings = [{'header': column} for column in df_time_values.columns]

    # Add the Excel table structure. Pandas will add the data.
    
    #worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.

    #worksheet.set_column(0, max_col - 1, 12)


    ###Create named ranges for time variables

    #for i in range(2, max_row+1):
    #     var_name = df_time_values.iloc[i-2,0]
    #     workbook.define_name(str(var_name), "='Time Values'!$A$"+str(i)) #""='Time Values'!$A$"+
    writer.close()

    writer = pd.ExcelWriter(configuration_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    df_month_values.to_excel(writer, sheet_name='Time Values', startrow=2, startcol=5, header=False, index = False)
    # Get the xlsxwriter workbook and worksheet objects.
    
    #worksheet = writer.sheets['Time Values']

    # Get the dimensions of the dataframe.
    #(max_row, max_col) = df_month_values.shape

    # Create a list of column headers, to use in add_table().
    #column_settings = [{'header': column} for column in df_month_values.columns]

    # Add the Excel table structure. Pandas will add the data.
    #worksheet.add_table(0, 5, max_row, max_col - 1 + 5, {'columns': column_settings})

    
    # Make the columns wider for clarity.
    #worksheet.set_column(0, max_col - 1 + 5, 12)

    # Close the Pandas Excel writer and output the Excel file.
    writer.close()
    
def write_range_to_csv(workbook,range_name,sheet,save_path):
        
        workbook = openpyxl.load_workbook(workbook)
        
        sheet = workbook[sheet]

        try:
            # writer object is created
            col = csv.writer(open(save_path,
                                'w', 
                                newline="",encoding='utf-8'))
            
            #Write only named range of crosstab to CSV
            #Get Range from Named range
            query_range = workbook.defined_names.get(range_name)
            address = list(query_range.destinations)
            for sheetname, cellAddress in address:
                #removing the $ from the address
                cellAddress = cellAddress.replace('$','')
                #Split address in starting and ending cell
                start, end = cellAddress.split(":")
                #Get Row and Column numbers for  starting and ending cell
                xy = coordinate_from_string(start) # returns ('A',4)
                start_col = column_index_from_string(xy[0]) # returns 1
                start_row = xy[1]
                xy = coordinate_from_string(end) # returns ('A',4)
                end_col = column_index_from_string(xy[0]) # returns 1
                end_row = xy[1]      

            #Loop rows and columns according to the named range address and write as string values to CSV
            for row_cells in sheet.iter_rows(min_row=start_row, max_row=end_row):
                col.writerow(str(cell.value) for cell in row_cells[:end_col])

        except Exception as e:
            # writer object is created
            col = csv.writer(open(save_path,
                                'w', 
                                newline="",encoding='utf-8'))
            # writing all data from sheet in csv file
            for r in sheet.rows:
                # row by row write 
                # operation is perform
                col.writerow([str(cell.value) for cell in r])


def check_refresh_status():
    global df_err
    df_err = 0
    try:
        df_refresh_status = pd.DataFrame.from_dict(refresh_status)       
        for index, row in df_refresh_status.iterrows():
            try:
                if row["Error"] == True or str(row["Error"]) == 'True':
                    df_err = df_err +1
            except:
                pass
        return df_err, True
    except:
        return df_err, False

def delete_tmp_files():
    df_temp_paths = pd.DataFrame.from_dict(temp_paths) 
    for index, row in df_temp_paths.iterrows():
        try:
            if os.path.exists(str(row["TempFilepath"])):
                os.remove(str(row["TempFilepath"]))
        except:
            pass


def Main_refresh():
    global refresh_status, ExcelApp, temp_paths, df_refresh_status


    #refresh_status = dict()
    refresh_status = []
    temp_paths = []
    Check_Queries_to_refresh()
    refresh_auto_reports()



    try:
    
        df_refresh_status = pd.DataFrame.from_dict(refresh_status)         
        print(df_refresh_status.head())

        if check_if_Excel_runs() == False:
            open_excel()
        else:
            if ExcelApp.xlApp == None or 'None' in str(ExcelApp.xlApp):
                del ExcelApp
                ExcelApp = Excel()
            xl_Instance = ExcelApp.xlApp 
            
        wb, xl_Instance =  open_workbook(xl_Instance, CONFIG_PATH)
        config_wb = wb
        config_ws = config_wb.Worksheets("Queries")
        #Iterate over all dates, add counter for workdays per month
        for i in range(2,queries_lastrow):
            # LOOP CELLS 
                for index, row in df_refresh_status.iterrows():
                    if config_ws.Cells(i,1).Value == str(row["Filename"]) and config_ws.Cells(i,5).Value == str(row["Datasource"]) and config_ws.Cells(i,9).Value == str(row["Crosstab"]):
                        #last refresh
                        config_ws.Cells(i,14).Value = str(row["Last refresh"])
                        if row["Error"] == True:
                            config_ws.Cells(i,14).Value = 'Error during refresh at ' + str(row["Last refresh"])
                        if row["DS exists"] == False:
                            config_ws.Cells(i,14).Value = 'Error during refresh (DS does not exist) at ' + str(row["Last refresh"])
        
        # Change CSV paths for Web / SharePoint files to local temp path
        try: 
            #Create dataframe with temporary file paths, paths are created during saving the file if the path is a Web or SharePoint path - saving a CSV there would fail
            df_temp_paths = pd.DataFrame.from_dict(temp_paths) 
            csv_paths_changed = False
            #Iterate over all filenames
            for i in range(2,queries_lastrow):
                # LOOP CELLS  - change CSV save path
                for index, row in df_temp_paths.iterrows():
                    Split_path(str(row["TempFilepath"]))
                    if config_ws.Cells(i,1).Value == str(row["Filename"]) and config_ws.Cells(i,15).Value !='' and config_ws.Cells(i,15).Value !=None and config_ws.Cells(i,15).Value !='nan':
                        #Change CSV path
                        config_ws.Cells(i,15).Value = str(row["TempPath"]) + str(split_nameonly) + "_" + config_ws.Cells(i,7).Value + ".csv"
                        print(str(row["TempPath"]) + str(split_nameonly) + "_" + config_ws.Cells(i,7).Value + ".csv")
                        csv_paths_changed = True
            if csv_paths_changed == True:
                logger.info(f"CSV save temp paths have been replaced, please check in configuration file on query sheet")
        except Exception as e:
            logger.info(f"Error updating CSV save temp path. " + str(e))

        config_wb.Save()
        config_wb.Close()
        time.sleep(2)
        if xl_Instance.Workbooks.Count == 0:
            xl_Instance.Quit()

    except Exception as e:
        logger.warning(f"Couldn't update last refresh time in configuration file. Error: "+ str(e))

    #Save datasources as CSV
    try:
        for index, row in data_sources.iterrows():
            #Only save csv if path is entered
            if str(row['Save as CSV Fullpath']) != '' and str(row['Save as CSV Fullpath']) != 'nan' and row['Save as CSV Fullpath'] != None and save_results_csv == True:
                query_range = 'SAP' + str(row['Crosstab'])

                if csv_paths_changed == True:
                    
                    for index_2, row_2 in df_temp_paths.iterrows():             
                            #If filename found in temp_path read from temp path and save in temp
                            if str(row_2["Filename"]) == str(row["Filename"]):                          
                                Split_path(str(row["Fullpath"]))
                                tm_pth = str(row_2['TempPath']) + str(split_nameonly) + "_" + str(row["Query technical name"]) + ".csv"
                                # select the active sheet
                                write_range_to_csv(str(row["Fullpath"]),query_range,str(row["Sheet"]),tm_pth)
                                
                            else:
                                tm_pth = str(row['Save as CSV Fullpath'])
                                write_range_to_csv(str(row["Fullpath"]),query_range,str(row["Sheet"]),tm_pth)
                else:
                        
                        if save_results_csv == True:
                            tm_pth = str(row['Save as CSV Fullpath'])

                            write_range_to_csv(str(row["Fullpath"]),query_range,str(row["Sheet"]),tm_pth)

    except Exception as ex:
        logger.info(f"Couldn't save " + str(row['Datasource']) + " as CSV. Error: " + str(ex))



def open_outlook():
    """Start a instance of Outlook application"""
    global ol_Instance
    outlook_open = False
    #Check if Outlook is running
    try:
        ol_Instance = win32.GetActiveObject("Outlook.Application")
        return ol_Instance

    except:
        outlook_open = False 

    if outlook_open == False:
        try:
            ol_Instance = win32.gencache.EnsureDispatch('Outlook.Application')
        except:
            #deletes gen_py folder in case of Attribute error
            ol_Instance = dispatch('Outlook.Application')
        return ol_Instance


def Send_Mail(Subject, Body, MailTo, Attachement_Path):

    ol = open_outlook()

    olmailitem=0x0 #size of the new email
    newmail=ol.CreateItem(olmailitem)
    newmail.Subject= Subject
    newmail.To=MailTo
    #newmail.CC=MailCC
    #newmail.Body= Body
    newmail.HTMLBody = Body
    if not Attachement_Path == None:
        try:
            for i in range(len(Attachement_Path)):
                newmail.Attachments.Add(Attachement_Path[i])
        except:
            pass
    # newmail.Attachments.Add(attach)
    # To display the mail before sending it
    #newmail.Display() 
    newmail.Send()

def Run_Macro_in_Excel(pth):

    #Split fullpath into Filename, Module name and Macro Name
    Macro_Fullpath = pth.split("!")[0]
    Macro_Filename = os.path.basename(Macro_Fullpath)
    Macro_Module = pth.split("!")[1]
    Module = Macro_Module.split(".")[0]
    Macro = Macro_Module.split(".")[1]

    xl=win32.Dispatch("Excel.Application")
    xl.Workbooks.Open(Macro_Fullpath)
    xl.Application.Visible = True
    cmd = Macro_Filename+"!"+Module+"."+Macro
    xl.Application.Run(cmd)
    ##xl.Application.Save() # if you want to save then uncomment this line and change delete the ", ReadOnly=1" part from the open function.
    #xl.Application.Quit() # Comment this out if your excel script closes
    del xl


def Run_Script_Exe(pth):
    global Err_Count
    try:      
        exe_filename, exe_file_extension = os.path.splitext(pth)
        if len(exe_file_extension) > 1:
            print('Starting process ' + str(pth))

            if exe_file_extension == ".vbs":
                subprocess_call = "cscript " + pth
            elif exe_file_extension == ".exe":
                subprocess_call = pth
            else:
                subprocess_call = pth
            #subprocess.call(subprocess_call)
            process = subprocess.Popen(subprocess_call)
            process.wait()

    except:
        print('Couldnt start process ' + str(pth))
        Err_Count = Err_Count +1
        logger.warning(f"Couldn't run process after refresh " + str(pth) +  " " + str(Exception))



def Main(): 



    global ExcelApp, capture_runtimes

    ExcelApp = Excel()

    if start_mode == "Scheduled" or update_mode == "Refresh":

        #try:
        #    update_time_values(CONFIG_PATH)
        #except:
        #    pass
        #quit()
        Main_refresh()
        #Run Script / EXE after refresh if path is filled in config file
        if executable_config:
            Run_Script_Exe(executable_path)
    elif update_mode == "Create Configuration single" or update_mode == "Create Configuration multiple":
        Main_Create_Configuration()
        
        try:
            #global refresh_running, Monitoring_Windows
            #Start Thread to monitor for BOA pop-up
            #refresh_running = True
            #thread_monitor_windows = Thread(target=search_boa_message_window)
            #thread_monitor_windows.start()
            logoff(xl_Instance)
        except:
            pass
        finally:
            #Stop thread after logoff
            #Monitoring_Windows = False
            #thread_monitor_windows.join()
            #refresh_running = False
            pass

    
if __name__ == '__main__':


    if clear_excel_sessions == True:
        kill_excel_instances()

    try:
        Main()


        
    except CredentialsException as cred_e:
        logger.critical('Logon credentials were not found, refresh stopped. Make sure password files are availabe and/or settings in Configuration File are maintained correctly.')
        #Close Excel
        try:
            SapReport.close_workbook()
        except:
            pass
        try:
            xl_Instance.Quit()
        except:
            pass
    except Exception as e:
        # send error to the logger
        try:
            logger.critical(f"Couldn't refresh the data. ({e.args[0]} | {e.args[1]}) + " + str(e))
        except:
            logger.critical(f"Couldn't refresh the data " + str(e))
        finally:
            #todo check refresh status and restart last active file df_refresh_status = pd.DataFrame.from_dict(refresh_status)       
            try:
                sys.exit()
            except SystemExit as ex:
                print ('caught SystemExit:', ex)
            

    except SystemExit as ex:
        print ('caught SystemExit:', ex)
        

    finally:
        end_time = datetime.fromtimestamp(time.time())

        #Delete tmp files if applicable and check refresh status only relavant when refresh is performed
        if start_mode == "Scheduled" or update_mode == "Refresh":
            try:
                delete_tmp_files()
            except:
                pass
            #Check refresh status
            try:
                df_err, df_err_check = check_refresh_status()
            except:
                print('Error reading refresh status.')
                logger.info('Error reading refresh status. ' + str(Exception))

            try:
                #Send Mail with error if error occured during refresh (in df_refresh_status)
                dirlist = []
                if Send_Mail_Error_config and df_err > 0:
                    #attachements saved appended in dirlist if they exist
                    
                    #Append logfile
                    if os.path.exists(location):
                        dirlist.append(str(location))
                    try:
                        #Append BOA messages
                        file_pth = CheckFiles(path_tmp,'boa_messages','.txt')
                        if os.path.exists(file_pth):
                            file_time = datetime.fromtimestamp(os.path.getmtime(file_pth))
                            #Only append the file if it has been updated after starting the update
                            if file_time>start_time:
                                dirlist.append(str(file_pth))
                    except Exception:
                        pass
            except:
                logger.info('Error preparing attachements for E-Mail. ' + str(Exception))


            #Send Mail if error occured
            if df_err_check and df_err > 0 and Send_Mail_Error_config:
                try:
                    #If DF Refresh status exists include in Mail
                    refresh_status_html = df_refresh_status.to_html()
                    if dirlist != []:
                        body =  '<p>An Error occured during the refresh ending at ' + str(end_time.strftime("%d/%m/%Y %H:%M:%S")) + '. Please check the attached files for details. Review the table below to identify which datasource caused the error. </p><br>' + refresh_status_html
                    else:
                        body =  '<p>An Error occured during the refresh ending at ' + str(end_time.strftime("%d/%m/%Y %H:%M:%S")) +'. Review the table below to identify which datasource caused the error.</p><br>' + refresh_status_html
                    
                    try:
                        if dirlist != []:
                            Send_Mail("Error during refresh",body,Send_Mail_Error, dirlist)
                        else:
                            Send_Mail("Error during refresh",body,Send_Mail_Error, None)
                    except:
                        logger.info('Error sending E-Mail. ' + str(Exception))

                except:
                    if dirlist != []:
                        body =  '<p>An Error occured during the refresh ending at ' + str(end_time.strftime("%d/%m/%Y %H:%M:%S")) + '. Please check the attached files for details</p><br>' 
                    else:
                        body =  '<p>An Error occured during the refresh ending at ' + str(end_time.strftime("%d/%m/%Y %H:%M:%S")) +'</p><br>' 
                    try:
                        if dirlist != []:
                            Send_Mail("Error during refresh",body,Send_Mail_Error, dirlist)
                        else:
                            Send_Mail("Error during refresh",body,Send_Mail_Error, None)
                    except:
                        logger.info('Error sending E-Mail. ' + str(Exception))
    
        if Err_Count == 0:
            #Log File will be deleted on no error
            logger.info("The Workbook refresh was done successfully!")
            logging.shutdown()
            try:
                if os.path.exists(location):
                    print("The Workbook refresh was done successfully! Logfile will be deleted.")
                    os.remove(location)
            except:
                pass
        else:
            logging.shutdown()
        
        #Save Query runtimes
        try:
            if update_mode == 'Refresh':
                if capture_runtimes:
                    df_runtimes = pd.DataFrame.from_dict(run_times)   
                    print(df_runtimes.head(100))
                    path_runtimes = os.path.join(file_path,'Runtimes.csv')
                    df_runtimes.to_csv(path_runtimes, mode = 'w', index = False)
        except:
            pass

        if Close_Excel_After_Completion:
            time.sleep(10)
            kill_excel_instances()
        
        #Optional clean ups
        try:
            #Clean up old temp files from pyinstaller (_MEIXXXX folders in temp)
            cleanup_mei()
        except:
            pass

        


